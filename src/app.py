#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Mighty Skill-Bridge: API & Web Server Component
Author: Antigravity 2.0 (AI Agent)

This FastAPI application provides:
1. Static hosting for index.html.
2. Dynamic Multimodal AI parsing of resumes & jobs (Gemini 1.5/2.5 Flash).
3. 4-Dimension AI Fit Evaluation.
4. Auto-syncing of matching results to Google Sheets (Mighty Match Logs) with visual decoration.
5. Quota-safe deterministic fallbacks when GEMINI_API_KEY is not configured, ensuring robust demo delivery.
"""

import os
import sys
import asyncio
import contextvars
import datetime
import json
import io
import csv
import re
import unicodedata
import hashlib
import uuid
import requests
import subprocess
import time
import base64
import secrets
import tempfile
import threading
from pathlib import Path
from urllib.parse import urlparse
from contextlib import asynccontextmanager
from dataclasses import dataclass, field, asdict
from typing import Any, Dict, List, Optional, Set, Tuple

# Session-only aptitude/motivation self-check logic (T876). Imported both as a
# top-level module (tests put src/ on sys.path) and as src.aptitude_demo
# (production imports the app via `from src.app import app`).
try:
    import aptitude_demo
except ImportError:  # pragma: no cover - production path
    from src import aptitude_demo


def env_int(name: str, default: int, minimum: int, maximum: int) -> int:
    try:
        value = int(os.environ.get(name, str(default)))
    except (TypeError, ValueError):
        return default
    return max(minimum, min(value, maximum))


def env_float(name: str, default: float, minimum: float, maximum: float) -> float:
    try:
        value = float(os.environ.get(name, str(default)))
    except (TypeError, ValueError):
        return default
    return max(minimum, min(value, maximum))


def env_flag(name: str, default: bool = False) -> bool:
    raw_value = os.environ.get(name)
    if raw_value is None:
        return default
    return raw_value.strip().lower() in {"1", "true", "yes", "on"}


# Set console encoding to UTF-8 to prevent encoding errors on Windows terminal
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Request, status, Query
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBasic, HTTPBasicCredentials, HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, Field
try:
    from .rate_limit import SlidingWindowRateLimiter, client_identifier
except ImportError:
    from rate_limit import SlidingWindowRateLimiter, client_identifier

try:
    from .stripe_customer_portal import (
        StripePortalError,
        build_customer_portal_payload,
        create_customer_portal_session,
        sanitized_payload as sanitize_stripe_portal_payload,
    )
except ImportError:
    from stripe_customer_portal import (
        StripePortalError,
        build_customer_portal_payload,
        create_customer_portal_session,
        sanitized_payload as sanitize_stripe_portal_payload,
    )

try:
    from .sales_email_match import (
        SearchCriteria,
        build_match_report_from_file,
        criteria_from_values,
    )
except ImportError:
    try:
        from sales_email_match import (
            SearchCriteria,
            build_match_report_from_file,
            criteria_from_values,
        )
    except ImportError:
        SearchCriteria = None
        build_match_report_from_file = None
        criteria_from_values = None

try:
    from .sales_email_review import (
        VALID_FEEDBACK_STATUSES as SALES_EMAIL_REVIEW_STATUSES,
        build_review_entry as build_sales_email_review_entry,
        build_review_report as build_sales_email_review_report,
        find_match as find_sales_email_match_for_review,
        load_review_report as load_sales_email_review_report,
        match_key as sales_email_match_key,
        project_by_key as sales_email_project_by_key,
        talent_by_key as sales_email_talent_by_key,
        upsert_review_entry as upsert_sales_email_review_entry,
        write_json_report as write_sales_email_review_json,
        write_markdown_report as write_sales_email_review_markdown,
    )
except ImportError:
    try:
        from sales_email_review import (
            VALID_FEEDBACK_STATUSES as SALES_EMAIL_REVIEW_STATUSES,
            build_review_entry as build_sales_email_review_entry,
            build_review_report as build_sales_email_review_report,
            find_match as find_sales_email_match_for_review,
            load_review_report as load_sales_email_review_report,
            match_key as sales_email_match_key,
            project_by_key as sales_email_project_by_key,
            talent_by_key as sales_email_talent_by_key,
            upsert_review_entry as upsert_sales_email_review_entry,
            write_json_report as write_sales_email_review_json,
            write_markdown_report as write_sales_email_review_markdown,
        )
    except ImportError:
        SALES_EMAIL_REVIEW_STATUSES = set()
        build_sales_email_review_entry = None
        build_sales_email_review_report = None
        find_sales_email_match_for_review = None
        load_sales_email_review_report = None
        sales_email_match_key = None
        sales_email_project_by_key = None
        sales_email_talent_by_key = None
        upsert_sales_email_review_entry = None
        write_sales_email_review_json = None
        write_sales_email_review_markdown = None

# Try loading optional libraries for Sheets & Gemini
try:
    import gspread
    from google.oauth2.service_account import Credentials as ServiceCredentials
    from google.oauth2.credentials import Credentials as UserCredentials
    import google.auth.transport.requests
    from google_workspace_account import (
        GoogleWorkspaceAccountError,
        assert_expected_google_account,
        credentials_from_gspread_client,
    )
    SHEETS_LIB_AVAILABLE = True
except ImportError:
    SHEETS_LIB_AVAILABLE = False

try:
    from google import genai
    from google.genai import types as genai_types
    GEMINI_LIB_AVAILABLE = True
except ImportError:
    genai = None
    genai_types = None
    GEMINI_LIB_AVAILABLE = False

# Firebase Admin SDK Initialization
try:
    import firebase_admin
    from firebase_admin import credentials, auth
    FIREBASE_ADMIN_AVAILABLE = True
except ImportError:
    FIREBASE_ADMIN_AVAILABLE = False

if FIREBASE_ADMIN_AVAILABLE:
    try:
        firebase_admin.get_app()
    except ValueError:
        try:
            firebase_admin.initialize_app()
            print("[+] Firebase Admin SDK initialized with default credentials.")
        except Exception as e:
            try:
                firebase_admin.initialize_app(credentials.BlankCredentials())
                print("[+] Firebase Admin SDK initialized with Blank Credentials (local emulator mode).")
            except Exception as ex:
                print(f"[-] Failed to initialize Firebase Admin: {ex}")
                FIREBASE_ADMIN_AVAILABLE = False

# Supabase Client SDK Integration (T731_4)
try:
    from supabase_client import (
        is_supabase_configured,
        get_engineers as sdk_get_engineers,
        get_engineer as sdk_get_engineer,
        insert_engineer as sdk_insert_engineer,
        get_jobs as sdk_get_jobs,
        get_job as sdk_get_job,
        insert_job as sdk_insert_job,
        get_match_results as sdk_get_match_results,
        insert_match_result as sdk_insert_match_result,
        get_supabase_client
    )
    SUPABASE_SDK_ACTIVE = is_supabase_configured()
except ImportError:
    SUPABASE_SDK_ACTIVE = False

security_bearer = HTTPBearer(auto_error=False)
USER_DATA_EXPORT_ALLOW_MOCK = env_flag("USER_DATA_EXPORT_ALLOW_MOCK", False)

def get_current_user(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security_bearer)) -> dict:
    """Dependency to get the authenticated user from Firebase Auth ID Token.
    
    If MOCK_AUTH environment variable is 1, or Firebase Admin SDK is not available,
    it returns a default mock user.
    """
    mock_user = {"uid": "user_9999", "email": "k-umezawa@ml-mightylink.com"}
    
    if env_flag("MOCK_AUTH", default=True):
        return mock_user

    if not FIREBASE_ADMIN_AVAILABLE:
        return mock_user

    if not credentials:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Missing Authorization Header",
            headers={"WWW-Authenticate": "Bearer"},
        )
        
    token = credentials.credentials
    try:
        decoded_token = auth.verify_id_token(token)
        return {
            "uid": decoded_token.get("uid"),
            "email": decoded_token.get("email"),
            "name": decoded_token.get("name", "User"),
            "decoded_token": decoded_token
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"Invalid Authorization Token: {str(e)}",
            headers={"WWW-Authenticate": "Bearer"},
        )


def get_current_user_for_data_export(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security_bearer),
) -> dict:
    """Strict Firebase identity for user data exports.

    The public demo can run with MOCK_AUTH enabled, but personal data export must
    not use that mock identity unless tests/local operators opt in explicitly.
    """
    if USER_DATA_EXPORT_ALLOW_MOCK:
        return get_current_user(credentials)

    if not FIREBASE_ADMIN_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Firebase Admin SDK is required for user data export.",
        )
    if not credentials:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Firebase ID token is required for user data export.",
            headers={"WWW-Authenticate": "Bearer"},
        )

    try:
        decoded_token = auth.verify_id_token(credentials.credentials)
        return {
            "uid": decoded_token.get("uid"),
            "email": decoded_token.get("email"),
            "name": decoded_token.get("name", "User"),
            "decoded_token": decoded_token,
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"Invalid Authorization Token: {str(e)}",
            headers={"WWW-Authenticate": "Bearer"},
        )

# Initialize FastAPI App
@asynccontextmanager
async def lifespan(app: FastAPI):
    # R114/T866: the schema must be guaranteed before the first request is
    # served. A daemon thread never completes on Cloud Run's request-scoped
    # CPU, so init_db now runs to completion before yield (cold start pays
    # the cost). to_thread keeps the event loop responsive while it runs.
    try:
        await asyncio.to_thread(init_db)
    except Exception as exc:
        # A transient DB outage must not crash-loop the container; storage
        # requests will surface classified 500s (record_storage_failure).
        print(f"[-] init_db failed during startup: {type(exc).__name__}: {exc}")
    yield


app = FastAPI(title="Mighty Skill-Bridge API Server", lifespan=lifespan)

# Dynamic Path Resolution (ensures app.py works robustly inside src/)
ROOT_DIR = os.path.dirname(os.path.abspath(__file__)) # This is src/
PROJECT_ROOT = os.path.dirname(ROOT_DIR)             # Project root
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
EXPORTS_DIR = os.path.join(PROJECT_ROOT, "exports")
AUDIT_DIR = os.path.join(DATA_DIR, "audit")
AUDIT_LOG_FILE = os.path.join(AUDIT_DIR, "ai_audit.jsonl")
EXTERNAL_API_USAGE_LOG_FILE = os.path.join(DATA_DIR, "external_api_usage.jsonl")
KNOWLEDGE_FLOW_DIR = os.path.join(EXPORTS_DIR, "knowledge_flow")
KNOWLEDGE_FLOW_MANIFEST = os.path.join(KNOWLEDGE_FLOW_DIR, "manifest.json")
KNOWLEDGE_FLOW_SCRIPT = os.path.join(PROJECT_ROOT, "scripts", "generate_knowledge_flow_demo.py")
SALES_EMAIL_MATCH_REPORT_FILE = os.environ.get(
    "SALES_EMAIL_MATCH_REPORT_FILE",
    os.path.join(EXPORTS_DIR, "sales_email_extraction_review.json"),
)
SALES_EMAIL_REVIEW_LOG_FILE = os.environ.get(
    "SALES_EMAIL_REVIEW_LOG_FILE",
    os.path.join(EXPORTS_DIR, "sales_email_review_log.json"),
)
SALES_EMAIL_REVIEW_MARKDOWN_FILE = os.environ.get(
    "SALES_EMAIL_REVIEW_MARKDOWN_FILE",
    os.path.join(EXPORTS_DIR, "sales_email_review_log.md"),
)
FAVICON_FILE = os.path.join(PROJECT_ROOT, "favicon.ico")
CHROME_DEVTOOLS_WORKSPACE_PATH = "/.well-known/appspecific/com.chrome.devtools.json"
SEEDANCE_DEMO_DIR = os.path.join(EXPORTS_DIR, "seedance_demo")
SEEDANCE_DEMO_VIDEO = os.path.join(SEEDANCE_DEMO_DIR, "mighty_skill_bridge_seedance_demo.mp4")
SEEDANCE_DEMO_MANIFEST = os.path.join(SEEDANCE_DEMO_DIR, "manifest.json")
SEEDANCE_MODEL = os.environ.get("SEEDANCE_MODEL", "seedance-1-0-pro")
SEEDANCE_API_URL = os.environ.get("SEEDANCE_API_URL", "").strip()
SEEDANCE_RESULT_API_URL_TEMPLATE = os.environ.get("SEEDANCE_RESULT_API_URL_TEMPLATE", "").strip()
SEEDANCE_PAYLOAD_STYLE = os.environ.get("SEEDANCE_PAYLOAD_STYLE", "content_task").strip().lower()
SEEDANCE_POLL_TIMEOUT_SECONDS = env_int("SEEDANCE_POLL_TIMEOUT_SECONDS", 30, 0, 600)
SEEDANCE_POLL_INTERVAL_SECONDS = env_int("SEEDANCE_POLL_INTERVAL_SECONDS", 5, 1, 60)
SEEDANCE_API_ENABLED = env_flag("SEEDANCE_API_ENABLED", False)
SEEDANCE_DAILY_GENERATION_LIMIT = env_int("SEEDANCE_DAILY_GENERATION_LIMIT", 1, 0, 1000)
SEEDANCE_DAILY_REPORTED_TOKEN_LIMIT = env_int("SEEDANCE_DAILY_REPORTED_TOKEN_LIMIT", 0, 0, 1_000_000_000)
SEEDANCE_API_KEY = (
    os.environ.get("SEEDANCE_API_KEY")
    or os.environ.get("ARK_API_KEY")
    or os.environ.get("BYTEPLUS_API_KEY")
)

CREDENTIALS_FILE = os.path.join(PROJECT_ROOT, "credentials.json")
CLIENT_SECRET_FILE = os.path.join(PROJECT_ROOT, "client_secret.json")
AUTHORIZED_USER_FILE = os.path.join(PROJECT_ROOT, "authorized_user.json")
SPREADSHEET_ID = "1L99HCBHr4IsVUWqnUuG6OgoUmxEQUdfaYQim1n6etB8"
USER_EMAIL = "k-umezawa@ml-mightylink.com"
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-3.5-flash")
GEMINI_DAILY_CALL_LIMIT = env_int("GEMINI_DAILY_CALL_LIMIT", 20, 0, 10000)
GEMINI_DAILY_REPORTED_TOKEN_LIMIT = env_int("GEMINI_DAILY_REPORTED_TOKEN_LIMIT", 100000, 0, 1_000_000_000)

# Basic authentication configuration
IS_MANAGED_RUNTIME = bool(os.environ.get("K_SERVICE") or os.environ.get("FUNCTION_TARGET"))
BASIC_AUTH_USERNAME = os.environ.get("BASIC_AUTH_USERNAME")
BASIC_AUTH_PASSWORD = os.environ.get("BASIC_AUTH_PASSWORD")

if not IS_MANAGED_RUNTIME:
    BASIC_AUTH_USERNAME = BASIC_AUTH_USERNAME or "admin"
    BASIC_AUTH_PASSWORD = BASIC_AUTH_PASSWORD or "mighty-link-pass"

security = HTTPBasic()
security_optional = HTTPBasic(auto_error=False)
api_rate_limiter = SlidingWindowRateLimiter()

RATE_LIMIT_ENABLED = env_flag("RATE_LIMIT_ENABLED", True)
RATE_LIMIT_WINDOW_SECONDS = env_int("RATE_LIMIT_WINDOW_SECONDS", 60, 1, 3600)
RATE_LIMIT_MAX_REQUESTS = env_int("RATE_LIMIT_MAX_REQUESTS", 120, 1, 100000)
RATE_LIMIT_AUTH_MAX_REQUESTS = env_int("RATE_LIMIT_AUTH_MAX_REQUESTS", 30, 1, 100000)
RATE_LIMIT_EXPENSIVE_MAX_REQUESTS = env_int("RATE_LIMIT_EXPENSIVE_MAX_REQUESTS", 20, 1, 100000)
RATE_LIMIT_GENERATION_MAX_REQUESTS = env_int("RATE_LIMIT_GENERATION_MAX_REQUESTS", 6, 1, 100000)
RATE_LIMIT_EXEMPT_PATHS = {
    "/",
    "/api/health",
    "/favicon.ico",
    CHROME_DEVTOOLS_WORKSPACE_PATH,
}
RATE_LIMIT_EXPENSIVE_API_PATHS = {
    "/api/parse",
    "/api/match",
    "/api/analytics/event",
    "/api/employee-assessment/responses",
    "/api/attendance/punch",
    "/api/attendance/timesheet/parse",
    "/api/attendance/timesheet/approve",
    "/api/feedback",
    "/api/support/request",
    "/api/billing/customer-portal/session",
    "/api/seedance/video-demo",
    "/api/sync",
}
RATE_LIMIT_GENERATION_API_PATHS = {
    "/api/knowledge-flow/generate",
}


def verify_credentials(credentials: HTTPBasicCredentials = Depends(security)):
    if not BASIC_AUTH_USERNAME or not BASIC_AUTH_PASSWORD:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Site authentication is not configured",
        )

    correct_username = secrets.compare_digest(credentials.username, BASIC_AUTH_USERNAME)
    correct_password = secrets.compare_digest(credentials.password, BASIC_AUTH_PASSWORD)
    if not (correct_username and correct_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username


def verify_credentials_optional(credentials: Optional[HTTPBasicCredentials] = Depends(security_optional)):
    """Read-only endpoints: allow unauthenticated access; validate if credentials are provided."""
    if credentials is None:
        return None
    correct_username = secrets.compare_digest(credentials.username, BASIC_AUTH_USERNAME)
    correct_password = secrets.compare_digest(credentials.password, BASIC_AUTH_PASSWORD)
    if not (correct_username and correct_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username


def rate_limit_rule_for_request(request: Request) -> Optional[Dict[str, object]]:
    """Return the rate-limit rule for this request, or None when exempt."""
    if not RATE_LIMIT_ENABLED:
        return None

    path = request.url.path
    if path in RATE_LIMIT_EXEMPT_PATHS:
        return None

    if path.startswith("/exports"):
        return {
            "name": "authenticated_exports",
            "path_key": "/exports/*",
            "limit": RATE_LIMIT_AUTH_MAX_REQUESTS,
            "window_seconds": RATE_LIMIT_WINDOW_SECONDS,
        }

    if path == "/admin" or path == "/admin/usage" or path.startswith("/api/admin") or path in {
        "/api/audit/recent",
        "/api/db-test",
        "/api/sales-email/reviews",
        "/api/sales-email/reviews/summary",
        "/api/user-data/export",
    }:
        return {
            "name": "authenticated_admin",
            "path_key": path,
            "limit": RATE_LIMIT_AUTH_MAX_REQUESTS,
            "window_seconds": RATE_LIMIT_WINDOW_SECONDS,
        }

    if request.method == "POST" and path in RATE_LIMIT_GENERATION_API_PATHS:
        return {
            "name": "artifact_generation",
            "path_key": path,
            "limit": RATE_LIMIT_GENERATION_MAX_REQUESTS,
            "window_seconds": RATE_LIMIT_WINDOW_SECONDS,
        }

    if request.method == "POST" and path in RATE_LIMIT_EXPENSIVE_API_PATHS:
        return {
            "name": "expensive_api",
            "path_key": path,
            "limit": RATE_LIMIT_EXPENSIVE_MAX_REQUESTS,
            "window_seconds": RATE_LIMIT_WINDOW_SECONDS,
        }

    if path.startswith("/api/seedance/video-task/"):
        return {
            "name": "api_polling",
            "path_key": "/api/seedance/video-task/*",
            "limit": RATE_LIMIT_MAX_REQUESTS,
            "window_seconds": RATE_LIMIT_WINDOW_SECONDS,
        }

    if path.startswith("/api/"):
        return {
            "name": "api_general",
            "path_key": path,
            "limit": RATE_LIMIT_MAX_REQUESTS,
            "window_seconds": RATE_LIMIT_WINDOW_SECONDS,
        }

    return None


@app.middleware("http")
async def enforce_api_rate_limits(request: Request, call_next):
    rule = rate_limit_rule_for_request(request)
    if rule is None:
        return await call_next(request)

    host = request.client.host if request.client else None
    client_id = client_identifier(request.headers, host)
    path_key = str(rule["path_key"])
    rule_name = str(rule["name"])
    limit = int(rule["limit"])
    window_seconds = int(rule["window_seconds"])
    key = f"{client_id}:{rule_name}:{path_key}"

    decision = api_rate_limiter.allow(
        key,
        limit=limit,
        window_seconds=window_seconds,
    )
    headers = {
        "X-RateLimit-Limit": str(decision.limit),
        "X-RateLimit-Remaining": str(decision.remaining),
        "X-RateLimit-Reset": str(decision.reset_epoch_seconds),
    }

    if not decision.allowed:
        headers["Retry-After"] = str(decision.retry_after_seconds)
        return JSONResponse(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            content={
                "detail": "Too many requests. Please retry after the advertised delay.",
                "rate_limit": {
                    "rule": rule_name,
                    "limit": decision.limit,
                    "window_seconds": window_seconds,
                    "retry_after_seconds": decision.retry_after_seconds,
                },
            },
            headers=headers,
        )

    response = await call_next(request)
    for header_name, header_value in headers.items():
        response.headers[header_name] = header_value
    return response


import sqlite3

# Try to import psycopg2 for Supabase PostgreSQL support, but fallback gracefully to SQLite
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    from psycopg2 import pool as psycopg2_pool
    POSTGRES_AVAILABLE = True
except ImportError:
    POSTGRES_AVAILABLE = False
    psycopg2_pool = None

DATABASE_URL = os.environ.get("SUPABASE_DB_URL", "").strip()
USE_SUPABASE = env_flag("USE_SUPABASE", False)
SUPABASE_DB_CONNECT_TIMEOUT_SECONDS = env_int("SUPABASE_DB_CONNECT_TIMEOUT_SECONDS", 3, 1, 30)
SUPABASE_DB_POOL_MIN = env_int("SUPABASE_DB_POOL_MIN", 1, 1, 10)
SUPABASE_DB_POOL_MAX = env_int("SUPABASE_DB_POOL_MAX", 4, SUPABASE_DB_POOL_MIN, 20)
SUPABASE_DB_POOL_RECYCLE_SECONDS = env_int("SUPABASE_DB_POOL_RECYCLE_SECONDS", 1800, 60, 86400)
SUPABASE_DB_POOL_PRE_PING = env_flag("SUPABASE_DB_POOL_PRE_PING", True)
SUPABASE_DB_APPLICATION_NAME = os.environ.get(
    "SUPABASE_DB_APPLICATION_NAME",
    "mighty-skill-bridge-functions",
).strip() or "mighty-skill-bridge-functions"

_postgres_pool = None
_postgres_pool_created_at = 0.0
_postgres_pool_lock = threading.Lock()


def _is_postgres_url(database_url: str) -> bool:
    return database_url.startswith(("postgresql://", "postgres://"))


def _database_url_pooler_mode(database_url: str) -> str:
    if not database_url:
        return "not_configured"
    try:
        parsed = urlparse(database_url)
    except ValueError:
        return "invalid"
    host = (parsed.hostname or "").lower()
    port = parsed.port
    if "pooler.supabase.com" in host and port == 6543:
        return "supavisor_transaction"
    if "pooler.supabase.com" in host and port == 5432:
        return "supavisor_session"
    if "supabase.co" in host and port == 5432:
        return "direct_ipv6_risk"
    return "custom_postgres"


def get_supabase_pool_status() -> dict:
    """Return non-secret connection pool configuration for health/debug views."""
    return {
        "enabled": bool(USE_SUPABASE and DATABASE_URL and _is_postgres_url(DATABASE_URL) and POSTGRES_AVAILABLE),
        "postgres_available": POSTGRES_AVAILABLE,
        "database_url_configured": bool(DATABASE_URL),
        "pooler_mode": _database_url_pooler_mode(DATABASE_URL),
        "min_connections": SUPABASE_DB_POOL_MIN,
        "max_connections": SUPABASE_DB_POOL_MAX,
        "connect_timeout_seconds": SUPABASE_DB_CONNECT_TIMEOUT_SECONDS,
        "recycle_seconds": SUPABASE_DB_POOL_RECYCLE_SECONDS,
        "pre_ping": SUPABASE_DB_POOL_PRE_PING,
        "application_name": SUPABASE_DB_APPLICATION_NAME,
        "pool_initialized": _postgres_pool is not None,
    }


class PooledPostgresConnection:
    """Small adapter so existing conn.close() calls return the handle to the pool."""

    def __init__(self, raw_connection: Any, pool: Any):
        self._raw_connection = raw_connection
        self._pool = pool
        self._returned = False

    def close(self) -> None:
        if not self._returned:
            self._pool.putconn(self._raw_connection)
            self._returned = True

    def __getattr__(self, name: str) -> Any:
        return getattr(self._raw_connection, name)


def _close_postgres_pool() -> None:
    global _postgres_pool, _postgres_pool_created_at
    with _postgres_pool_lock:
        if _postgres_pool is not None:
            _postgres_pool.closeall()
        _postgres_pool = None
        _postgres_pool_created_at = 0.0


def _get_or_create_postgres_pool():
    global _postgres_pool, _postgres_pool_created_at
    if not POSTGRES_AVAILABLE or psycopg2_pool is None:
        raise RuntimeError("psycopg2 pool support is not available")

    now = time.monotonic()
    with _postgres_pool_lock:
        expired = (
            _postgres_pool is not None
            and now - _postgres_pool_created_at > SUPABASE_DB_POOL_RECYCLE_SECONDS
        )
        if expired:
            _postgres_pool.closeall()
            _postgres_pool = None
            _postgres_pool_created_at = 0.0
        if _postgres_pool is None:
            _postgres_pool = psycopg2_pool.ThreadedConnectionPool(
                SUPABASE_DB_POOL_MIN,
                SUPABASE_DB_POOL_MAX,
                DATABASE_URL,
                connect_timeout=SUPABASE_DB_CONNECT_TIMEOUT_SECONDS,
                application_name=SUPABASE_DB_APPLICATION_NAME,
            )
            _postgres_pool_created_at = now
        return _postgres_pool


def _borrow_postgres_connection():
    pool = _get_or_create_postgres_pool()
    conn = pool.getconn()
    try:
        if SUPABASE_DB_POOL_PRE_PING:
            cur = conn.cursor()
            try:
                cur.execute("SELECT 1;")
                cur.fetchone()
            finally:
                cur.close()
        return PooledPostgresConnection(conn, pool)
    except Exception:
        pool.putconn(conn, close=True)
        raise

def get_db_connection():
    if USE_SUPABASE and DATABASE_URL and _is_postgres_url(DATABASE_URL) and POSTGRES_AVAILABLE:
        try:
            conn = _borrow_postgres_connection()
            return conn, "postgres"
        except Exception as e:
            print(f"[-] Failed to borrow Supabase PostgreSQL pooled connection: {e}. Falling back to SQLite.")
    
    # SQLite Fallback
    # Check if we are running in a Serverless/Container environment with a read-only filesystem
    is_serverless = os.environ.get("K_SERVICE") is not None
    tmp_db_path = os.path.join(tempfile.gettempdir(), "mighty.db")
    
    if is_serverless:
        db_path = tmp_db_path
    else:
        # Check if local directory is writable, fallback to /tmp if not
        try:
            test_file = os.path.join(DATA_DIR, f".write_test_{threading.get_ident()}_{uuid.uuid4().hex}")
            with open(test_file, "w") as f:
                f.write("test")
            os.remove(test_file)
            db_path = os.path.join(DATA_DIR, "mighty.db")
        except Exception:
            db_path = tmp_db_path
            
    conn = sqlite3.connect(db_path, timeout=30.0)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA busy_timeout=30000;")
    # Ensure row-factory is dictionary-like
    conn.row_factory = sqlite3.Row
    return conn, "sqlite"


def init_sales_email_review_tables(cursor: Any, db_type: str) -> None:
    """Create the T817_6 review tables when app-level migrations are absent."""
    if db_type == "postgres":
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS project_requirements (
            id SERIAL PRIMARY KEY,
            title VARCHAR(255) NOT NULL,
            summary TEXT,
            required_skills JSONB NOT NULL DEFAULT '[]'::jsonb,
            nice_to_have_skills JSONB NOT NULL DEFAULT '[]'::jsonb,
            skill_categories JSONB NOT NULL DEFAULT '{}'::jsonb,
            rate_min INTEGER CHECK (rate_min IS NULL OR rate_min >= 0),
            rate_max INTEGER CHECK (rate_max IS NULL OR rate_max >= 0),
            location VARCHAR(160),
            remote_type VARCHAR(32) CHECK (remote_type IS NULL OR remote_type IN ('onsite', 'hybrid', 'remote', 'unknown')),
            start_date_text VARCHAR(120),
            evidence_excerpt TEXT,
            review_status VARCHAR(32) NOT NULL DEFAULT 'pending' CHECK (review_status IN ('pending', 'confirmed', 'corrected', 'rejected')),
            metadata JSONB DEFAULT '{}'::jsonb,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS talent_profiles_from_email (
            id SERIAL PRIMARY KEY,
            anonymized_talent_key VARCHAR(120) NOT NULL UNIQUE,
            summary TEXT,
            skills JSONB NOT NULL DEFAULT '[]'::jsonb,
            skill_categories JSONB NOT NULL DEFAULT '{}'::jsonb,
            desired_rate_min INTEGER CHECK (desired_rate_min IS NULL OR desired_rate_min >= 0),
            desired_rate_max INTEGER CHECK (desired_rate_max IS NULL OR desired_rate_max >= 0),
            desired_location VARCHAR(160),
            remote_preference VARCHAR(32) CHECK (remote_preference IS NULL OR remote_preference IN ('onsite', 'hybrid', 'remote', 'unknown')),
            availability_text VARCHAR(160),
            evidence_excerpt TEXT,
            review_status VARCHAR(32) NOT NULL DEFAULT 'pending' CHECK (review_status IN ('pending', 'confirmed', 'corrected', 'rejected')),
            metadata JSONB DEFAULT '{}'::jsonb,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS email_match_results (
            id SERIAL PRIMARY KEY,
            project_requirement_id INTEGER REFERENCES project_requirements(id) ON DELETE CASCADE,
            talent_profile_id INTEGER REFERENCES talent_profiles_from_email(id) ON DELETE CASCADE,
            engineer_id INTEGER,
            direction VARCHAR(32) NOT NULL CHECK (direction IN ('engineer_to_project', 'project_to_talent')),
            match_score REAL NOT NULL CHECK (match_score >= 0 AND match_score <= 100),
            matched_skills JSONB NOT NULL DEFAULT '[]'::jsonb,
            missing_skills JSONB NOT NULL DEFAULT '[]'::jsonb,
            mismatch_reasons JSONB NOT NULL DEFAULT '[]'::jsonb,
            evidence_summary TEXT,
            review_status VARCHAR(32) NOT NULL DEFAULT 'pending' CHECK (review_status IN ('pending', 'accepted', 'rejected', 'corrected')),
            metadata JSONB DEFAULT '{}'::jsonb,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CHECK (project_requirement_id IS NOT NULL OR talent_profile_id IS NOT NULL OR engineer_id IS NOT NULL)
        );
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS email_match_feedback (
            id SERIAL PRIMARY KEY,
            match_result_id INTEGER NOT NULL REFERENCES email_match_results(id) ON DELETE CASCADE,
            reviewer_id VARCHAR(255),
            feedback_status VARCHAR(32) NOT NULL CHECK (feedback_status IN ('accepted', 'rejected', 'needs_review', 'corrected')),
            corrected_score REAL CHECK (corrected_score IS NULL OR (corrected_score >= 0 AND corrected_score <= 100)),
            corrected_notes TEXT,
            metadata JSONB DEFAULT '{}'::jsonb,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_project_requirements_review_status ON project_requirements(review_status);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_talent_profiles_review_status ON talent_profiles_from_email(review_status);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_email_match_results_score ON email_match_results(match_score DESC);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_email_match_feedback_status ON email_match_feedback(feedback_status);")
        cursor.execute("ALTER TABLE project_requirements ENABLE ROW LEVEL SECURITY;")
        cursor.execute("ALTER TABLE talent_profiles_from_email ENABLE ROW LEVEL SECURITY;")
        cursor.execute("ALTER TABLE email_match_results ENABLE ROW LEVEL SECURITY;")
        cursor.execute("ALTER TABLE email_match_feedback ENABLE ROW LEVEL SECURITY;")
    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS project_requirements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title VARCHAR(255) NOT NULL,
            summary TEXT,
            required_skills TEXT NOT NULL DEFAULT '[]',
            nice_to_have_skills TEXT NOT NULL DEFAULT '[]',
            skill_categories TEXT NOT NULL DEFAULT '{}',
            rate_min INTEGER CHECK (rate_min IS NULL OR rate_min >= 0),
            rate_max INTEGER CHECK (rate_max IS NULL OR rate_max >= 0),
            location VARCHAR(160),
            remote_type VARCHAR(32) CHECK (remote_type IS NULL OR remote_type IN ('onsite', 'hybrid', 'remote', 'unknown')),
            start_date_text VARCHAR(120),
            evidence_excerpt TEXT,
            review_status VARCHAR(32) NOT NULL DEFAULT 'pending' CHECK (review_status IN ('pending', 'confirmed', 'corrected', 'rejected')),
            metadata TEXT DEFAULT '{}',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS talent_profiles_from_email (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            anonymized_talent_key VARCHAR(120) NOT NULL UNIQUE,
            summary TEXT,
            skills TEXT NOT NULL DEFAULT '[]',
            skill_categories TEXT NOT NULL DEFAULT '{}',
            desired_rate_min INTEGER CHECK (desired_rate_min IS NULL OR desired_rate_min >= 0),
            desired_rate_max INTEGER CHECK (desired_rate_max IS NULL OR desired_rate_max >= 0),
            desired_location VARCHAR(160),
            remote_preference VARCHAR(32) CHECK (remote_preference IS NULL OR remote_preference IN ('onsite', 'hybrid', 'remote', 'unknown')),
            availability_text VARCHAR(160),
            evidence_excerpt TEXT,
            review_status VARCHAR(32) NOT NULL DEFAULT 'pending' CHECK (review_status IN ('pending', 'confirmed', 'corrected', 'rejected')),
            metadata TEXT DEFAULT '{}',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS email_match_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_requirement_id INTEGER,
            talent_profile_id INTEGER,
            engineer_id INTEGER,
            direction VARCHAR(32) NOT NULL CHECK (direction IN ('engineer_to_project', 'project_to_talent')),
            match_score REAL NOT NULL CHECK (match_score >= 0 AND match_score <= 100),
            matched_skills TEXT NOT NULL DEFAULT '[]',
            missing_skills TEXT NOT NULL DEFAULT '[]',
            mismatch_reasons TEXT NOT NULL DEFAULT '[]',
            evidence_summary TEXT,
            review_status VARCHAR(32) NOT NULL DEFAULT 'pending' CHECK (review_status IN ('pending', 'accepted', 'rejected', 'corrected')),
            metadata TEXT DEFAULT '{}',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CHECK (project_requirement_id IS NOT NULL OR talent_profile_id IS NOT NULL OR engineer_id IS NOT NULL),
            FOREIGN KEY(project_requirement_id) REFERENCES project_requirements(id) ON DELETE CASCADE,
            FOREIGN KEY(talent_profile_id) REFERENCES talent_profiles_from_email(id) ON DELETE CASCADE,
            FOREIGN KEY(engineer_id) REFERENCES engineers(id) ON DELETE SET NULL
        );
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS email_match_feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            match_result_id INTEGER NOT NULL,
            reviewer_id VARCHAR(255),
            feedback_status VARCHAR(32) NOT NULL CHECK (feedback_status IN ('accepted', 'rejected', 'needs_review', 'corrected')),
            corrected_score REAL CHECK (corrected_score IS NULL OR (corrected_score >= 0 AND corrected_score <= 100)),
            corrected_notes TEXT,
            metadata TEXT DEFAULT '{}',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(match_result_id) REFERENCES email_match_results(id) ON DELETE CASCADE
        );
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_project_requirements_review_status ON project_requirements(review_status);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_talent_profiles_review_status ON talent_profiles_from_email(review_status);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_email_match_results_score ON email_match_results(match_score);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_email_match_feedback_status ON email_match_feedback(feedback_status);")


# --- Storage failure classification (T866 / R114 postmortem) -----------------
# Insert helpers used to swallow exceptions into a generic {"id": 0}, turning
# every schema/connection/constraint problem into an opaque 500. Each failure
# is now classified and tagged with a correlation ID that appears both in the
# server log and in the HTTP 500 detail (no SQL text or personal data leaves
# the server — only the category and the ID).
_LAST_STORAGE_FAILURE: contextvars.ContextVar = contextvars.ContextVar(
    "last_storage_failure", default=None
)

STORAGE_ERROR_CATEGORIES = ("relation_missing", "connection", "constraint", "unknown")


def classify_storage_error(exc: BaseException) -> str:
    text = f"{type(exc).__name__}: {exc}".lower()
    if "no such table" in text or "undefinedtable" in text or (
        "relation" in text and "does not exist" in text
    ):
        return "relation_missing"
    if any(marker in text for marker in (
        "could not connect", "connection refused", "connection reset",
        "server closed the connection", "timeout", "pool", "network",
    )):
        return "connection"
    if "integrityerror" in text or "constraint" in text or "violates" in text:
        return "constraint"
    return "unknown"


def record_storage_failure(operation: str, exc: BaseException) -> str:
    """Classify + log a storage failure; returns the correlation ID."""
    category = classify_storage_error(exc)
    correlation_id = f"st-{uuid.uuid4().hex[:12]}"
    _LAST_STORAGE_FAILURE.set(
        {"operation": operation, "category": category, "correlation_id": correlation_id}
    )
    print(
        f"[-] Storage failure [{correlation_id}] op={operation} category={category} "
        f"error={type(exc).__name__}: {exc}"
    )
    return correlation_id


def storage_failure_detail(message: str) -> str:
    """HTTP 500 detail with category + correlation ID when a failure was recorded."""
    info = _LAST_STORAGE_FAILURE.get()
    if not info:
        return message
    return f"{message} (category={info['category']}, correlation_id={info['correlation_id']})"


def init_db():
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        if db_type == "postgres":
            # PostgreSQL DDL
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS engineers (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) NOT NULL,
                resume_raw TEXT,
                parsed_skills TEXT,
                career_goals TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS jobs (
                id SERIAL PRIMARY KEY,
                title VARCHAR(255) NOT NULL,
                company VARCHAR(100),
                job_description TEXT,
                parsed_requirements TEXT,
                company_culture TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS match_results (
                id SERIAL PRIMARY KEY,
                engineer_id INTEGER REFERENCES engineers(id) ON DELETE CASCADE,
                job_id INTEGER REFERENCES jobs(id) ON DELETE CASCADE,
                fit_ratio REAL NOT NULL,
                score_skill INTEGER,
                score_culture INTEGER,
                score_growth INTEGER,
                score_performing INTEGER,
                match_summary TEXT,
                interview_questions TEXT,
                analyzed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS feedback_events (
                id SERIAL PRIMARY KEY,
                match_result_id INTEGER REFERENCES match_results(id) ON DELETE SET NULL,
                rating VARCHAR(32) NOT NULL CHECK (rating IN ('helpful', 'not_helpful')),
                nps_score INTEGER CHECK (nps_score BETWEEN 0 AND 10),
                comment TEXT,
                source VARCHAR(80) NOT NULL DEFAULT 'diagnosis_report',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata JSONB DEFAULT '{}'::jsonb,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_feedback_events_match_result_id ON feedback_events(match_result_id);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_feedback_events_created_at ON feedback_events(created_at);")
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS support_requests (
                id SERIAL PRIMARY KEY,
                category VARCHAR(32) NOT NULL CHECK (category IN ('general', 'technical', 'billing', 'privacy', 'feedback')),
                priority VARCHAR(16) NOT NULL DEFAULT 'normal' CHECK (priority IN ('normal', 'high', 'urgent')),
                contact_email VARCHAR(254) NOT NULL,
                subject VARCHAR(160) NOT NULL,
                message TEXT NOT NULL,
                status VARCHAR(32) NOT NULL DEFAULT 'new' CHECK (status IN ('new', 'triaged', 'in_progress', 'escalated', 'closed')),
                source VARCHAR(80) NOT NULL DEFAULT 'support_form',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata JSONB DEFAULT '{}'::jsonb,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_support_requests_status_priority ON support_requests(status, priority);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_support_requests_created_at ON support_requests(created_at);")
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS usage_analytics_events (
                id SERIAL PRIMARY KEY,
                event_name VARCHAR(80) NOT NULL
                    CHECK (event_name IN ('page_view', 'section_view', 'cta_click', 'form_submit', 'form_success', 'form_error', 'dashboard_export')),
                event_surface VARCHAR(80) NOT NULL DEFAULT 'public_demo'
                    CHECK (event_surface IN ('public_demo', 'firebase_app', 'internal_console')),
                page_path TEXT,
                session_pseudonym VARCHAR(120) NOT NULL,
                user_agent_family VARCHAR(40) NOT NULL DEFAULT 'unknown',
                metadata JSONB DEFAULT '{}'::jsonb,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_usage_analytics_event_name ON usage_analytics_events(event_name);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_usage_analytics_created_at ON usage_analytics_events(created_at);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_usage_analytics_session ON usage_analytics_events(session_pseudonym);")
            # Supabase exposes public-schema tables through the anon REST API.
            # Enabling RLS with no policies denies anon access entirely, while
            # the app (table owner via the postgres role) bypasses RLS.
            cursor.execute("ALTER TABLE engineers ENABLE ROW LEVEL SECURITY;")
            cursor.execute("ALTER TABLE jobs ENABLE ROW LEVEL SECURITY;")
            cursor.execute("ALTER TABLE match_results ENABLE ROW LEVEL SECURITY;")
            cursor.execute("ALTER TABLE feedback_events ENABLE ROW LEVEL SECURITY;")
            cursor.execute("ALTER TABLE support_requests ENABLE ROW LEVEL SECURITY;")
            cursor.execute("ALTER TABLE usage_analytics_events ENABLE ROW LEVEL SECURITY;")
            cursor.execute("""
            DO $$
            BEGIN
                IF EXISTS (SELECT 1 FROM pg_roles WHERE rolname = 'anon') THEN
                    REVOKE ALL ON TABLE usage_analytics_events FROM anon;
                END IF;
                IF EXISTS (SELECT 1 FROM pg_roles WHERE rolname = 'authenticated') THEN
                    REVOKE ALL ON TABLE usage_analytics_events FROM authenticated;
                END IF;
            END $$;
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS employee_assessment_responses (
                id SERIAL PRIMARY KEY,
                subject_pseudonym VARCHAR(120) NOT NULL,
                department_bucket VARCHAR(80) NOT NULL,
                motivation_level INTEGER NOT NULL CHECK (motivation_level BETWEEN 1 AND 5),
                culture_level INTEGER NOT NULL CHECK (culture_level BETWEEN 1 AND 5),
                growth_support_excerpt TEXT NOT NULL DEFAULT '',
                consent_version VARCHAR(80) NOT NULL,
                consented_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status VARCHAR(32) NOT NULL DEFAULT 'pending_review'
                    CHECK (status IN ('pending_review', 'reviewed', 'deleted')),
                source VARCHAR(80) NOT NULL DEFAULT 'employee_assessment_form',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata JSONB DEFAULT '{}'::jsonb,
                deletion_due_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_employee_assessment_subject ON employee_assessment_responses(subject_pseudonym);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_employee_assessment_created_at ON employee_assessment_responses(created_at);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_employee_assessment_status ON employee_assessment_responses(status);")
            cursor.execute("ALTER TABLE employee_assessment_responses ENABLE ROW LEVEL SECURITY;")
            cursor.execute("""
            DO $$
            BEGIN
                IF EXISTS (SELECT 1 FROM pg_roles WHERE rolname = 'anon') THEN
                    REVOKE ALL ON TABLE employee_assessment_responses FROM anon;
                END IF;
                IF EXISTS (SELECT 1 FROM pg_roles WHERE rolname = 'authenticated') THEN
                    REVOKE ALL ON TABLE employee_assessment_responses FROM authenticated;
                END IF;
            END $$;
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_punch_events (
                id SERIAL PRIMARY KEY,
                subject_pseudonym VARCHAR(120) NOT NULL,
                event_type VARCHAR(32) NOT NULL
                    CHECK (event_type IN ('clock_in', 'clock_out', 'break_start', 'break_end')),
                recorded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                source VARCHAR(80) NOT NULL DEFAULT 'attendance_widget',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata JSONB DEFAULT '{}'::jsonb,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_timesheet_imports (
                id SERIAL PRIMARY KEY,
                subject_pseudonym VARCHAR(120) NOT NULL,
                file_digest VARCHAR(80) NOT NULL,
                file_extension VARCHAR(16) NOT NULL,
                work_minutes INTEGER NOT NULL DEFAULT 0 CHECK (work_minutes >= 0),
                overtime_minutes INTEGER NOT NULL DEFAULT 0 CHECK (overtime_minutes >= 0),
                holiday_work_days INTEGER NOT NULL DEFAULT 0 CHECK (holiday_work_days >= 0),
                midnight_minutes INTEGER NOT NULL DEFAULT 0 CHECK (midnight_minutes >= 0),
                anomaly_count INTEGER NOT NULL DEFAULT 0 CHECK (anomaly_count >= 0),
                status VARCHAR(32) NOT NULL DEFAULT 'pending_approval'
                    CHECK (status IN ('pending_approval', 'approved', 'rejected', 'manual_review')),
                consent_version VARCHAR(80) NOT NULL,
                source VARCHAR(80) NOT NULL DEFAULT 'attendance_timesheet_upload',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata JSONB DEFAULT '{}'::jsonb,
                approved_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_punch_subject ON attendance_punch_events(subject_pseudonym);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_punch_recorded_at ON attendance_punch_events(recorded_at);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_timesheet_subject ON attendance_timesheet_imports(subject_pseudonym);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_timesheet_status ON attendance_timesheet_imports(status);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_timesheet_created_at ON attendance_timesheet_imports(created_at);")
            cursor.execute("ALTER TABLE attendance_punch_events ENABLE ROW LEVEL SECURITY;")
            cursor.execute("ALTER TABLE attendance_timesheet_imports ENABLE ROW LEVEL SECURITY;")
            cursor.execute("""
            DO $$
            BEGIN
                IF EXISTS (SELECT 1 FROM pg_roles WHERE rolname = 'anon') THEN
                    REVOKE ALL ON TABLE attendance_punch_events FROM anon;
                    REVOKE ALL ON TABLE attendance_timesheet_imports FROM anon;
                END IF;
                IF EXISTS (SELECT 1 FROM pg_roles WHERE rolname = 'authenticated') THEN
                    REVOKE ALL ON TABLE attendance_punch_events FROM authenticated;
                    REVOKE ALL ON TABLE attendance_timesheet_imports FROM authenticated;
                END IF;
            END $$;
            """)
        else:
            # SQLite DDL
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS engineers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(100) NOT NULL,
                resume_raw TEXT,
                parsed_skills TEXT,
                career_goals TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title VARCHAR(255) NOT NULL,
                company VARCHAR(100),
                job_description TEXT,
                parsed_requirements TEXT,
                company_culture TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS match_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                engineer_id INTEGER,
                job_id INTEGER,
                fit_ratio REAL NOT NULL,
                score_skill INTEGER,
                score_culture INTEGER,
                score_growth INTEGER,
                score_performing INTEGER,
                match_summary TEXT,
                interview_questions TEXT,
                analyzed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(engineer_id) REFERENCES engineers(id) ON DELETE CASCADE,
                FOREIGN KEY(job_id) REFERENCES jobs(id) ON DELETE CASCADE
            );
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS feedback_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                match_result_id INTEGER,
                rating VARCHAR(32) NOT NULL CHECK (rating IN ('helpful', 'not_helpful')),
                nps_score INTEGER CHECK (nps_score IS NULL OR (nps_score BETWEEN 0 AND 10)),
                comment TEXT,
                source VARCHAR(80) NOT NULL DEFAULT 'diagnosis_report',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata TEXT DEFAULT '{}',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(match_result_id) REFERENCES match_results(id) ON DELETE SET NULL
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_feedback_events_match_result_id ON feedback_events(match_result_id);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_feedback_events_created_at ON feedback_events(created_at);")
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS support_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category VARCHAR(32) NOT NULL CHECK (category IN ('general', 'technical', 'billing', 'privacy', 'feedback')),
                priority VARCHAR(16) NOT NULL DEFAULT 'normal' CHECK (priority IN ('normal', 'high', 'urgent')),
                contact_email VARCHAR(254) NOT NULL,
                subject VARCHAR(160) NOT NULL,
                message TEXT NOT NULL,
                status VARCHAR(32) NOT NULL DEFAULT 'new' CHECK (status IN ('new', 'triaged', 'in_progress', 'escalated', 'closed')),
                source VARCHAR(80) NOT NULL DEFAULT 'support_form',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata TEXT DEFAULT '{}',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_support_requests_status_priority ON support_requests(status, priority);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_support_requests_created_at ON support_requests(created_at);")
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS usage_analytics_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_name VARCHAR(80) NOT NULL
                    CHECK (event_name IN ('page_view', 'section_view', 'cta_click', 'form_submit', 'form_success', 'form_error', 'dashboard_export')),
                event_surface VARCHAR(80) NOT NULL DEFAULT 'public_demo'
                    CHECK (event_surface IN ('public_demo', 'firebase_app', 'internal_console')),
                page_path TEXT,
                session_pseudonym VARCHAR(120) NOT NULL,
                user_agent_family VARCHAR(40) NOT NULL DEFAULT 'unknown',
                metadata TEXT DEFAULT '{}',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_usage_analytics_event_name ON usage_analytics_events(event_name);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_usage_analytics_created_at ON usage_analytics_events(created_at);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_usage_analytics_session ON usage_analytics_events(session_pseudonym);")
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS employee_assessment_responses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                subject_pseudonym VARCHAR(120) NOT NULL,
                department_bucket VARCHAR(80) NOT NULL,
                motivation_level INTEGER NOT NULL CHECK (motivation_level BETWEEN 1 AND 5),
                culture_level INTEGER NOT NULL CHECK (culture_level BETWEEN 1 AND 5),
                growth_support_excerpt TEXT NOT NULL DEFAULT '',
                consent_version VARCHAR(80) NOT NULL,
                consented_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status VARCHAR(32) NOT NULL DEFAULT 'pending_review'
                    CHECK (status IN ('pending_review', 'reviewed', 'deleted')),
                source VARCHAR(80) NOT NULL DEFAULT 'employee_assessment_form',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata TEXT DEFAULT '{}',
                deletion_due_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_employee_assessment_subject ON employee_assessment_responses(subject_pseudonym);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_employee_assessment_created_at ON employee_assessment_responses(created_at);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_employee_assessment_status ON employee_assessment_responses(status);")
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_punch_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                subject_pseudonym VARCHAR(120) NOT NULL,
                event_type VARCHAR(32) NOT NULL
                    CHECK (event_type IN ('clock_in', 'clock_out', 'break_start', 'break_end')),
                recorded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                source VARCHAR(80) NOT NULL DEFAULT 'attendance_widget',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata TEXT DEFAULT '{}',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_timesheet_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                subject_pseudonym VARCHAR(120) NOT NULL,
                file_digest VARCHAR(80) NOT NULL,
                file_extension VARCHAR(16) NOT NULL,
                work_minutes INTEGER NOT NULL DEFAULT 0 CHECK (work_minutes >= 0),
                overtime_minutes INTEGER NOT NULL DEFAULT 0 CHECK (overtime_minutes >= 0),
                holiday_work_days INTEGER NOT NULL DEFAULT 0 CHECK (holiday_work_days >= 0),
                midnight_minutes INTEGER NOT NULL DEFAULT 0 CHECK (midnight_minutes >= 0),
                anomaly_count INTEGER NOT NULL DEFAULT 0 CHECK (anomaly_count >= 0),
                status VARCHAR(32) NOT NULL DEFAULT 'pending_approval'
                    CHECK (status IN ('pending_approval', 'approved', 'rejected', 'manual_review')),
                consent_version VARCHAR(80) NOT NULL,
                source VARCHAR(80) NOT NULL DEFAULT 'attendance_timesheet_upload',
                page_url TEXT,
                session_id VARCHAR(120),
                metadata TEXT DEFAULT '{}',
                approved_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_punch_subject ON attendance_punch_events(subject_pseudonym);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_punch_recorded_at ON attendance_punch_events(recorded_at);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_timesheet_subject ON attendance_timesheet_imports(subject_pseudonym);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_timesheet_status ON attendance_timesheet_imports(status);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_timesheet_created_at ON attendance_timesheet_imports(created_at);")
        init_sales_email_review_tables(cursor, db_type)
        conn.commit()
        print(f"[+] Database tables initialized successfully ({db_type}).")
    except Exception as e:
        print(f"[-] Database initialization error: {e}")
    finally:
        cursor.close()
        conn.close()

def db_insert_engineer(name: str, resume_raw: str, parsed_skills: dict, career_goals: dict) -> int:
    if SUPABASE_SDK_ACTIVE:
        inserted_id = sdk_insert_engineer(name, resume_raw, parsed_skills, career_goals)
        if inserted_id:
            return inserted_id

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        if db_type == "postgres":
            cursor.execute(
                "INSERT INTO engineers (name, resume_raw, parsed_skills, career_goals) VALUES (%s, %s, %s, %s) RETURNING id;",
                (name, resume_raw, json.dumps(parsed_skills, ensure_ascii=False), json.dumps(career_goals, ensure_ascii=False))
            )
            inserted_id = cursor.fetchone()[0]
        else:
            cursor.execute(
                "INSERT INTO engineers (name, resume_raw, parsed_skills, career_goals) VALUES (?, ?, ?, ?);",
                (name, resume_raw, json.dumps(parsed_skills, ensure_ascii=False), json.dumps(career_goals, ensure_ascii=False))
            )
            inserted_id = cursor.lastrowid
        conn.commit()
        return inserted_id
    except Exception as e:
        record_storage_failure("insert_engineer", e)
        return 0
    finally:
        cursor.close()
        conn.close()

def db_insert_job(title: str, company: str, job_description: str, parsed_requirements: dict, company_culture: dict) -> int:
    if SUPABASE_SDK_ACTIVE:
        inserted_id = sdk_insert_job(title, company, job_description, parsed_requirements, company_culture)
        if inserted_id:
            return inserted_id

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        if db_type == "postgres":
            cursor.execute(
                "INSERT INTO jobs (title, company, job_description, parsed_requirements, company_culture) VALUES (%s, %s, %s, %s, %s) RETURNING id;",
                (title, company, job_description, json.dumps(parsed_requirements, ensure_ascii=False), json.dumps(company_culture, ensure_ascii=False))
            )
            inserted_id = cursor.fetchone()[0]
        else:
            cursor.execute(
                "INSERT INTO jobs (title, company, job_description, parsed_requirements, company_culture) VALUES (?, ?, ?, ?, ?);",
                (title, company, job_description, json.dumps(parsed_requirements, ensure_ascii=False), json.dumps(company_culture, ensure_ascii=False))
            )
            inserted_id = cursor.lastrowid
        conn.commit()
        return inserted_id
    except Exception as e:
        record_storage_failure("insert_job", e)
        return 0
    finally:
        cursor.close()
        conn.close()

def db_insert_match_result(engineer_id: int, job_id: int, fit_ratio: float, score_skill: int, score_culture: int, score_growth: int, score_performing: int, match_summary: str, interview_questions: list) -> int:
    if SUPABASE_SDK_ACTIVE:
        inserted_id = sdk_insert_match_result(
            engineer_id, job_id, fit_ratio, score_skill, score_culture, score_growth, score_performing, match_summary, interview_questions
        )
        if inserted_id:
            return inserted_id

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        if db_type == "postgres":
            cursor.execute(
                "INSERT INTO match_results (engineer_id, job_id, fit_ratio, score_skill, score_culture, score_growth, score_performing, match_summary, interview_questions) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;",
                (engineer_id, job_id, fit_ratio, score_skill, score_culture, score_growth, score_performing, match_summary, json.dumps(interview_questions, ensure_ascii=False))
            )
            inserted_id = cursor.fetchone()[0]
        else:
            cursor.execute(
                "INSERT INTO match_results (engineer_id, job_id, fit_ratio, score_skill, score_culture, score_growth, score_performing, match_summary, interview_questions) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);",
                (engineer_id, job_id, fit_ratio, score_skill, score_culture, score_growth, score_performing, match_summary, json.dumps(interview_questions, ensure_ascii=False))
            )
            inserted_id = cursor.lastrowid
        conn.commit()
        return inserted_id
    except Exception as e:
        record_storage_failure("insert_match_result", e)
        return 0
    finally:
        cursor.close()
        conn.close()


VALID_FEEDBACK_RATINGS = {"helpful", "not_helpful"}
MAX_FEEDBACK_COMMENT_LENGTH = 1000
VALID_SUPPORT_CATEGORIES = {"general", "technical", "billing", "privacy", "feedback"}
VALID_SUPPORT_PRIORITIES = {"normal", "high", "urgent"}
SUPPORT_STATUSES = {"new", "triaged", "in_progress", "escalated", "closed"}
MAX_SUPPORT_SUBJECT_LENGTH = 160
MAX_SUPPORT_MESSAGE_LENGTH = 3000
SUPPORT_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
LEGAL_CONSENT_VERSION = "MSB-LEGAL-2026-07-GA"
LEGAL_CONSENT_DOCS = (
    "TERMS_OF_SERVICE.md",
    "PRIVACY_POLICY.md",
    "TOKUSHOHO_NOTATION.md",
    "BILLING_AND_REFUND_POLICY.md",
)
# --------------------------------------------------------------------------- #
# Onboarding / activation flow (T752, gate PUBLIC-06)
#
# Internal GA issues accounts administratively (T833), so the flow that matters
# is activation of an issued account rather than self-signup. The step catalogue
# is served to the wizard (GET /api/onboarding/state) and is the same list the
# activation validator checks, so the UI and the gate cannot drift apart.
#
# Progress is kept client-side (localStorage) and only a pseudonymized subject is
# audited on activation — no per-user row is stored, keeping this flow clear of
# the still-open T798 legal review.
# --------------------------------------------------------------------------- #
ONBOARDING_FLOW_VERSION = "MSB-ONBOARDING-2026-07"
ONBOARDING_PSEUDONYM_SALT = os.environ.get(
    "ONBOARDING_PSEUDONYM_SALT",
    "mighty-link-onboarding-local-salt-v1",
)
MAX_ONBOARDING_IDENTIFIER_LENGTH = 120
ONBOARDING_STEPS: list[dict] = [
    {
        "id": "account",
        "title": "アカウントの受領とサインイン",
        "description": "管理者が発行したアカウントでサインインし、本人のメールアドレスであることを確認します。",
        "required": True,
    },
    {
        "id": "legal_consent",
        "title": "利用規約・プライバシーポリシーへの同意",
        "description": "画面上部の同意欄で、利用規約・プライバシーポリシー・特商法表記・課金規約に同意します。",
        "required": True,
    },
    {
        "id": "profile",
        "title": "表示名と所属の初期設定",
        "description": "社内で識別できる表示名と所属部署を設定します。氏名などの個人情報は入力不要です。",
        "required": True,
    },
    {
        "id": "first_analysis",
        "title": "スキルシートで初回診断を試す",
        "description": "サンプルまたは自社のスキルシートを取り込み、フィット分析を一度実行して結果の見方を確認します。",
        "required": False,
    },
    {
        "id": "guide",
        "title": "利用ガイドとサポート窓口の確認",
        "description": "利用ガイド・FAQ とサポート窓口の連絡先を確認します。",
        "required": False,
    },
]
ONBOARDING_STEP_IDS = {step["id"] for step in ONBOARDING_STEPS}
ONBOARDING_REQUIRED_STEP_IDS = [step["id"] for step in ONBOARDING_STEPS if step["required"]]

EMPLOYEE_ASSESSMENT_CONSENT_VERSION = "MSB-EMP-ASSESS-2026-06"
EMPLOYEE_ASSESSMENT_PSEUDONYM_SALT = os.environ.get(
    "EMPLOYEE_ASSESSMENT_PSEUDONYM_SALT",
    "mighty-link-employee-assessment-local-salt-v1",
)
MAX_EMPLOYEE_IDENTIFIER_LENGTH = 120
MAX_EMPLOYEE_DEPARTMENT_LENGTH = 80
MAX_EMPLOYEE_ASSESSMENT_FEEDBACK_LENGTH = 1000
EMPLOYEE_ASSESSMENT_RETENTION_DAYS = env_int("EMPLOYEE_ASSESSMENT_RETENTION_DAYS", 180, 1, 1095)
ATTENDANCE_CONSENT_VERSION = "MSB-ATTENDANCE-2026-06"
ATTENDANCE_PSEUDONYM_SALT = os.environ.get(
    "ATTENDANCE_PSEUDONYM_SALT",
    "mighty-link-attendance-local-salt-v1",
)
MAX_ATTENDANCE_IDENTIFIER_LENGTH = 120
MAX_ATTENDANCE_FILE_BYTES = env_int("MAX_ATTENDANCE_FILE_BYTES", 200_000, 1_000, 2_000_000)
ATTENDANCE_EVENT_TYPE_ALIASES = {
    "in": "clock_in",
    "clock_in": "clock_in",
    "clock-in": "clock_in",
    "out": "clock_out",
    "clock_out": "clock_out",
    "clock-out": "clock_out",
    "rest-start": "break_start",
    "rest_start": "break_start",
    "break-start": "break_start",
    "break_start": "break_start",
    "rest-end": "break_end",
    "rest_end": "break_end",
    "break-end": "break_end",
    "break_end": "break_end",
}
VALID_ATTENDANCE_DECISIONS = {"approved", "rejected"}
VALID_USAGE_ANALYTICS_EVENTS = {
    "page_view",
    "section_view",
    "cta_click",
    "form_submit",
    "form_success",
    "form_error",
    "dashboard_export",
}
VALID_USAGE_ANALYTICS_SURFACES = {"public_demo", "firebase_app", "internal_console"}
MAX_USAGE_ANALYTICS_METADATA_KEYS = 12
USAGE_ANALYTICS_PSEUDONYM_SALT = os.environ.get(
    "USAGE_ANALYTICS_PSEUDONYM_SALT",
    "mighty-link-usage-analytics-local-salt-v1",
)
SENSITIVE_EMAIL_RE = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
SENSITIVE_PHONE_RE = re.compile(r"(?<!\d)(?:\+?\d[\d\s().-]{7,}\d)(?!\d)")
SENSITIVE_SECRET_RE = re.compile(
    r"(Bearer\s+[A-Za-z0-9._=-]+|(?:api[_-]?key|token|secret|password)\s*[:=]\s*[A-Za-z0-9._=-]+)",
    re.IGNORECASE,
)


def clean_feedback_text(value: Optional[str], limit: int) -> str:
    if value is None:
        return ""
    normalized = re.sub(r"\s+", " ", str(value)).strip()
    return normalized[:limit]


def redact_sensitive_text(value: Optional[str], limit: int) -> str:
    text = clean_feedback_text(value, limit)
    text = SENSITIVE_EMAIL_RE.sub("<email:redacted>", text)
    text = SENSITIVE_PHONE_RE.sub("<phone:redacted>", text)
    text = SENSITIVE_SECRET_RE.sub("<secret:redacted>", text)
    return text[:limit]


def validate_legal_consent(accepted: bool, consent_version: Optional[str], source: str) -> dict:
    version = clean_feedback_text(consent_version, 80)
    if not accepted:
        raise HTTPException(
            status_code=400,
            detail="Terms of Service and Privacy Policy consent is required before running this API.",
        )
    if version != LEGAL_CONSENT_VERSION:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid legal consent version. Expected {LEGAL_CONSENT_VERSION}.",
        )
    return {
        "accepted": True,
        "version": version,
        "source": clean_feedback_text(source, 80) or "unknown",
        "docs": list(LEGAL_CONSENT_DOCS),
    }


def db_insert_feedback_event(
    match_result_id: Optional[int],
    rating: str,
    nps_score: Optional[int],
    comment: Optional[str],
    source: str,
    page_url: Optional[str],
    session_id: Optional[str],
    metadata: Optional[dict] = None,
) -> int:
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    clean_comment = clean_feedback_text(comment, MAX_FEEDBACK_COMMENT_LENGTH)
    clean_source = clean_feedback_text(source, 80) or "diagnosis_report"
    clean_page_url = clean_feedback_text(page_url, 500)
    clean_session_id = clean_feedback_text(session_id, 120)
    metadata_json = json.dumps(metadata or {}, ensure_ascii=False)
    try:
        if db_type == "postgres":
            cursor.execute(
                """
                INSERT INTO feedback_events
                    (match_result_id, rating, nps_score, comment, source, page_url, session_id, metadata)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                RETURNING id;
                """,
                (
                    match_result_id,
                    rating,
                    nps_score,
                    clean_comment,
                    clean_source,
                    clean_page_url,
                    clean_session_id,
                    metadata_json,
                ),
            )
            inserted_id = cursor.fetchone()[0]
        else:
            cursor.execute(
                """
                INSERT INTO feedback_events
                    (match_result_id, rating, nps_score, comment, source, page_url, session_id, metadata)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?);
                """,
                (
                    match_result_id,
                    rating,
                    nps_score,
                    clean_comment,
                    clean_source,
                    clean_page_url,
                    clean_session_id,
                    metadata_json,
                ),
            )
            inserted_id = cursor.lastrowid
        conn.commit()
        return inserted_id
    except Exception as e:
        record_storage_failure("insert_feedback_event", e)
        return 0
    finally:
        cursor.close()
        conn.close()


def _scalar(row: Any, index: int = 0) -> Any:
    return row[index] if row is not None else None


def db_get_feedback_summary(limit: int = 20) -> dict:
    limit = max(1, min(limit, 100))
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM feedback_events;")
        total = int(_scalar(cursor.fetchone()) or 0)

        cursor.execute("SELECT rating, COUNT(*) FROM feedback_events GROUP BY rating;")
        rating_counts = {"helpful": 0, "not_helpful": 0}
        for row in cursor.fetchall():
            rating_counts[str(row[0])] = int(row[1])

        cursor.execute("SELECT AVG(nps_score), COUNT(nps_score) FROM feedback_events WHERE nps_score IS NOT NULL;")
        nps_row = cursor.fetchone()
        nps_average = float(nps_row[0]) if nps_row and nps_row[0] is not None else None
        nps_count = int(nps_row[1]) if nps_row else 0

        columns = [
            "id",
            "match_result_id",
            "rating",
            "nps_score",
            "comment",
            "source",
            "page_url",
            "session_id",
            "created_at",
        ]
        if db_type == "postgres":
            cursor.execute(
                """
                SELECT id, match_result_id, rating, nps_score, comment, source, page_url, session_id, created_at
                FROM feedback_events
                ORDER BY id DESC
                LIMIT %s;
                """,
                (limit,),
            )
        else:
            cursor.execute(
                """
                SELECT id, match_result_id, rating, nps_score, comment, source, page_url, session_id, created_at
                FROM feedback_events
                ORDER BY id DESC
                LIMIT ?;
                """,
                (limit,),
            )

        recent = []
        for row in cursor.fetchall():
            row_dict = dict(row) if isinstance(row, sqlite3.Row) else dict(zip(columns, row))
            recent.append({
                "id": row_dict.get("id"),
                "match_result_id": row_dict.get("match_result_id"),
                "rating": row_dict.get("rating"),
                "nps_score": row_dict.get("nps_score"),
                "comment_excerpt": clean_feedback_text(row_dict.get("comment"), 160),
                "source": row_dict.get("source"),
                "page_url": row_dict.get("page_url"),
                "session_id": row_dict.get("session_id"),
                "created_at": str(row_dict.get("created_at") or ""),
            })

        return {
            "total": total,
            "rating_counts": rating_counts,
            "nps": {
                "average": round(nps_average, 2) if nps_average is not None else None,
                "count": nps_count,
            },
            "recent": recent,
        }
    except Exception as e:
        print(f"[-] Database feedback summary failed: {e}")
        return {
            "total": 0,
            "rating_counts": {"helpful": 0, "not_helpful": 0},
            "nps": {"average": None, "count": 0},
            "recent": [],
            "error": str(e),
        }
    finally:
        cursor.close()
        conn.close()


def usage_analytics_session_pseudonym(session_id: Optional[str]) -> str:
    clean_session = clean_feedback_text(session_id, 120)
    if not clean_session:
        clean_session = "anonymous-session"
    digest = stable_digest(f"{USAGE_ANALYTICS_PSEUDONYM_SALT}:{clean_session.lower()}")
    return f"usage-{digest}"


def usage_analytics_page_path(page_url: Optional[str]) -> str:
    clean_page_url = clean_feedback_text(page_url, 500)
    if not clean_page_url:
        return "/"
    parsed = urlparse(clean_page_url)
    if parsed.scheme or parsed.netloc:
        path = parsed.path or "/"
        fragment = f"#{parsed.fragment}" if parsed.fragment else ""
        return clean_feedback_text(f"{path}{fragment}", 500)
    path_without_query = clean_page_url.split("?", 1)[0]
    return clean_feedback_text(path_without_query or "/", 500)


def usage_analytics_user_agent_family(user_agent: Optional[str]) -> str:
    clean_user_agent = clean_feedback_text(user_agent, 300).lower()
    if not clean_user_agent:
        return "unknown"
    if "bot" in clean_user_agent or "crawler" in clean_user_agent or "spider" in clean_user_agent:
        return "bot"
    if "edg/" in clean_user_agent or "edge/" in clean_user_agent:
        return "edge"
    if "chrome/" in clean_user_agent or "crios/" in clean_user_agent:
        return "chrome"
    if "firefox/" in clean_user_agent or "fxios/" in clean_user_agent:
        return "firefox"
    if "safari/" in clean_user_agent:
        return "safari"
    return "other"


def sanitize_usage_analytics_metadata(metadata: Optional[dict]) -> dict:
    if not isinstance(metadata, dict):
        return {}

    safe_metadata: dict = {}
    for key, value in metadata.items():
        if len(safe_metadata) >= MAX_USAGE_ANALYTICS_METADATA_KEYS:
            break
        safe_key = re.sub(r"[^a-zA-Z0-9_.:-]", "_", str(key).strip())[:40]
        if not safe_key:
            continue
        lower_key = safe_key.lower()
        if any(token in lower_key for token in ("email", "phone", "name", "token", "secret", "password")):
            safe_metadata[safe_key] = "<redacted>"
            continue
        if isinstance(value, bool) or isinstance(value, int) or isinstance(value, float):
            safe_metadata[safe_key] = value
        elif value is None:
            safe_metadata[safe_key] = None
        else:
            safe_metadata[safe_key] = redact_sensitive_text(value, 160)
    return safe_metadata


def db_insert_usage_analytics_event(
    event_name: str,
    event_surface: str,
    page_url: Optional[str],
    session_id: Optional[str],
    user_agent: Optional[str],
    metadata: Optional[dict] = None,
) -> dict:
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    clean_event_name = clean_feedback_text(event_name, 80).lower()
    clean_event_surface = clean_feedback_text(event_surface, 80).lower() or "public_demo"
    page_path = usage_analytics_page_path(page_url)
    session_pseudonym = usage_analytics_session_pseudonym(session_id)
    user_agent_family = usage_analytics_user_agent_family(user_agent)
    metadata_json = json.dumps(sanitize_usage_analytics_metadata(metadata), ensure_ascii=False)
    try:
        if db_type == "postgres":
            cursor.execute(
                """
                INSERT INTO usage_analytics_events
                    (event_name, event_surface, page_path, session_pseudonym, user_agent_family, metadata)
                VALUES (%s, %s, %s, %s, %s, %s::jsonb)
                RETURNING id;
                """,
                (
                    clean_event_name,
                    clean_event_surface,
                    page_path,
                    session_pseudonym,
                    user_agent_family,
                    metadata_json,
                ),
            )
            inserted_id = int(cursor.fetchone()[0])
        else:
            cursor.execute(
                """
                INSERT INTO usage_analytics_events
                    (event_name, event_surface, page_path, session_pseudonym, user_agent_family, metadata)
                VALUES (?, ?, ?, ?, ?, ?);
                """,
                (
                    clean_event_name,
                    clean_event_surface,
                    page_path,
                    session_pseudonym,
                    user_agent_family,
                    metadata_json,
                ),
            )
            inserted_id = int(cursor.lastrowid)
        conn.commit()
        return {
            "id": inserted_id,
            "event_name": clean_event_name,
            "event_surface": clean_event_surface,
            "page_path": page_path,
            "session_pseudonym": session_pseudonym,
            "user_agent_family": user_agent_family,
        }
    except Exception as e:
        record_storage_failure("insert_usage_analytics_event", e)
        return {"id": 0, "error": str(e)}
    finally:
        cursor.close()
        conn.close()


def db_get_usage_analytics_summary(limit: int = 20) -> dict:
    limit = max(1, min(limit, 100))
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM usage_analytics_events;")
        total = int(_scalar(cursor.fetchone()) or 0)

        cursor.execute("SELECT event_name, COUNT(*) FROM usage_analytics_events GROUP BY event_name;")
        event_counts = {event_name: 0 for event_name in sorted(VALID_USAGE_ANALYTICS_EVENTS)}
        for row in cursor.fetchall():
            event_counts[str(row[0])] = int(row[1])

        cursor.execute("SELECT event_surface, COUNT(*) FROM usage_analytics_events GROUP BY event_surface;")
        surface_counts = {surface_name: 0 for surface_name in sorted(VALID_USAGE_ANALYTICS_SURFACES)}
        for row in cursor.fetchall():
            surface_counts[str(row[0])] = int(row[1])

        if db_type == "postgres":
            cursor.execute(
                """
                SELECT COUNT(*), COUNT(DISTINCT session_pseudonym)
                FROM usage_analytics_events
                WHERE created_at >= CURRENT_TIMESTAMP - INTERVAL '7 days';
                """
            )
        else:
            cursor.execute(
                """
                SELECT COUNT(*), COUNT(DISTINCT session_pseudonym)
                FROM usage_analytics_events
                WHERE datetime(created_at) >= datetime('now', '-7 days');
                """
            )
        recent_row = cursor.fetchone()
        events_last_7_days = int(recent_row[0]) if recent_row and recent_row[0] is not None else 0
        unique_sessions_last_7_days = int(recent_row[1]) if recent_row and recent_row[1] is not None else 0

        columns = [
            "id",
            "event_name",
            "event_surface",
            "page_path",
            "session_pseudonym",
            "user_agent_family",
            "created_at",
        ]
        if db_type == "postgres":
            cursor.execute(
                """
                SELECT id, event_name, event_surface, page_path, session_pseudonym, user_agent_family, created_at
                FROM usage_analytics_events
                ORDER BY id DESC
                LIMIT %s;
                """,
                (limit,),
            )
        else:
            cursor.execute(
                """
                SELECT id, event_name, event_surface, page_path, session_pseudonym, user_agent_family, created_at
                FROM usage_analytics_events
                ORDER BY id DESC
                LIMIT ?;
                """,
                (limit,),
            )

        recent = []
        for row in cursor.fetchall():
            row_dict = dict(row) if isinstance(row, sqlite3.Row) else dict(zip(columns, row))
            recent.append({
                "id": row_dict.get("id"),
                "event_name": row_dict.get("event_name"),
                "event_surface": row_dict.get("event_surface"),
                "page_path": row_dict.get("page_path"),
                "session_pseudonym": row_dict.get("session_pseudonym"),
                "user_agent_family": row_dict.get("user_agent_family"),
                "created_at": str(row_dict.get("created_at") or ""),
            })

        return {
            "total": total,
            "events_last_7_days": events_last_7_days,
            "unique_sessions_last_7_days": unique_sessions_last_7_days,
            "event_counts": event_counts,
            "surface_counts": surface_counts,
            "recent": recent,
            "privacy_controls": {
                "raw_session_id_stored": False,
                "ip_address_stored": False,
                "raw_user_agent_stored": False,
                "form_contents_stored": False,
                "metadata_sensitive_fields_redacted": True,
            },
        }
    except Exception as e:
        print(f"[-] Database usage analytics summary failed: {e}")
        return {
            "total": 0,
            "events_last_7_days": 0,
            "unique_sessions_last_7_days": 0,
            "event_counts": {event_name: 0 for event_name in sorted(VALID_USAGE_ANALYTICS_EVENTS)},
            "surface_counts": {surface_name: 0 for surface_name in sorted(VALID_USAGE_ANALYTICS_SURFACES)},
            "recent": [],
            "privacy_controls": {
                "raw_session_id_stored": False,
                "ip_address_stored": False,
                "raw_user_agent_stored": False,
                "form_contents_stored": False,
                "metadata_sensitive_fields_redacted": True,
            },
            "error": str(e),
        }
    finally:
        cursor.close()
        conn.close()


def employee_assessment_pseudonym(employee_identifier: str) -> str:
    clean_identifier = clean_feedback_text(employee_identifier, MAX_EMPLOYEE_IDENTIFIER_LENGTH).lower()
    digest = stable_digest(f"{EMPLOYEE_ASSESSMENT_PSEUDONYM_SALT}:{clean_identifier}")
    return f"emp-assess-{digest}"


def db_insert_employee_assessment_response(
    employee_identifier: str,
    department: str,
    motivation_level: int,
    culture_level: int,
    growth_feedback: Optional[str],
    consent_version: str,
    source: str,
    page_url: Optional[str],
    session_id: Optional[str],
    metadata: Optional[dict] = None,
) -> dict:
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    subject_pseudonym = employee_assessment_pseudonym(employee_identifier)
    clean_department = clean_feedback_text(department, MAX_EMPLOYEE_DEPARTMENT_LENGTH)
    clean_feedback = redact_sensitive_text(growth_feedback, MAX_EMPLOYEE_ASSESSMENT_FEEDBACK_LENGTH)
    clean_consent_version = clean_feedback_text(consent_version, 80) or EMPLOYEE_ASSESSMENT_CONSENT_VERSION
    clean_source = clean_feedback_text(source, 80) or "employee_assessment_form"
    clean_page_url = clean_feedback_text(page_url, 500)
    clean_session_id = clean_feedback_text(session_id, 120)
    deletion_due_at = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(
        days=EMPLOYEE_ASSESSMENT_RETENTION_DAYS
    )
    metadata_payload = {
        "api_version": "2026-06-24",
        "wbs_task": "T840",
        "assessment_scope": "work_support_self_report",
        "raw_identifier_stored": False,
        "sensitive_text_redacted": True,
        "pseudonym_salt_configured": bool(os.environ.get("EMPLOYEE_ASSESSMENT_PSEUDONYM_SALT")),
        **(metadata or {}),
    }
    metadata_json = json.dumps(metadata_payload, ensure_ascii=False)
    try:
        if db_type == "postgres":
            cursor.execute(
                """
                INSERT INTO employee_assessment_responses
                    (subject_pseudonym, department_bucket, motivation_level, culture_level,
                     growth_support_excerpt, consent_version, source, page_url, session_id,
                     metadata, deletion_due_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                RETURNING id;
                """,
                (
                    subject_pseudonym,
                    clean_department,
                    motivation_level,
                    culture_level,
                    clean_feedback,
                    clean_consent_version,
                    clean_source,
                    clean_page_url,
                    clean_session_id,
                    metadata_json,
                    deletion_due_at,
                ),
            )
            inserted_id = int(cursor.fetchone()[0])
        else:
            cursor.execute(
                """
                INSERT INTO employee_assessment_responses
                    (subject_pseudonym, department_bucket, motivation_level, culture_level,
                     growth_support_excerpt, consent_version, source, page_url, session_id,
                     metadata, deletion_due_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                """,
                (
                    subject_pseudonym,
                    clean_department,
                    motivation_level,
                    culture_level,
                    clean_feedback,
                    clean_consent_version,
                    clean_source,
                    clean_page_url,
                    clean_session_id,
                    metadata_json,
                    deletion_due_at.isoformat(timespec="seconds"),
                ),
            )
            inserted_id = int(cursor.lastrowid)
        conn.commit()
        return {
            "id": inserted_id,
            "subject_pseudonym": subject_pseudonym,
            "deletion_due_at": deletion_due_at.isoformat(timespec="seconds"),
            "stored_feedback_excerpt": clean_feedback,
        }
    except Exception as e:
        record_storage_failure("insert_employee_assessment_response", e)
        return {"id": 0, "subject_pseudonym": subject_pseudonym, "error": str(e)}
    finally:
        cursor.close()
        conn.close()


def db_get_employee_assessment_summary(limit: int = 20) -> dict:
    limit = max(1, min(limit, 100))
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM employee_assessment_responses;")
        total = int(_scalar(cursor.fetchone()) or 0)

        cursor.execute("SELECT AVG(motivation_level), AVG(culture_level) FROM employee_assessment_responses;")
        avg_row = cursor.fetchone()
        motivation_average = float(avg_row[0]) if avg_row and avg_row[0] is not None else None
        culture_average = float(avg_row[1]) if avg_row and avg_row[1] is not None else None

        cursor.execute("SELECT department_bucket, COUNT(*) FROM employee_assessment_responses GROUP BY department_bucket;")
        department_counts = {}
        for row in cursor.fetchall():
            department_counts[str(row[0])] = int(row[1])

        columns = [
            "id",
            "subject_pseudonym",
            "department_bucket",
            "motivation_level",
            "culture_level",
            "growth_support_excerpt",
            "consent_version",
            "status",
            "created_at",
            "deletion_due_at",
        ]
        if db_type == "postgres":
            cursor.execute(
                """
                SELECT id, subject_pseudonym, department_bucket, motivation_level, culture_level,
                       growth_support_excerpt, consent_version, status, created_at, deletion_due_at
                FROM employee_assessment_responses
                ORDER BY id DESC
                LIMIT %s;
                """,
                (limit,),
            )
        else:
            cursor.execute(
                """
                SELECT id, subject_pseudonym, department_bucket, motivation_level, culture_level,
                       growth_support_excerpt, consent_version, status, created_at, deletion_due_at
                FROM employee_assessment_responses
                ORDER BY id DESC
                LIMIT ?;
                """,
                (limit,),
            )

        recent = []
        for row in cursor.fetchall():
            row_dict = dict(row) if isinstance(row, sqlite3.Row) else dict(zip(columns, row))
            recent.append({
                "id": row_dict.get("id"),
                "subject_pseudonym": row_dict.get("subject_pseudonym"),
                "department_bucket": row_dict.get("department_bucket"),
                "motivation_level": row_dict.get("motivation_level"),
                "culture_level": row_dict.get("culture_level"),
                "growth_support_excerpt": clean_feedback_text(row_dict.get("growth_support_excerpt"), 180),
                "consent_version": row_dict.get("consent_version"),
                "status": row_dict.get("status"),
                "created_at": str(row_dict.get("created_at") or ""),
                "deletion_due_at": str(row_dict.get("deletion_due_at") or ""),
            })

        return {
            "total": total,
            "averages": {
                "motivation_level": round(motivation_average, 2) if motivation_average is not None else None,
                "culture_level": round(culture_average, 2) if culture_average is not None else None,
            },
            "department_counts": department_counts,
            "recent": recent,
            "privacy_controls": {
                "raw_identifier_stored": False,
                "sensitive_text_redacted": True,
                "admin_summary_requires_basic_auth": True,
            },
        }
    except Exception as e:
        print(f"[-] Database employee assessment summary failed: {e}")
        return {
            "total": 0,
            "averages": {"motivation_level": None, "culture_level": None},
            "department_counts": {},
            "recent": [],
            "privacy_controls": {
                "raw_identifier_stored": False,
                "sensitive_text_redacted": True,
                "admin_summary_requires_basic_auth": True,
            },
            "error": str(e),
        }
    finally:
        cursor.close()
        conn.close()


def onboarding_pseudonym(account_identifier: str) -> str:
    """Stable, non-reversible subject id for onboarding audit events (T752)."""
    clean_identifier = clean_feedback_text(
        account_identifier, MAX_ONBOARDING_IDENTIFIER_LENGTH
    ).lower()
    return f"onb-{stable_digest(f'{ONBOARDING_PSEUDONYM_SALT}:{clean_identifier}')}"


def split_onboarding_step_ids(step_ids: Optional[List[str]]) -> tuple[list[str], list[str]]:
    """Partition submitted step ids into known (deduped, canonical order) and ignored.

    Unknown ids are reported rather than silently dropped so a wizard that drifts
    from the server catalogue is visible instead of quietly under-reporting.
    """
    submitted = [clean_feedback_text(s, 64) for s in (step_ids or [])]
    known = [step["id"] for step in ONBOARDING_STEPS if step["id"] in submitted]
    ignored = sorted({s for s in submitted if s and s not in ONBOARDING_STEP_IDS})
    return known, ignored


def build_onboarding_progress(step_ids: Optional[List[str]]) -> dict:
    """Progress computed against the canonical step catalogue."""
    completed, ignored = split_onboarding_step_ids(step_ids)
    remaining_required = [s for s in ONBOARDING_REQUIRED_STEP_IDS if s not in completed]
    total = len(ONBOARDING_STEPS) or 1
    return {
        "status": "success",
        "flow_version": ONBOARDING_FLOW_VERSION,
        "completed_step_ids": completed,
        "ignored_step_ids": ignored,
        "remaining_required_step_ids": remaining_required,
        "progress_pct": round(100.0 * len(completed) / total, 1),
        "can_activate": not remaining_required,
    }


def attendance_pseudonym(employee_identifier: str) -> str:
    clean_identifier = clean_feedback_text(employee_identifier, MAX_ATTENDANCE_IDENTIFIER_LENGTH).lower()
    digest = stable_digest(f"{ATTENDANCE_PSEUDONYM_SALT}:{clean_identifier}")
    return f"att-{digest}"


def normalize_attendance_event_type(event_type: str) -> str:
    normalized = clean_feedback_text(event_type, 32).lower().replace(" ", "_")
    return ATTENDANCE_EVENT_TYPE_ALIASES.get(normalized, "")


def minutes_to_hours(minutes: int) -> float:
    return round(max(0, int(minutes)) / 60.0, 2)


def hours_to_minutes(value: Any) -> int:
    try:
        numeric = float(str(value).replace("時間", "").replace("h", "").replace("H", "").strip())
    except (TypeError, ValueError):
        return 0
    return max(0, int(round(numeric * 60)))


def truthy_attendance_flag(value: Any) -> bool:
    text = clean_feedback_text(str(value or ""), 40).lower()
    return text not in {"", "0", "false", "no", "none", "なし", "-", "無"}


def normalize_attendance_key(text: Any) -> str:
    """Normalize a header cell for matching: NFKC (full-width -> half-width),
    strip all whitespace/newlines, drop underscores, lowercase."""
    normalized = unicodedata.normalize("NFKC", str(text or ""))
    return "".join(normalized.split()).replace("_", "").replace("　", "").lower()


ATTENDANCE_HOURS_HEADER_PREFIXES = ("作業時間", "実労働時間", "労働時間", "勤務時間")
ATTENDANCE_HOURS_HEADER_KEYS = {"workhours", "actualhours", "workinghours", "worktime"}
ATTENDANCE_SUMMARY_ROW_KEYS = {"合計", "小計", "総計", "平均", "total", "subtotal"}


def is_attendance_header_row(values: List[Any]) -> bool:
    for value in values:
        key = normalize_attendance_key(value)
        if not key:
            continue
        if key in ATTENDANCE_HOURS_HEADER_KEYS or key.startswith(ATTENDANCE_HOURS_HEADER_PREFIXES):
            return True
    return False


def attendance_rows_from_matrix(matrix: List[List[Any]]) -> List[Dict[str, Any]]:
    """Build CSV-like row dicts from a cell matrix (T874).

    Real-world timesheets (e.g. SRA 作業報告書様式) have title/preamble rows before
    the header and summary/footer rows after the data. Detect the header row by its
    hours column, and drop 合計/小計 style summary rows so totals are not doubled.
    """
    header: List[str] = []
    header_index = -1
    first_non_empty = -1
    for index, raw_row in enumerate(matrix[:80]):
        values = ["" if value is None else value for value in raw_row]
        if not any(str(value).strip() for value in values):
            continue
        if first_non_empty < 0:
            first_non_empty = index
        if is_attendance_header_row(values):
            header = [str(value or "").strip() for value in values]
            header_index = index
            break
    if header_index < 0:
        if first_non_empty < 0:
            return []
        header_index = first_non_empty
        header = [str(value or "").strip() for value in matrix[first_non_empty]]

    rows: List[Dict[str, Any]] = []
    for raw_row in matrix[header_index + 1:]:
        values = ["" if value is None else value for value in raw_row]
        if not any(str(value).strip() for value in values):
            continue
        first_cell = next((value for value in values if str(value).strip()), "")
        if normalize_attendance_key(first_cell) in ATTENDANCE_SUMMARY_ROW_KEYS:
            continue
        rows.append({header[i]: values[i] for i in range(min(len(header), len(values)))})
    return rows


def row_value(row: Dict[str, Any], candidates: List[str]) -> Any:
    normalized = {normalize_attendance_key(key): value for key, value in row.items()}
    for candidate in candidates:
        key = normalize_attendance_key(candidate)
        if key in normalized:
            return normalized[key]
        for row_key, value in normalized.items():
            if key and row_key.startswith(key):
                return value
    return ""


def parse_attendance_csv_bytes(raw_bytes: bytes) -> dict:
    text = raw_bytes.decode("utf-8-sig", errors="ignore")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    matrix = [list(record) for record in csv.reader(io.StringIO(text))]
    rows = attendance_rows_from_matrix(matrix)
    if not rows:
        raise ValueError("CSV header and at least one data row are required")
    return aggregate_attendance_rows(rows)


def parse_attendance_xlsx_bytes(raw_bytes: bytes) -> dict:
    """Convert the first worksheet of an .xlsx timesheet into CSV-like rows (T873).

    The workbook is read in-memory only; the raw file is never persisted.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel support requires openpyxl on the server") from exc
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Excel file could not be read; save it as .xlsx and retry") from exc
    try:
        worksheet = workbook.worksheets[0]
        matrix = [list(excel_row) for excel_row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    rows = attendance_rows_from_matrix(matrix)
    if not rows:
        raise ValueError("Excel sheet needs a header row and at least one data row")
    return aggregate_attendance_rows(rows)


def parse_attendance_xls_bytes(raw_bytes: bytes) -> dict:
    """Convert a legacy .xls timesheet via xlrd into CSV-like rows (T874)."""
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("legacy Excel support requires xlrd on the server") from exc
    try:
        workbook = xlrd.open_workbook(file_contents=raw_bytes)
        worksheet = workbook.sheet_by_index(0)
        matrix = [worksheet.row_values(index) for index in range(worksheet.nrows)]
    except Exception as exc:
        raise ValueError("Excel(.xls) file could not be read; save it as .xlsx or CSV and retry") from exc
    rows = attendance_rows_from_matrix(matrix)
    if not rows:
        raise ValueError("Excel sheet needs a header row and at least one data row")
    return aggregate_attendance_rows(rows)


def aggregate_attendance_rows(rows: List[Dict[str, Any]]) -> dict:
    work_minutes = 0
    overtime_minutes = 0
    midnight_minutes = 0
    holiday_work_days = 0
    anomaly_count = 0
    parsed_rows = 0

    for row in rows:
        if not any(str(value or "").strip() for value in row.values()):
            continue
        parsed_rows += 1
        work_minutes += hours_to_minutes(row_value(row, ["work_hours", "actual_hours", "working_hours", "実労働時間", "労働時間", "勤務時間", "作業時間"]))
        overtime_minutes += hours_to_minutes(row_value(row, ["overtime_hours", "overtime", "残業時間", "時間外労働", "時間外労働時間"]))
        midnight_minutes += hours_to_minutes(row_value(row, ["midnight_hours", "late_night_hours", "深夜労働", "深夜労働時間"]))
        if truthy_attendance_flag(row_value(row, ["holiday_work", "holiday_worked", "休日出勤", "休日出勤日数"])):
            raw_holiday = row_value(row, ["holiday_work", "holiday_worked", "休日出勤", "休日出勤日数"])
            try:
                holiday_work_days += max(0, int(float(str(raw_holiday))))
            except (TypeError, ValueError):
                holiday_work_days += 1
        anomaly_text = clean_feedback_text(str(row_value(row, ["anomaly", "anomalies", "打刻漏れ", "異常", "異常値判定"]) or ""), 120)
        if truthy_attendance_flag(anomaly_text):
            anomaly_count += 1

    if parsed_rows == 0:
        raise ValueError("timesheet has no usable rows")

    return {
        "work_minutes": work_minutes,
        "overtime_minutes": overtime_minutes,
        "holiday_work_days": holiday_work_days,
        "midnight_minutes": midnight_minutes,
        "anomaly_count": anomaly_count,
        "parsed_rows": parsed_rows,
        "parser": "csv_header_sum_v1",
    }


def attendance_summary_payload(row: dict) -> dict:
    return {
        "work_hours": minutes_to_hours(int(row.get("work_minutes") or 0)),
        "overtime_hours": minutes_to_hours(int(row.get("overtime_minutes") or 0)),
        "holiday_work_days": int(row.get("holiday_work_days") or 0),
        "midnight_hours": minutes_to_hours(int(row.get("midnight_minutes") or 0)),
        "anomaly_count": int(row.get("anomaly_count") or 0),
    }


def db_insert_attendance_punch(
    employee_identifier: str,
    event_type: str,
    source: str,
    page_url: Optional[str],
    session_id: Optional[str],
    metadata: Optional[dict] = None,
) -> dict:
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    subject_pseudonym = attendance_pseudonym(employee_identifier)
    normalized_event_type = normalize_attendance_event_type(event_type)
    recorded_at = datetime.datetime.now(datetime.timezone.utc)
    metadata_json = json.dumps({
        "api_version": "2026-06-24",
        "wbs_task": "T841",
        "raw_identifier_stored": False,
        **(metadata or {}),
    }, ensure_ascii=False)
    try:
        if db_type == "postgres":
            cursor.execute(
                """
                INSERT INTO attendance_punch_events
                    (subject_pseudonym, event_type, recorded_at, source, page_url, session_id, metadata)
                VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb)
                RETURNING id;
                """,
                (
                    subject_pseudonym,
                    normalized_event_type,
                    recorded_at,
                    clean_feedback_text(source, 80) or "attendance_widget",
                    clean_feedback_text(page_url, 500),
                    clean_feedback_text(session_id, 120),
                    metadata_json,
                ),
            )
            inserted_id = int(cursor.fetchone()[0])
        else:
            cursor.execute(
                """
                INSERT INTO attendance_punch_events
                    (subject_pseudonym, event_type, recorded_at, source, page_url, session_id, metadata)
                VALUES (?, ?, ?, ?, ?, ?, ?);
                """,
                (
                    subject_pseudonym,
                    normalized_event_type,
                    recorded_at.isoformat(timespec="seconds"),
                    clean_feedback_text(source, 80) or "attendance_widget",
                    clean_feedback_text(page_url, 500),
                    clean_feedback_text(session_id, 120),
                    metadata_json,
                ),
            )
            inserted_id = int(cursor.lastrowid)
        conn.commit()
        return {
            "id": inserted_id,
            "subject_pseudonym": subject_pseudonym,
            "event_type": normalized_event_type,
            "recorded_at": recorded_at.isoformat(timespec="seconds"),
        }
    except Exception as e:
        record_storage_failure("insert_attendance_punch", e)
        return {"id": 0, "subject_pseudonym": subject_pseudonym, "error": str(e)}
    finally:
        cursor.close()
        conn.close()


def db_insert_attendance_timesheet_import(
    employee_identifier: str,
    file_name: str,
    file_bytes: bytes,
    parse_result: dict,
    consent_version: str,
    source: str,
    page_url: Optional[str],
    session_id: Optional[str],
) -> dict:
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    subject_pseudonym = attendance_pseudonym(employee_identifier)
    file_extension = Path(file_name or "").suffix.lower().lstrip(".")[:16] or "csv"
    file_digest = stable_digest(file_bytes.hex())
    metadata_json = json.dumps({
        "api_version": "2026-06-24",
        "wbs_task": "T841",
        "parser": parse_result.get("parser"),
        "parsed_rows": parse_result.get("parsed_rows", 0),
        "raw_file_stored": False,
        "original_filename_stored": False,
        "jobcan_equivalent_status": "approval_log_ready",
    }, ensure_ascii=False)
    values = (
        subject_pseudonym,
        file_digest,
        file_extension,
        int(parse_result.get("work_minutes") or 0),
        int(parse_result.get("overtime_minutes") or 0),
        int(parse_result.get("holiday_work_days") or 0),
        int(parse_result.get("midnight_minutes") or 0),
        int(parse_result.get("anomaly_count") or 0),
        clean_feedback_text(consent_version, 80) or ATTENDANCE_CONSENT_VERSION,
        clean_feedback_text(source, 80) or "attendance_timesheet_upload",
        clean_feedback_text(page_url, 500),
        clean_feedback_text(session_id, 120),
        metadata_json,
    )
    try:
        if db_type == "postgres":
            cursor.execute(
                """
                INSERT INTO attendance_timesheet_imports
                    (subject_pseudonym, file_digest, file_extension, work_minutes, overtime_minutes,
                     holiday_work_days, midnight_minutes, anomaly_count, consent_version, source,
                     page_url, session_id, metadata)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                RETURNING id;
                """,
                values,
            )
            inserted_id = int(cursor.fetchone()[0])
        else:
            cursor.execute(
                """
                INSERT INTO attendance_timesheet_imports
                    (subject_pseudonym, file_digest, file_extension, work_minutes, overtime_minutes,
                     holiday_work_days, midnight_minutes, anomaly_count, consent_version, source,
                     page_url, session_id, metadata)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                """,
                values,
            )
            inserted_id = int(cursor.lastrowid)
        conn.commit()
        return {
            "id": inserted_id,
            "subject_pseudonym": subject_pseudonym,
            "status": "pending_approval",
            "summary": attendance_summary_payload(parse_result),
        }
    except Exception as e:
        record_storage_failure("insert_attendance_timesheet", e)
        return {"id": 0, "subject_pseudonym": subject_pseudonym, "error": str(e)}
    finally:
        cursor.close()
        conn.close()


def db_update_attendance_timesheet_decision(import_id: int, decision: str) -> dict:
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    approved_at = datetime.datetime.now(datetime.timezone.utc)
    try:
        if db_type == "postgres":
            cursor.execute(
                """
                UPDATE attendance_timesheet_imports
                SET status = %s, approved_at = %s
                WHERE id = %s
                RETURNING id, subject_pseudonym, work_minutes, overtime_minutes,
                          holiday_work_days, midnight_minutes, anomaly_count, status, approved_at;
                """,
                (decision, approved_at, import_id),
            )
        else:
            cursor.execute(
                """
                UPDATE attendance_timesheet_imports
                SET status = ?, approved_at = ?
                WHERE id = ?;
                """,
                (decision, approved_at.isoformat(timespec="seconds"), import_id),
            )
            cursor.execute(
                """
                SELECT id, subject_pseudonym, work_minutes, overtime_minutes,
                       holiday_work_days, midnight_minutes, anomaly_count, status, approved_at
                FROM attendance_timesheet_imports
                WHERE id = ?;
                """,
                (import_id,),
            )
        row = cursor.fetchone()
        if not row:
            return {"id": 0, "error": "not_found"}
        columns = [
            "id",
            "subject_pseudonym",
            "work_minutes",
            "overtime_minutes",
            "holiday_work_days",
            "midnight_minutes",
            "anomaly_count",
            "status",
            "approved_at",
        ]
        row_dict = dict(row) if isinstance(row, sqlite3.Row) else dict(zip(columns, row))
        conn.commit()
        return {
            "id": int(row_dict["id"]),
            "subject_pseudonym": row_dict["subject_pseudonym"],
            "status": row_dict["status"],
            "approved_at": str(row_dict.get("approved_at") or ""),
            "summary": attendance_summary_payload(row_dict),
        }
    except Exception as e:
        print(f"[-] Database update attendance decision failed: {e}")
        return {"id": 0, "error": str(e)}
    finally:
        cursor.close()
        conn.close()


def db_get_attendance_summary(limit: int = 20) -> dict:
    limit = max(1, min(limit, 100))
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM attendance_punch_events;")
        punch_total = int(_scalar(cursor.fetchone()) or 0)
        cursor.execute("SELECT COUNT(*) FROM attendance_timesheet_imports;")
        import_total = int(_scalar(cursor.fetchone()) or 0)
        cursor.execute("SELECT status, COUNT(*) FROM attendance_timesheet_imports GROUP BY status;")
        status_counts = {"pending_approval": 0, "approved": 0, "rejected": 0, "manual_review": 0}
        for row in cursor.fetchall():
            status_counts[str(row[0])] = int(row[1])
        cursor.execute("SELECT AVG(overtime_minutes), AVG(work_minutes), SUM(anomaly_count) FROM attendance_timesheet_imports WHERE status = 'approved';")
        aggregate_row = cursor.fetchone()
        avg_overtime = float(aggregate_row[0]) if aggregate_row and aggregate_row[0] is not None else None
        avg_work = float(aggregate_row[1]) if aggregate_row and aggregate_row[1] is not None else None
        anomaly_sum = int(aggregate_row[2]) if aggregate_row and aggregate_row[2] is not None else 0

        columns = [
            "id",
            "subject_pseudonym",
            "work_minutes",
            "overtime_minutes",
            "holiday_work_days",
            "midnight_minutes",
            "anomaly_count",
            "status",
            "created_at",
            "approved_at",
        ]
        if db_type == "postgres":
            cursor.execute(
                """
                SELECT id, subject_pseudonym, work_minutes, overtime_minutes, holiday_work_days,
                       midnight_minutes, anomaly_count, status, created_at, approved_at
                FROM attendance_timesheet_imports
                ORDER BY id DESC
                LIMIT %s;
                """,
                (limit,),
            )
        else:
            cursor.execute(
                """
                SELECT id, subject_pseudonym, work_minutes, overtime_minutes, holiday_work_days,
                       midnight_minutes, anomaly_count, status, created_at, approved_at
                FROM attendance_timesheet_imports
                ORDER BY id DESC
                LIMIT ?;
                """,
                (limit,),
            )
        recent_imports = []
        for row in cursor.fetchall():
            row_dict = dict(row) if isinstance(row, sqlite3.Row) else dict(zip(columns, row))
            recent_imports.append({
                "id": row_dict.get("id"),
                "subject_pseudonym": row_dict.get("subject_pseudonym"),
                "status": row_dict.get("status"),
                "summary": attendance_summary_payload(row_dict),
                "created_at": str(row_dict.get("created_at") or ""),
                "approved_at": str(row_dict.get("approved_at") or ""),
            })

        return {
            "punch_total": punch_total,
            "import_total": import_total,
            "status_counts": status_counts,
            "approved_averages": {
                "overtime_hours": minutes_to_hours(avg_overtime or 0) if avg_overtime is not None else None,
                "work_hours": minutes_to_hours(avg_work or 0) if avg_work is not None else None,
            },
            "approved_anomaly_count": anomaly_sum,
            "recent_imports": recent_imports,
            "privacy_controls": {
                "raw_identifier_stored": False,
                "raw_file_stored": False,
                "original_filename_stored": False,
                "admin_summary_requires_basic_auth": True,
            },
        }
    except Exception as e:
        print(f"[-] Database attendance summary failed: {e}")
        return {
            "punch_total": 0,
            "import_total": 0,
            "status_counts": {"pending_approval": 0, "approved": 0, "rejected": 0, "manual_review": 0},
            "approved_averages": {"overtime_hours": None, "work_hours": None},
            "approved_anomaly_count": 0,
            "recent_imports": [],
            "privacy_controls": {
                "raw_identifier_stored": False,
                "raw_file_stored": False,
                "original_filename_stored": False,
                "admin_summary_requires_basic_auth": True,
            },
            "error": str(e),
        }
    finally:
        cursor.close()
        conn.close()


def db_json_value(value: Any) -> str:
    return json.dumps(value if value is not None else {}, ensure_ascii=False)


def db_json_array(value: Any) -> str:
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False)
    return json.dumps([], ensure_ascii=False)


def sales_email_db_direction(direction: str) -> str:
    return "engineer_to_project" if direction == "talent_to_project" else "project_to_talent"


def sales_email_match_review_status(feedback_status: str) -> str:
    return "pending" if feedback_status == "needs_review" else feedback_status


def db_get_or_create_sales_project(cursor: Any, db_type: str, project: dict, project_key: str) -> int:
    if db_type == "postgres":
        cursor.execute("SELECT id FROM project_requirements WHERE metadata->>'project_key' = %s LIMIT 1;", (project_key,))
    else:
        cursor.execute("SELECT id FROM project_requirements WHERE metadata LIKE ? LIMIT 1;", (f"%{project_key}%",))
    row = cursor.fetchone()
    if row:
        return int(row[0])

    metadata = {"project_key": project_key, "source": "sales_email_human_review", "wbs_task": "T817_6"}
    values = (
        clean_feedback_text(project.get("title") or "Sales email project", 255),
        clean_feedback_text(project.get("summary") or "", 1000),
        db_json_array(project.get("required_skills")),
        db_json_array(project.get("nice_to_have_skills")),
        db_json_value(project.get("skill_categories") if isinstance(project.get("skill_categories"), dict) else {}),
        project.get("rate_min") if isinstance(project.get("rate_min"), int) else None,
        project.get("rate_max") if isinstance(project.get("rate_max"), int) else None,
        clean_feedback_text(project.get("location") or "", 160),
        clean_feedback_text(project.get("remote_type") or "unknown", 32) or "unknown",
        clean_feedback_text(project.get("start_date_text") or "", 120),
        clean_feedback_text(project.get("evidence_excerpt") or "", 1000),
        db_json_value(metadata),
    )
    if db_type == "postgres":
        cursor.execute(
            """
            INSERT INTO project_requirements
                (title, summary, required_skills, nice_to_have_skills, skill_categories, rate_min, rate_max,
                 location, remote_type, start_date_text, evidence_excerpt, metadata)
            VALUES (%s, %s, %s::jsonb, %s::jsonb, %s::jsonb, %s, %s, %s, %s, %s, %s, %s::jsonb)
            RETURNING id;
            """,
            values,
        )
        return int(cursor.fetchone()[0])

    cursor.execute(
        """
        INSERT INTO project_requirements
            (title, summary, required_skills, nice_to_have_skills, skill_categories, rate_min, rate_max,
             location, remote_type, start_date_text, evidence_excerpt, metadata)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """,
        values,
    )
    return int(cursor.lastrowid)


def db_get_or_create_sales_talent(cursor: Any, db_type: str, talent: dict, talent_key: str) -> int:
    if db_type == "postgres":
        cursor.execute("SELECT id FROM talent_profiles_from_email WHERE anonymized_talent_key = %s LIMIT 1;", (talent_key,))
    else:
        cursor.execute("SELECT id FROM talent_profiles_from_email WHERE anonymized_talent_key = ? LIMIT 1;", (talent_key,))
    row = cursor.fetchone()
    if row:
        return int(row[0])

    metadata = {"talent_key": talent_key, "source": "sales_email_human_review", "wbs_task": "T817_6"}
    values = (
        clean_feedback_text(talent_key or "anonymous_talent", 120),
        clean_feedback_text(talent.get("summary") or "", 1000),
        db_json_array(talent.get("skills")),
        db_json_value(talent.get("skill_categories") if isinstance(talent.get("skill_categories"), dict) else {}),
        talent.get("desired_rate_min") if isinstance(talent.get("desired_rate_min"), int) else None,
        talent.get("desired_rate_max") if isinstance(talent.get("desired_rate_max"), int) else None,
        clean_feedback_text(talent.get("desired_location") or "", 160),
        clean_feedback_text(talent.get("remote_preference") or "unknown", 32) or "unknown",
        clean_feedback_text(talent.get("availability_text") or "", 160),
        clean_feedback_text(talent.get("evidence_excerpt") or "", 1000),
        db_json_value(metadata),
    )
    if db_type == "postgres":
        cursor.execute(
            """
            INSERT INTO talent_profiles_from_email
                (anonymized_talent_key, summary, skills, skill_categories, desired_rate_min, desired_rate_max,
                 desired_location, remote_preference, availability_text, evidence_excerpt, metadata)
            VALUES (%s, %s, %s::jsonb, %s::jsonb, %s, %s, %s, %s, %s, %s, %s::jsonb)
            RETURNING id;
            """,
            values,
        )
        return int(cursor.fetchone()[0])

    cursor.execute(
        """
        INSERT INTO talent_profiles_from_email
            (anonymized_talent_key, summary, skills, skill_categories, desired_rate_min, desired_rate_max,
             desired_location, remote_preference, availability_text, evidence_excerpt, metadata)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """,
        values,
    )
    return int(cursor.lastrowid)


def db_get_or_create_sales_email_match_result(
    cursor: Any,
    db_type: str,
    match_row: dict,
    project_id: int,
    talent_id: int,
    match_key_value: str,
    report_direction: str,
) -> int:
    if db_type == "postgres":
        cursor.execute("SELECT id FROM email_match_results WHERE metadata->>'match_key' = %s LIMIT 1;", (match_key_value,))
    else:
        cursor.execute("SELECT id FROM email_match_results WHERE metadata LIKE ? LIMIT 1;", (f"%{match_key_value}%",))
    row = cursor.fetchone()
    if row:
        return int(row[0])

    metadata = {
        "match_key": match_key_value,
        "match_reason": match_row.get("match_reason"),
        "matched_conditions": match_row.get("matched_conditions", []),
        "score_breakdown": match_row.get("score_breakdown", {}),
        "report_direction": report_direction,
        "source": "sales_email_human_review",
        "wbs_task": "T817_6",
    }
    values = (
        project_id,
        talent_id,
        sales_email_db_direction(report_direction),
        float(match_row.get("score") or 0),
        db_json_array(match_row.get("matched_skills")),
        db_json_array(match_row.get("missing_skills")),
        db_json_array(match_row.get("mismatch_reasons")),
        clean_feedback_text(match_row.get("match_reason") or "", 1000),
        db_json_value(metadata),
    )
    if db_type == "postgres":
        cursor.execute(
            """
            INSERT INTO email_match_results
                (project_requirement_id, talent_profile_id, direction, match_score, matched_skills,
                 missing_skills, mismatch_reasons, evidence_summary, metadata)
            VALUES (%s, %s, %s, %s, %s::jsonb, %s::jsonb, %s::jsonb, %s, %s::jsonb)
            RETURNING id;
            """,
            values,
        )
        return int(cursor.fetchone()[0])

    cursor.execute(
        """
        INSERT INTO email_match_results
            (project_requirement_id, talent_profile_id, direction, match_score, matched_skills,
             missing_skills, mismatch_reasons, evidence_summary, metadata)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
        """,
        values,
    )
    return int(cursor.lastrowid)


def db_insert_sales_email_match_review(
    match_row: dict,
    project: dict,
    talent: dict,
    review_entry: dict,
    report_direction: str,
) -> dict:
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        project_id = db_get_or_create_sales_project(cursor, db_type, project, review_entry["project_key"])
        talent_id = db_get_or_create_sales_talent(cursor, db_type, talent, review_entry["talent_key"])
        match_result_id = db_get_or_create_sales_email_match_result(
            cursor,
            db_type,
            match_row,
            project_id,
            talent_id,
            review_entry["match_key"],
            report_direction,
        )
        feedback_status = review_entry["feedback_status"]
        match_status = sales_email_match_review_status(feedback_status)
        metadata = {
            "review_id": review_entry.get("review_id"),
            "match_key": review_entry.get("match_key"),
            "corrected_fields": review_entry.get("corrected_fields", {}),
            "next_action": review_entry.get("next_action", ""),
            "privacy_controls": review_entry.get("privacy_controls", []),
            "wbs_task": "T817_6",
        }
        corrected_notes = clean_feedback_text(review_entry.get("reviewer_notes") or "", MAX_FEEDBACK_COMMENT_LENGTH)
        if db_type == "postgres":
            cursor.execute(
                """
                UPDATE email_match_results
                SET review_status = %s,
                    match_score = COALESCE(%s, match_score),
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s;
                """,
                (match_status, review_entry.get("corrected_score"), match_result_id),
            )
            cursor.execute(
                """
                INSERT INTO email_match_feedback
                    (match_result_id, reviewer_id, feedback_status, corrected_score, corrected_notes, metadata)
                VALUES (%s, %s, %s, %s, %s, %s::jsonb)
                RETURNING id;
                """,
                (
                    match_result_id,
                    review_entry.get("reviewer_id"),
                    feedback_status,
                    review_entry.get("corrected_score"),
                    corrected_notes,
                    db_json_value(metadata),
                ),
            )
            feedback_id = int(cursor.fetchone()[0])
        else:
            cursor.execute(
                """
                UPDATE email_match_results
                SET review_status = ?,
                    match_score = COALESCE(?, match_score),
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?;
                """,
                (match_status, review_entry.get("corrected_score"), match_result_id),
            )
            cursor.execute(
                """
                INSERT INTO email_match_feedback
                    (match_result_id, reviewer_id, feedback_status, corrected_score, corrected_notes, metadata)
                VALUES (?, ?, ?, ?, ?, ?);
                """,
                (
                    match_result_id,
                    review_entry.get("reviewer_id"),
                    feedback_status,
                    review_entry.get("corrected_score"),
                    corrected_notes,
                    db_json_value(metadata),
                ),
            )
            feedback_id = int(cursor.lastrowid)
        conn.commit()
        return {
            "project_requirement_id": project_id,
            "talent_profile_id": talent_id,
            "match_result_id": match_result_id,
            "feedback_id": feedback_id,
            "db_type": db_type,
        }
    except Exception as e:
        conn.rollback()
        record_storage_failure("insert_sales_email_match_review", e)
        return {"error": str(e), "db_type": db_type}
    finally:
        cursor.close()
        conn.close()


def db_get_sales_email_review_summary(limit: int = 20) -> dict:
    limit = max(1, min(limit, 100))
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM email_match_feedback;")
        total = int(_scalar(cursor.fetchone()) or 0)

        cursor.execute("SELECT feedback_status, COUNT(*) FROM email_match_feedback GROUP BY feedback_status;")
        status_counts = {status: 0 for status in sorted(SALES_EMAIL_REVIEW_STATUSES or [])}
        for row in cursor.fetchall():
            status_counts[str(row[0])] = int(row[1])

        columns = [
            "id",
            "match_result_id",
            "reviewer_id",
            "feedback_status",
            "corrected_score",
            "corrected_notes",
            "metadata",
            "created_at",
        ]
        if db_type == "postgres":
            cursor.execute(
                """
                SELECT id, match_result_id, reviewer_id, feedback_status, corrected_score, corrected_notes, metadata, created_at
                FROM email_match_feedback
                ORDER BY id DESC
                LIMIT %s;
                """,
                (limit,),
            )
        else:
            cursor.execute(
                """
                SELECT id, match_result_id, reviewer_id, feedback_status, corrected_score, corrected_notes, metadata, created_at
                FROM email_match_feedback
                ORDER BY id DESC
                LIMIT ?;
                """,
                (limit,),
            )
        recent = []
        for row in cursor.fetchall():
            row_dict = dict(row) if isinstance(row, sqlite3.Row) else dict(zip(columns, row))
            metadata = row_dict.get("metadata") or {}
            if isinstance(metadata, str):
                try:
                    metadata = json.loads(metadata)
                except json.JSONDecodeError:
                    metadata = {}
            recent.append({
                "id": row_dict.get("id"),
                "match_result_id": row_dict.get("match_result_id"),
                "reviewer_id": clean_feedback_text(row_dict.get("reviewer_id"), 120),
                "feedback_status": row_dict.get("feedback_status"),
                "corrected_score": row_dict.get("corrected_score"),
                "notes_excerpt": clean_feedback_text(row_dict.get("corrected_notes"), 160),
                "match_key": metadata.get("match_key"),
                "next_action": clean_feedback_text(metadata.get("next_action"), 160),
                "created_at": str(row_dict.get("created_at") or ""),
            })
        return {
            "total": total,
            "status_counts": status_counts,
            "recent": recent,
        }
    except Exception as e:
        print(f"[-] Database sales email review summary failed: {e}")
        return {
            "total": 0,
            "status_counts": {status: 0 for status in sorted(SALES_EMAIL_REVIEW_STATUSES or [])},
            "recent": [],
            "error": str(e),
        }
    finally:
        cursor.close()
        conn.close()


def build_operations_dashboard_summary(limit: int = 20) -> dict:
    limit = max(1, min(limit, 100))
    employee_assessment = db_get_employee_assessment_summary(limit=limit)
    attendance = db_get_attendance_summary(limit=limit)
    sales_email_review = db_get_sales_email_review_summary(limit=limit)
    usage_analytics = db_get_usage_analytics_summary(limit=limit)

    sales_status_counts = sales_email_review.get("status_counts", {}) or {}
    reviewed_sales_count = sum(
        int(sales_status_counts.get(status_name, 0) or 0)
        for status_name in ("accepted", "rejected", "corrected")
    )
    sales_total = int(sales_email_review.get("total", 0) or 0)
    sales_review_completion_rate = round((reviewed_sales_count / sales_total) * 100, 2) if sales_total else 0.0

    return {
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
        "wbs_task": "T842;T800",
        "kpis": {
            "employee_assessment_responses": int(employee_assessment.get("total", 0) or 0),
            "motivation_average": employee_assessment.get("averages", {}).get("motivation_level"),
            "culture_average": employee_assessment.get("averages", {}).get("culture_level"),
            "attendance_punch_events": int(attendance.get("punch_total", 0) or 0),
            "attendance_timesheet_imports": int(attendance.get("import_total", 0) or 0),
            "attendance_pending_approval": int(attendance.get("status_counts", {}).get("pending_approval", 0) or 0),
            "attendance_approved_overtime_average": attendance.get("approved_averages", {}).get("overtime_hours"),
            "attendance_approved_anomaly_count": int(attendance.get("approved_anomaly_count", 0) or 0),
            "sales_email_reviews": sales_total,
            "sales_email_needs_review": int(sales_status_counts.get("needs_review", 0) or 0),
            "sales_email_review_completion_rate": sales_review_completion_rate,
            "usage_analytics_events": int(usage_analytics.get("total", 0) or 0),
            "usage_events_last_7_days": int(usage_analytics.get("events_last_7_days", 0) or 0),
            "usage_unique_sessions_last_7_days": int(usage_analytics.get("unique_sessions_last_7_days", 0) or 0),
            "usage_page_views": int(usage_analytics.get("event_counts", {}).get("page_view", 0) or 0),
        },
        "sources": {
            "employee_assessment": employee_assessment,
            "attendance": attendance,
            "sales_email_review": sales_email_review,
            "usage_analytics": usage_analytics,
        },
        "security": {
            "admin_summary_requires_basic_auth": True,
            "report_export_requires_basic_auth": True,
            "raw_identifiers_excluded": True,
            "raw_attendance_files_excluded": True,
            "sales_email_body_excluded": True,
            "usage_analytics_pseudonymized_sessions": True,
            "usage_analytics_ip_address_excluded": True,
            "usage_analytics_raw_user_agent_excluded": True,
            "usage_analytics_form_contents_excluded": True,
            "secret_like_values_redacted": True,
        },
    }


def build_operations_dashboard_csv(summary: dict) -> str:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["category", "metric", "value", "source"])

    kpis = summary.get("kpis", {})
    for metric_name in sorted(kpis):
        writer.writerow(["kpi", metric_name, kpis.get(metric_name), "operations_dashboard"])

    sources = summary.get("sources", {})
    employee_assessment = sources.get("employee_assessment", {})
    attendance = sources.get("attendance", {})
    sales_email_review = sources.get("sales_email_review", {})
    usage_analytics = sources.get("usage_analytics", {})

    writer.writerow(["employee_assessment", "responses", employee_assessment.get("total", 0), "api_summary"])
    for dept_name, count in sorted((employee_assessment.get("department_counts") or {}).items()):
        writer.writerow(["employee_assessment_department", dept_name, count, "api_summary"])

    writer.writerow(["attendance", "punch_events", attendance.get("punch_total", 0), "api_summary"])
    writer.writerow(["attendance", "timesheet_imports", attendance.get("import_total", 0), "api_summary"])
    for status_name, count in sorted((attendance.get("status_counts") or {}).items()):
        writer.writerow(["attendance_status", status_name, count, "api_summary"])

    writer.writerow(["sales_email_review", "reviews", sales_email_review.get("total", 0), "api_summary"])
    for status_name, count in sorted((sales_email_review.get("status_counts") or {}).items()):
        writer.writerow(["sales_email_review_status", status_name, count, "api_summary"])

    writer.writerow(["usage_analytics", "events", usage_analytics.get("total", 0), "api_summary"])
    writer.writerow(["usage_analytics", "unique_sessions_last_7_days", usage_analytics.get("unique_sessions_last_7_days", 0), "api_summary"])
    for event_name, count in sorted((usage_analytics.get("event_counts") or {}).items()):
        writer.writerow(["usage_analytics_event", event_name, count, "api_summary"])
    for surface_name, count in sorted((usage_analytics.get("surface_counts") or {}).items()):
        writer.writerow(["usage_analytics_surface", surface_name, count, "api_summary"])

    security = summary.get("security", {})
    for flag_name in sorted(security):
        writer.writerow(["security", flag_name, security.get(flag_name), "operations_dashboard"])

    return output.getvalue()


def db_insert_support_request(
    category: str,
    priority: str,
    contact_email: str,
    subject: str,
    message: str,
    source: str,
    page_url: Optional[str],
    session_id: Optional[str],
    metadata: Optional[dict] = None,
) -> int:
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    clean_email = clean_feedback_text(contact_email, 254).lower()
    clean_subject = clean_feedback_text(subject, MAX_SUPPORT_SUBJECT_LENGTH)
    clean_message = clean_feedback_text(message, MAX_SUPPORT_MESSAGE_LENGTH)
    clean_source = clean_feedback_text(source, 80) or "support_form"
    clean_page_url = clean_feedback_text(page_url, 500)
    clean_session_id = clean_feedback_text(session_id, 120)
    metadata_json = json.dumps(metadata or {}, ensure_ascii=False)
    try:
        if db_type == "postgres":
            cursor.execute(
                """
                INSERT INTO support_requests
                    (category, priority, contact_email, subject, message, source, page_url, session_id, metadata)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                RETURNING id;
                """,
                (
                    category,
                    priority,
                    clean_email,
                    clean_subject,
                    clean_message,
                    clean_source,
                    clean_page_url,
                    clean_session_id,
                    metadata_json,
                ),
            )
            inserted_id = cursor.fetchone()[0]
        else:
            cursor.execute(
                """
                INSERT INTO support_requests
                    (category, priority, contact_email, subject, message, source, page_url, session_id, metadata)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
                """,
                (
                    category,
                    priority,
                    clean_email,
                    clean_subject,
                    clean_message,
                    clean_source,
                    clean_page_url,
                    clean_session_id,
                    metadata_json,
                ),
            )
            inserted_id = cursor.lastrowid
        conn.commit()
        return inserted_id
    except Exception as e:
        record_storage_failure("insert_support_request", e)
        return 0
    finally:
        cursor.close()
        conn.close()


def db_get_support_summary(limit: int = 20) -> dict:
    limit = max(1, min(limit, 100))
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM support_requests;")
        total = int(_scalar(cursor.fetchone()) or 0)

        cursor.execute("SELECT status, COUNT(*) FROM support_requests GROUP BY status;")
        status_counts = {status_name: 0 for status_name in sorted(SUPPORT_STATUSES)}
        for row in cursor.fetchall():
            status_counts[str(row[0])] = int(row[1])

        cursor.execute("SELECT priority, COUNT(*) FROM support_requests GROUP BY priority;")
        priority_counts = {priority_name: 0 for priority_name in sorted(VALID_SUPPORT_PRIORITIES)}
        for row in cursor.fetchall():
            priority_counts[str(row[0])] = int(row[1])

        cursor.execute("SELECT category, COUNT(*) FROM support_requests GROUP BY category;")
        category_counts = {category_name: 0 for category_name in sorted(VALID_SUPPORT_CATEGORIES)}
        for row in cursor.fetchall():
            category_counts[str(row[0])] = int(row[1])

        columns = [
            "id",
            "category",
            "priority",
            "contact_email",
            "subject",
            "message",
            "status",
            "source",
            "page_url",
            "session_id",
            "created_at",
        ]
        if db_type == "postgres":
            cursor.execute(
                """
                SELECT id, category, priority, contact_email, subject, message, status, source, page_url, session_id, created_at
                FROM support_requests
                ORDER BY id DESC
                LIMIT %s;
                """,
                (limit,),
            )
        else:
            cursor.execute(
                """
                SELECT id, category, priority, contact_email, subject, message, status, source, page_url, session_id, created_at
                FROM support_requests
                ORDER BY id DESC
                LIMIT ?;
                """,
                (limit,),
            )

        recent = []
        for row in cursor.fetchall():
            row_dict = dict(row) if isinstance(row, sqlite3.Row) else dict(zip(columns, row))
            recent.append({
                "id": row_dict.get("id"),
                "category": row_dict.get("category"),
                "priority": row_dict.get("priority"),
                "contact_email": row_dict.get("contact_email"),
                "subject": row_dict.get("subject"),
                "message_excerpt": clean_feedback_text(row_dict.get("message"), 180),
                "status": row_dict.get("status"),
                "source": row_dict.get("source"),
                "page_url": row_dict.get("page_url"),
                "session_id": row_dict.get("session_id"),
                "created_at": str(row_dict.get("created_at") or ""),
            })

        return {
            "total": total,
            "status_counts": status_counts,
            "priority_counts": priority_counts,
            "category_counts": category_counts,
            "recent": recent,
        }
    except Exception as e:
        print(f"[-] Database support summary failed: {e}")
        return {
            "total": 0,
            "status_counts": {status_name: 0 for status_name in sorted(SUPPORT_STATUSES)},
            "priority_counts": {priority_name: 0 for priority_name in sorted(VALID_SUPPORT_PRIORITIES)},
            "category_counts": {category_name: 0 for category_name in sorted(VALID_SUPPORT_CATEGORIES)},
            "recent": [],
            "error": str(e),
        }
    finally:
        cursor.close()
        conn.close()


USER_DATA_EXPORT_MAX_ROWS = 200
USER_DATA_EXPORT_JSON_FIELDS = {
    "parsed_skills",
    "career_goals",
    "parsed_requirements",
    "company_culture",
    "interview_questions",
    "metadata",
}


def db_placeholder(db_type: str) -> str:
    return "%s" if db_type == "postgres" else "?"


def json_safe_value(column: str, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    if isinstance(value, (dict, list, int, float, bool)):
        return value
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if column in USER_DATA_EXPORT_JSON_FIELDS and isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return value
    return str(value)


def db_row_to_export_dict(row: Any, columns: List[str]) -> dict:
    row_dict = dict(row) if isinstance(row, sqlite3.Row) else dict(zip(columns, row))
    return {key: json_safe_value(key, value) for key, value in row_dict.items()}


def db_fetch_by_ids(
    cursor: Any,
    db_type: str,
    table: str,
    columns: List[str],
    ids: List[int],
) -> List[dict]:
    if not ids:
        return []
    placeholder = db_placeholder(db_type)
    placeholders = ",".join([placeholder] * len(ids))
    sql = f"SELECT {', '.join(columns)} FROM {table} WHERE id IN ({placeholders}) ORDER BY id DESC;"
    cursor.execute(sql, tuple(ids))
    return [db_row_to_export_dict(row, columns) for row in cursor.fetchall()]


def unique_ints(values: List[Any]) -> List[int]:
    result = []
    seen = set()
    for value in values:
        if value is None:
            continue
        try:
            normalized = int(value)
        except (TypeError, ValueError):
            continue
        if normalized not in seen:
            result.append(normalized)
            seen.add(normalized)
    return result


def build_user_data_export(current_user: dict, session_id: str = "") -> dict:
    """Build a scoped JSON export for the authenticated user."""
    user_email = clean_feedback_text(current_user.get("email"), 254).lower()
    user_uid = clean_feedback_text(current_user.get("uid"), 120)
    clean_session_id = clean_feedback_text(session_id, 120)

    if not user_email and not clean_session_id:
        raise ValueError("user email or browser session_id is required for export scope")

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    placeholder = db_placeholder(db_type)
    try:
        support_columns = [
            "id",
            "category",
            "priority",
            "contact_email",
            "subject",
            "message",
            "status",
            "source",
            "page_url",
            "session_id",
            "metadata",
            "created_at",
            "updated_at",
        ]
        support_requests: List[dict] = []
        if user_email:
            cursor.execute(
                f"""
                SELECT {', '.join(support_columns)}
                FROM support_requests
                WHERE LOWER(contact_email) = LOWER({placeholder})
                ORDER BY id DESC
                LIMIT {placeholder};
                """,
                (user_email, USER_DATA_EXPORT_MAX_ROWS),
            )
            support_requests = [
                db_row_to_export_dict(row, support_columns) for row in cursor.fetchall()
            ]

        feedback_columns = [
            "id",
            "match_result_id",
            "rating",
            "nps_score",
            "comment",
            "source",
            "page_url",
            "session_id",
            "metadata",
            "created_at",
        ]
        feedback_events: List[dict] = []
        if clean_session_id:
            cursor.execute(
                f"""
                SELECT {', '.join(feedback_columns)}
                FROM feedback_events
                WHERE session_id = {placeholder}
                ORDER BY id DESC
                LIMIT {placeholder};
                """,
                (clean_session_id, USER_DATA_EXPORT_MAX_ROWS),
            )
            feedback_events = [
                db_row_to_export_dict(row, feedback_columns) for row in cursor.fetchall()
            ]

        match_columns = [
            "id",
            "engineer_id",
            "job_id",
            "fit_ratio",
            "score_skill",
            "score_culture",
            "score_growth",
            "score_performing",
            "match_summary",
            "interview_questions",
            "analyzed_at",
        ]
        match_ids = unique_ints([item.get("match_result_id") for item in feedback_events])
        match_results = db_fetch_by_ids(cursor, db_type, "match_results", match_columns, match_ids)

        engineer_columns = [
            "id",
            "name",
            "resume_raw",
            "parsed_skills",
            "career_goals",
            "created_at",
        ]
        engineer_ids = unique_ints([item.get("engineer_id") for item in match_results])
        engineers = db_fetch_by_ids(cursor, db_type, "engineers", engineer_columns, engineer_ids)

        job_columns = [
            "id",
            "title",
            "company",
            "job_description",
            "parsed_requirements",
            "company_culture",
            "created_at",
        ]
        job_ids = unique_ints([item.get("job_id") for item in match_results])
        jobs = db_fetch_by_ids(cursor, db_type, "jobs", job_columns, job_ids)

        records = {
            "support_requests": support_requests,
            "feedback_events": feedback_events,
            "match_results": match_results,
            "engineers": engineers,
            "jobs": jobs,
        }

        return {
            "status": "success",
            "schema_version": "2026-06-20.T781",
            "generated_at": (
                datetime.datetime.now(datetime.timezone.utc)
                .replace(microsecond=0)
                .isoformat()
                .replace("+00:00", "Z")
            ),
            "user": {
                "uid": user_uid,
                "email": user_email,
            },
            "scope": {
                "support_email": user_email or None,
                "browser_session_id": clean_session_id or None,
                "match_source": "feedback_events.match_result_id for the supplied browser session",
                "max_rows_per_collection": USER_DATA_EXPORT_MAX_ROWS,
            },
            "record_counts": {name: len(value) for name, value in records.items()},
            "records": records,
            "ownership_gaps": [
                "engineers/jobs/match_results created before owner_uid support are exported only when tied to the supplied feedback session_id.",
                "T752 onboarding must add stable owner_uid columns before public paid launch self-service export.",
            ],
        }
    finally:
        cursor.close()
        conn.close()


def resolve_or_insert_engineer(content: str) -> int:
    profile = build_profile(content, "engineer")
    name = profile.title
    if SUPABASE_SDK_ACTIVE:
        try:
            client = get_supabase_client()
            if client:
                res = client.table("engineers").select("id").eq("name", name).limit(1).execute()
                if res and res.data:
                    return res.data[0].get("id", 0)
        except Exception as e:
            print(f"[-] Supabase SDK query in resolve_or_insert_engineer failed: {e}")

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        if db_type == "postgres":
            cursor.execute("SELECT id FROM engineers WHERE name = %s LIMIT 1;", (name,))
        else:
            cursor.execute("SELECT id FROM engineers WHERE name = ? LIMIT 1;", (name,))
        row = cursor.fetchone()
        if row:
            return row[0] if db_type == "postgres" else row["id"]
    except Exception as e:
        print(f"[-] Database query in resolve_or_insert_engineer failed: {e}")
    finally:
        cursor.close()
        conn.close()
    return db_insert_engineer(name, content, profile.skills_by_category, {"strengths": profile.strengths, "risk_flags": profile.risk_flags})

def resolve_or_insert_job(content: str) -> int:
    profile = build_profile(content, "job")
    title = profile.title
    if SUPABASE_SDK_ACTIVE:
        try:
            client = get_supabase_client()
            if client:
                res = client.table("jobs").select("id").eq("title", title).limit(1).execute()
                if res and res.data:
                    return res.data[0].get("id", 0)
        except Exception as e:
            print(f"[-] Supabase SDK query in resolve_or_insert_job failed: {e}")

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        if db_type == "postgres":
            cursor.execute("SELECT id FROM jobs WHERE title = %s LIMIT 1;", (title,))
        else:
            cursor.execute("SELECT id FROM jobs WHERE title = ? LIMIT 1;", (title,))
        row = cursor.fetchone()
        if row:
            return row[0] if db_type == "postgres" else row["id"]
    except Exception as e:
        print(f"[-] Database query in resolve_or_insert_job failed: {e}")
    finally:
        cursor.close()
        conn.close()
    return db_insert_job(title, "Mighty-Link", content, {"mandatory": profile.all_skills, "preferred": []}, {"summary": profile.summary})


# Custom StaticFiles subclass to enforce Basic Authentication
class BasicAuthStaticFiles(StaticFiles):
    def __init__(self, *args, username: str = "admin", password: str = "mighty-link-pass", **kwargs):
        super().__init__(*args, **kwargs)
        self.username = username
        self.password = password

    async def __call__(self, scope, receive, send):
        request = Request(scope, receive)
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            await self._unauthorized(send)
            return

        try:
            auth_type, credentials = auth_header.split(" ")
            if auth_type.lower() != "basic":
                await self._unauthorized(send)
                return
            
            decoded = base64.b64decode(credentials).decode("utf-8")
            username, password = decoded.split(":", 1)
            correct_username = secrets.compare_digest(username, self.username)
            correct_password = secrets.compare_digest(password, self.password)
            if not (correct_username and correct_password):
                await self._unauthorized(send)
                return
        except Exception:
            await self._unauthorized(send)
            return

        await super().__call__(scope, receive, send)

    async def _unauthorized(self, send):
        response_headers = [
            (b"content-type", b"text/plain; charset=utf-8"),
            (b"www-authenticate", b'Basic realm="Mighty-Link Demo exports"'),
        ]
        await send({
            "type": "http.response.start",
            "status": 401,
            "headers": response_headers,
        })
        await send({
            "type": "http.response.body",
            "body": b"Unauthorized",
        })


os.makedirs(EXPORTS_DIR, exist_ok=True)
app.mount("/exports", BasicAuthStaticFiles(directory=EXPORTS_DIR, username=BASIC_AUTH_USERNAME, password=BASIC_AUTH_PASSWORD), name="exports")


def deterministic_uuid4(seed: str) -> str:
    digest = bytearray(hashlib.sha256(seed.encode("utf-8")).digest()[:16])
    digest[6] = (digest[6] & 0x0F) | 0x40
    digest[8] = (digest[8] & 0x3F) | 0x80
    return str(uuid.UUID(bytes=bytes(digest)))


DEVTOOLS_WORKSPACE_UUID = deterministic_uuid4(os.path.normcase(os.path.abspath(PROJECT_ROOT)))


# Mighty-Link Color Palette (Normalized for Sheets API)
COLORS = {
    "header_bg": {"red": 26/255, "green": 115/255, "blue": 232/255},   # #1A73E8 (Mighty Blue)
    "header_text": {"red": 1.0, "green": 1.0, "blue": 1.0},            # White
    "accent_green": {"red": 52/255, "green": 168/255, "blue": 83/255},  # #34A853 (Mighty Green)
    "row_even": {"red": 248/255, "green": 250/255, "blue": 252/255},    # Slate 50
    "border_gray": {"red": 226/255, "green": 232/255, "blue": 240/255}  # Slate 200
}

SKILL_TAXONOMY = {
    "backend": [
        "python", "fastapi", "django", "flask", "node.js", "node", "express",
        "rest api", "api", "graphql", "java", "spring", "go", "golang"
    ],
    "frontend": [
        "javascript", "typescript", "react", "react.js", "next.js", "vue",
        "html", "css", "tailwind", "chart.js", "ui", "ux"
    ],
    "ai": [
        "gemini", "openai", "llm", "rag", "prompt", "vertex ai", "生成ai",
        "ai", "マルチモーダル", "エージェント", "自律"
    ],
    "google_workspace": [
        "google sheets", "sheets api", "google drive", "drive api",
        "google calendar", "calendar api", "gspread", "oauth", "docs api",
        "workspace", "スプレッドシート", "カレンダー"
    ],
    "cloud": [
        "google cloud", "gcp", "aws", "azure", "cloud run", "docker",
        "github actions", "ci/cd", "vertex ai"
    ],
    "database": [
        "postgresql", "sqlite", "mysql", "sql", "pinecone", "vector db",
        "redis", "bigquery"
    ],
    "delivery": [
        "agile", "scrum", "アジャイル", "スクラム", "要件定義", "設計",
        "顧客折衝", "リード", "レビュー", "テスト", "運用"
    ],
}

ROLE_PATTERNS = [
    r"シニア[^\n、。]*",
    r"リード[^\n、。]*",
    r"フルスタック[^\n、。]*",
    r"ソリューションアーキテクト",
    r"バックエンド[^\n、。]*",
    r"フロントエンド[^\n、。]*",
    r"AI[^\n、。]*エンジニア",
]

SAMPLE_ENGINEER_TEXT = (
    "【氏名】佐藤 賢太 (さとう けんた)\n"
    "【職種】シニアAIソリューションアーキテクト / フルスタックエンジニア\n"
    "【概要】IT業界経験8年。クラウドネイティブなWebアプリケーション開発、"
    "Python、JavaScript(TypeScript)、FastAPI、React、Google Cloud API、"
    "OpenAI、Gemini、gspread を用いた自律エージェント開発をリード。\n"
    "【主要スキル】Python, JavaScript, TypeScript, FastAPI, Django, React.js, Next.js, "
    "Vertex AI, Gemini API, gspread, SQL\n"
    "【インフラ/データベース】AWS, Google Cloud, PostgreSQL, Pinecone (Vector DB)\n"
    "【キャリア志向】生成AIを活用したプロダクト開発でビジネス価値を創造すること。"
)

SAMPLE_JOB_TEXT = (
    "【案件名】大手ITソリューション企業：LLM自律エージェント＆データ連携基盤開発\n"
    "【業務内容】生成AI(Gemini, GPT)を活用した業務プロセスの自動化・自律化エージェントの実装、"
    "Google API (Sheets API, Docs API) と連携した文書作成自動同期システムの構築、"
    "FastAPI / React.js を用いたWebアプリケーションの設計・開発。\n"
    "【必須スキル】Python/TypeScript実務開発、REST API(FastAPI等)設計構築、React.js実装実績。\n"
    "【歓迎スキル】Gemini/OpenAI API等のLLM連携実績、Google Cloud API (gspread, Drive API) 等のOAuth認証による連携実績。"
)


@dataclass
class ParsedProfile:
    doc_type: str
    title: str
    role: str
    summary: str
    experience_years: int = 0
    skills_by_category: Dict[str, List[str]] = field(default_factory=dict)
    all_skills: List[str] = field(default_factory=list)
    strengths: List[str] = field(default_factory=list)
    risk_flags: List[str] = field(default_factory=list)
    raw_excerpt: str = ""


def clamp(value: float, min_value: int = 50, max_value: int = 100) -> int:
    return max(min_value, min(max_value, round(value)))


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def now_utc_iso() -> str:
    return datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds")


def stable_digest(value: str) -> str:
    return hashlib.sha256((value or "").encode("utf-8")).hexdigest()[:16]


def safe_excerpt(value: str, limit: int = 220) -> str:
    normalized = normalize_text(value)
    return normalized[:limit]


def ensure_audit_dir():
    os.makedirs(AUDIT_DIR, exist_ok=True)


def write_audit_event(event_type: str, payload: dict) -> dict:
    """Append a privacy-conscious local audit event for later AI tuning."""
    ensure_audit_dir()
    timestamp = now_utc_iso()
    event = {
        "event_id": stable_digest(f"{event_type}:{timestamp}:{json.dumps(payload, ensure_ascii=False, sort_keys=True)}"),
        "timestamp_utc": timestamp,
        "event_type": event_type,
        "payload": payload,
    }
    with open(AUDIT_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(event, ensure_ascii=False) + "\n")
    return event


def read_recent_audit_events(limit: int = 20) -> List[dict]:
    if not os.path.exists(AUDIT_LOG_FILE):
        return []
    limit = max(1, min(limit, 100))
    with open(AUDIT_LOG_FILE, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f if line.strip()]
    events = []
    for line in lines[-limit:]:
        try:
            events.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return list(reversed(events))


def today_key() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d")


def utc_now_iso() -> str:
    return (
        datetime.datetime.now(datetime.timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


def read_external_api_events(limit: Optional[int] = None) -> List[dict]:
    if not os.path.exists(EXTERNAL_API_USAGE_LOG_FILE):
        return []
    with open(EXTERNAL_API_USAGE_LOG_FILE, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f if line.strip()]
    if limit:
        lines = lines[-max(1, min(limit, 1000)):]
    events = []
    for line in lines:
        try:
            events.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return events


def append_external_api_event(event: dict) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    safe_event = {
        "timestamp": utc_now_iso(),
        "day": today_key(),
        "provider": event.get("provider", "unknown"),
        "operation": event.get("operation", "unknown"),
        "billable": bool(event.get("billable", False)),
        "outcome": event.get("outcome", "unknown"),
        "model": event.get("model"),
        "task_id": event.get("task_id"),
        "http_status": event.get("http_status"),
        "reported_total_tokens": event.get("reported_total_tokens"),
        "reported_input_tokens": event.get("reported_input_tokens"),
        "reported_output_tokens": event.get("reported_output_tokens"),
        "token_source": event.get("token_source", "provider_not_reported"),
        "reason": event.get("reason"),
        "prompt_digest": event.get("prompt_digest"),
    }
    with open(EXTERNAL_API_USAGE_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(safe_event, ensure_ascii=False) + "\n")


def find_token_usage(value) -> dict:
    """Extract provider-reported usage tokens from common API response shapes."""
    usage = {
        "reported_total_tokens": None,
        "reported_input_tokens": None,
        "reported_output_tokens": None,
        "token_source": "provider_not_reported",
    }
    if value is None:
        return usage
    if not isinstance(value, (dict, list)):
        value = {
            "total_token_count": getattr(value, "total_token_count", None),
            "prompt_token_count": getattr(value, "prompt_token_count", None),
            "candidates_token_count": getattr(value, "candidates_token_count", None),
            "total_tokens": getattr(value, "total_tokens", None),
            "input_tokens": getattr(value, "input_tokens", None),
            "output_tokens": getattr(value, "output_tokens", None),
        }
    if isinstance(value, list):
        for item in value:
            found = find_token_usage(item)
            if found["reported_total_tokens"] is not None:
                return found
        return usage
    if isinstance(value, dict):
        candidates = []
        for key in ("usage", "usage_metadata", "token_usage", "billing", "data", "result"):
            if isinstance(value.get(key), dict):
                candidates.append(value[key])
        candidates.append(value)
        for candidate in candidates:
            total = (
                candidate.get("total_tokens")
                or candidate.get("total_token_count")
                or candidate.get("tokens")
                or candidate.get("total")
            )
            input_tokens = (
                candidate.get("input_tokens")
                or candidate.get("prompt_tokens")
                or candidate.get("prompt_token_count")
            )
            output_tokens = (
                candidate.get("output_tokens")
                or candidate.get("completion_tokens")
                or candidate.get("candidates_token_count")
            )
            if total is None and (input_tokens is not None or output_tokens is not None):
                total = int(input_tokens or 0) + int(output_tokens or 0)
            if total is not None:
                return {
                    "reported_total_tokens": int(total),
                    "reported_input_tokens": int(input_tokens) if input_tokens is not None else None,
                    "reported_output_tokens": int(output_tokens) if output_tokens is not None else None,
                    "token_source": "provider_response",
                }
        for item in value.values():
            found = find_token_usage(item)
            if found["reported_total_tokens"] is not None:
                return found
    return usage


def external_api_daily_stats(provider: str, operation: Optional[str] = None) -> dict:
    events = [
        event for event in read_external_api_events()
        if event.get("day") == today_key()
        and event.get("provider") == provider
        and (operation is None or event.get("operation") == operation)
    ]
    billable_events = [event for event in events if event.get("billable")]
    return {
        "events": len(events),
        "billable_calls": len(billable_events),
        "blocked_calls": sum(1 for event in events if event.get("outcome") == "blocked"),
        "reported_total_tokens": sum(int(event.get("reported_total_tokens") or 0) for event in events),
    }


def check_external_api_circuit(provider: str, operation: str, call_limit: int, token_limit: int = 0) -> Tuple[bool, str, dict]:
    stats = external_api_daily_stats(provider, operation)
    if call_limit <= 0:
        return False, f"{provider}:{operation} is disabled by daily call limit 0.", stats
    if stats["billable_calls"] >= call_limit:
        return False, f"{provider}:{operation} daily call limit reached ({stats['billable_calls']}/{call_limit}).", stats
    if token_limit > 0 and stats["reported_total_tokens"] >= token_limit:
        return False, f"{provider}:{operation} daily reported token limit reached ({stats['reported_total_tokens']}/{token_limit}).", stats
    return True, "allowed", stats


def build_external_api_usage_summary() -> dict:
    events = read_external_api_events()
    today = today_key()
    providers = {}
    for provider in ("seedance_api", "gemini_api"):
        provider_events = [event for event in events if event.get("provider") == provider]
        today_events = [event for event in provider_events if event.get("day") == today]
        providers[provider] = {
            "total_events": len(provider_events),
            "today_events": len(today_events),
            "today_billable_calls": sum(1 for event in today_events if event.get("billable")),
            "today_blocked_calls": sum(1 for event in today_events if event.get("outcome") == "blocked"),
            "today_reported_total_tokens": sum(int(event.get("reported_total_tokens") or 0) for event in today_events),
        }
    seedance_create_stats = external_api_daily_stats("seedance_api", "generation_create")
    gemini_parse_stats = external_api_daily_stats("gemini_api", "parse")
    gemini_match_stats = external_api_daily_stats("gemini_api", "match")
    return {
        "status": "success",
        "day": today,
        "usage_log": os.path.relpath(EXTERNAL_API_USAGE_LOG_FILE, PROJECT_ROOT),
        "providers": providers,
        "circuit_breakers": {
            "seedance_generation_create": {
                "enabled": SEEDANCE_API_ENABLED,
                "configured": SEEDANCE_CONFIGURED,
                "daily_call_limit": SEEDANCE_DAILY_GENERATION_LIMIT,
                "today_billable_calls": seedance_create_stats["billable_calls"],
                "today_reported_total_tokens": seedance_create_stats["reported_total_tokens"],
                "daily_reported_token_limit": SEEDANCE_DAILY_REPORTED_TOKEN_LIMIT or None,
                "state": "open" if (
                    not SEEDANCE_API_ENABLED
                    or seedance_create_stats["billable_calls"] >= SEEDANCE_DAILY_GENERATION_LIMIT
                    or (
                        SEEDANCE_DAILY_REPORTED_TOKEN_LIMIT > 0
                        and seedance_create_stats["reported_total_tokens"] >= SEEDANCE_DAILY_REPORTED_TOKEN_LIMIT
                    )
                ) else "closed",
            },
            "gemini_parse": {
                "enabled": GEMINI_READY,
                "daily_call_limit": GEMINI_DAILY_CALL_LIMIT,
                "today_billable_calls": gemini_parse_stats["billable_calls"],
                "today_reported_total_tokens": gemini_parse_stats["reported_total_tokens"],
            },
            "gemini_match": {
                "enabled": GEMINI_READY,
                "daily_call_limit": GEMINI_DAILY_CALL_LIMIT,
                "today_billable_calls": gemini_match_stats["billable_calls"],
                "today_reported_total_tokens": gemini_match_stats["reported_total_tokens"],
            },
        },
        "seedance_saved_default": read_seedance_manifest(),
        "recent_events": list(reversed(read_external_api_events(limit=100))),
        "usage_note": "Provider-reported token usage is shown when the API response includes it. Seedance video responses may not include tokens; use BytePlus Console > ModelArk > Usage as the spend source of truth.",
    }


def seedance_demo_video_url() -> Optional[str]:
    if os.path.exists(SEEDANCE_DEMO_VIDEO):
        return "/exports/seedance_demo/mighty_skill_bridge_seedance_demo.mp4"
    return None


def read_seedance_manifest() -> Optional[dict]:
    if not os.path.exists(SEEDANCE_DEMO_MANIFEST):
        return None
    with open(SEEDANCE_DEMO_MANIFEST, "r", encoding="utf-8-sig") as f:
        return json.load(f)


def find_video_url(value):
    """Return the first plausible video URL from nested Seedance-like payloads."""
    if isinstance(value, str):
        lowered = value.lower()
        if lowered.startswith(("http://", "https://", "/exports/")) and any(
            marker in lowered for marker in (".mp4", ".webm", ".mov", "video", "output")
        ):
            return value
        return None
    if isinstance(value, list):
        for item in value:
            found = find_video_url(item)
            if found:
                return found
        return None
    if isinstance(value, dict):
        preferred_keys = (
            "video_url",
            "url",
            "output_url",
            "download_url",
            "content_url",
            "video",
            "videos",
            "output",
            "data",
            "result",
            "results",
            "assets",
        )
        for key in preferred_keys:
            if key in value:
                found = find_video_url(value[key])
                if found:
                    return found
        for item in value.values():
            found = find_video_url(item)
            if found:
                return found
    return None


def find_task_id(value) -> Optional[str]:
    if isinstance(value, dict):
        for key in ("task_id", "id", "request_id", "job_id"):
            if key in value and value[key]:
                return str(value[key])
        for item in value.values():
            found = find_task_id(item)
            if found:
                return found
    if isinstance(value, list):
        for item in value:
            found = find_task_id(item)
            if found:
                return found
    return None


def find_task_status(value) -> Optional[str]:
    if isinstance(value, dict):
        for key in ("status", "task_status", "TaskStatus", "state", "State", "phase"):
            if key in value and value[key] is not None and not isinstance(value[key], (dict, list)):
                return str(value[key])
        for item in value.values():
            found = find_task_status(item)
            if found:
                return found
    if isinstance(value, list):
        for item in value:
            found = find_task_status(item)
            if found:
                return found
    return None


def seedance_fallback_response(reason: str, prompt: str, task_id: Optional[str] = None) -> dict:
    video_url = seedance_demo_video_url()
    return {
        "status": "success",
        "mode": "fallback",
        "provider": "local_seedance_demo_asset",
        "model": SEEDANCE_MODEL,
        "video_url": video_url,
        "task_id": task_id,
        "fallback_reason": reason,
        "manifest": read_seedance_manifest(),
        "prompt_digest": stable_digest(prompt),
    }


def seedance_pending_response(reason: str, prompt: str, task_id: str, raw_status: Optional[str] = None) -> dict:
    return {
        "status": "success",
        "mode": "pending",
        "provider": "seedance_api",
        "model": SEEDANCE_MODEL,
        "video_url": seedance_demo_video_url(),
        "task_id": task_id,
        "task_status": raw_status or "running",
        "fallback_reason": reason,
        "manifest": read_seedance_manifest(),
        "prompt_digest": stable_digest(prompt),
    }


def summarize_seedance_http_error(response: requests.Response) -> str:
    """Expose enough provider error detail for setup debugging without logging credentials."""
    try:
        error_payload = response.json()
    except ValueError:
        error_payload = response.text
    if isinstance(error_payload, (dict, list)):
        detail = json.dumps(error_payload, ensure_ascii=False)
    else:
        detail = str(error_payload)
    detail = detail.replace(SEEDANCE_API_KEY or "", "[redacted]")
    if len(detail) > 1200:
        detail = detail[:1200] + "...[truncated]"
    return f"{response.status_code} {response.reason}: {detail}"


def build_seedance_payload(prompt: str, req: "SeedanceVideoRequest") -> dict:
    """Build a ModelArk Seedance task payload.

    ModelArk's current Seedance task API expects text input under `content`.
    Keep a `prompt_legacy` escape hatch because BytePlus has multiple media endpoints.
    """
    if SEEDANCE_PAYLOAD_STYLE == "prompt_legacy":
        return {
            "model": SEEDANCE_MODEL,
            "prompt": prompt,
            "aspect_ratio": req.aspect_ratio,
            "duration": req.duration_seconds,
        }
    return {
        "model": SEEDANCE_MODEL,
        "content": [
            {
                "type": "text",
                "text": prompt,
            }
        ],
        "ratio": req.aspect_ratio,
        "duration": req.duration_seconds,
        "generate_audio": False,
    }


def poll_seedance_result(task_id: str, headers: dict) -> Tuple[Optional[str], Optional[dict], str]:
    if not SEEDANCE_RESULT_API_URL_TEMPLATE:
        return (
            None,
            None,
            "Seedance task was accepted, but SEEDANCE_RESULT_API_URL_TEMPLATE is not configured for result polling.",
        )
    if SEEDANCE_POLL_TIMEOUT_SECONDS <= 0:
        return None, None, "Seedance task was accepted, but result polling is disabled."

    result_url = SEEDANCE_RESULT_API_URL_TEMPLATE.format(task_id=task_id)
    deadline = time.monotonic() + SEEDANCE_POLL_TIMEOUT_SECONDS
    attempts = 0
    last_status = "unknown"
    last_payload = None

    while True:
        attempts += 1
        response = requests.get(result_url, headers=headers, timeout=45)
        if not response.ok:
            append_external_api_event({
                "provider": "seedance_api",
                "operation": "task_poll",
                "billable": False,
                "outcome": "http_error",
                "model": SEEDANCE_MODEL,
                "task_id": task_id,
                "http_status": response.status_code,
                "reason": summarize_seedance_http_error(response),
            })
            return (
                None,
                last_payload,
                f"Seedance result request failed: {summarize_seedance_http_error(response)}",
            )

        result_payload = response.json()
        last_payload = result_payload
        usage = find_token_usage(result_payload)
        video_url = find_video_url(result_payload)
        if video_url:
            append_external_api_event({
                "provider": "seedance_api",
                "operation": "task_poll",
                "billable": False,
                "outcome": "live",
                "model": SEEDANCE_MODEL,
                "task_id": task_id,
                "http_status": response.status_code,
                **usage,
            })
            return video_url, result_payload, ""

        last_status = find_task_status(result_payload) or last_status
        append_external_api_event({
            "provider": "seedance_api",
            "operation": "task_poll",
            "billable": False,
            "outcome": "pending",
            "model": SEEDANCE_MODEL,
            "task_id": task_id,
            "http_status": response.status_code,
            "reason": f"task_status={last_status}",
            **usage,
        })
        if time.monotonic() >= deadline:
            break
        time.sleep(SEEDANCE_POLL_INTERVAL_SECONDS)

    return (
        None,
        last_payload,
        f"Seedance task was accepted but no video URL was returned within {SEEDANCE_POLL_TIMEOUT_SECONDS}s. "
        f"task_status={last_status}; attempts={attempts}",
    )


def profile_audit_payload(profile: ParsedProfile, ai_mode: str, fallback_reason: str, source_text: str, file_name: Optional[str] = None) -> dict:
    return {
        "ai_mode": ai_mode,
        "fallback_reason": fallback_reason,
        "doc_type": profile.doc_type,
        "title": profile.title,
        "role": profile.role,
        "experience_years": profile.experience_years,
        "all_skills": profile.all_skills,
        "strengths": profile.strengths,
        "risk_flags": profile.risk_flags,
        "source_digest": stable_digest(source_text),
        "source_length": len(source_text or ""),
        "source_excerpt": safe_excerpt(source_text),
        "file_name": file_name,
    }


def match_audit_payload(match_data: dict) -> dict:
    structured = match_data.get("structured", {})
    candidate = structured.get("candidate", {})
    job = structured.get("job", {})
    return {
        "ai_mode": match_data.get("ai_mode"),
        "fallback_reason": match_data.get("fallback_reason"),
        "final_score": match_data.get("final_score"),
        "scores": match_data.get("scores", {}),
        "candidate_title": candidate.get("title"),
        "job_title": job.get("title"),
        "candidate_role": candidate.get("role"),
        "job_role": job.get("role"),
        "matched_skills": structured.get("matched_skills", []),
        "missing_skills": structured.get("missing_skills", []),
        "summary_excerpt": safe_excerpt(match_data.get("summary", "")),
    }


def decode_uploaded_text(file_bytes: bytes) -> str:
    if not file_bytes:
        return ""
    for encoding in ("utf-8", "cp932", "shift_jis"):
        try:
            decoded = file_bytes.decode(encoding)
            if decoded.strip():
                return decoded
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="ignore")


def extract_labeled_value(text: str, labels: List[str]) -> str:
    for label in labels:
        pattern = rf"【{re.escape(label)}】\s*([^\n]+)"
        match = re.search(pattern, text)
        if match:
            return match.group(1).strip()
    return ""


def extract_experience_years(text: str) -> int:
    candidates = []
    patterns = [
        r"([0-9]{1,2})\s*年以上",
        r"([0-9]{1,2})\s*年",
        r"([0-9]{1,2})\s*(?:years?|yrs?)\b",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            candidates.append(int(match.group(1)))
    return max(candidates) if candidates else 0


def detect_skills(text: str) -> Tuple[Dict[str, List[str]], List[str]]:
    lower_text = text.lower()
    skills_by_category = {}
    all_skills: Set[str] = set()
    for category, keywords in SKILL_TAXONOMY.items():
        detected = []
        for keyword in keywords:
            if keyword_matches(keyword, lower_text):
                detected.append(keyword)
                all_skills.add(keyword)
        skills_by_category[category] = sorted(set(detected), key=str.lower)
    return skills_by_category, sorted(all_skills, key=str.lower)


def keyword_matches(keyword: str, lower_text: str) -> bool:
    key = keyword.lower()
    if re.fullmatch(r"[a-z0-9.+#/-]+", key):
        return re.search(rf"(?<![a-z0-9]){re.escape(key)}(?![a-z0-9])", lower_text) is not None
    return key in lower_text


def detect_role(text: str, doc_type: str) -> str:
    labeled = extract_labeled_value(text, ["職種", "役割", "ポジション", "案件名"])
    if labeled:
        return labeled
    for pattern in ROLE_PATTERNS:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(0).strip()
    return "AIソリューションエンジニア" if doc_type == "engineer" else "AI / Google Workspace 連携案件"


def detect_title(text: str, doc_type: str) -> str:
    labels = ["氏名", "名前"] if doc_type == "engineer" else ["案件名", "求人名", "タイトル"]
    labeled = extract_labeled_value(text, labels)
    if labeled:
        return labeled
    return "候補者プロフィール" if doc_type == "engineer" else "対象案件"


def build_profile(text: str, doc_type: str) -> ParsedProfile:
    source = text.strip() or (SAMPLE_ENGINEER_TEXT if doc_type == "engineer" else SAMPLE_JOB_TEXT)
    skills_by_category, all_skills = detect_skills(source)
    role = detect_role(source, doc_type)
    title = detect_title(source, doc_type)
    years = extract_experience_years(source)
    normalized = normalize_text(source)

    category_strengths = [
        category for category, skills in skills_by_category.items() if len(skills) >= 2
    ]
    strengths = []
    if years:
        strengths.append(f"実務経験 {years} 年相当")
    strengths.extend([f"{category} 領域の実装経験" for category in category_strengths[:4]])
    if not strengths and all_skills:
        strengths.append(f"{', '.join(all_skills[:3])} を中心とした技術経験")

    risk_flags = []
    if doc_type == "engineer" and "google_workspace" not in category_strengths:
        risk_flags.append("Google Workspace API 連携経験の深掘り確認")
    if doc_type == "engineer" and "ai" not in category_strengths:
        risk_flags.append("LLM / 生成AI 連携経験の具体性確認")
    if doc_type == "job" and not all_skills:
        risk_flags.append("必須スキル要件が未構造化")

    return ParsedProfile(
        doc_type=doc_type,
        title=title,
        role=role,
        summary=normalized[:260],
        experience_years=years,
        skills_by_category=skills_by_category,
        all_skills=all_skills,
        strengths=strengths[:5],
        risk_flags=risk_flags[:4],
        raw_excerpt=normalized[:600],
    )


def format_profile(profile: ParsedProfile) -> str:
    label = "氏名" if profile.doc_type == "engineer" else "案件名"
    skill_lines = []
    for category, skills in profile.skills_by_category.items():
        if skills:
            skill_lines.append(f"- {category}: {', '.join(skills)}")
    strengths = "\n".join(f"- {item}" for item in profile.strengths) or "- 入力内容から詳細確認が必要"
    risks = "\n".join(f"- {item}" for item in profile.risk_flags) or "- 重大な未確認リスクなし"
    return (
        f"【{label}】{profile.title}\n"
        f"【役割】{profile.role}\n"
        f"【経験年数】{profile.experience_years or '未記載'}\n"
        f"【抽出スキル】\n{chr(10).join(skill_lines) if skill_lines else '- 未抽出'}\n"
        f"【強み】\n{strengths}\n"
        f"【確認ポイント】\n{risks}\n"
        f"【要約】{profile.summary}"
    )


def overlap_ratio(candidate_skills: List[str], job_skills: List[str]) -> float:
    if not job_skills:
        return 0.55
    candidate_set = {item.lower() for item in candidate_skills}
    job_set = {item.lower() for item in job_skills}
    return len(candidate_set & job_set) / max(len(job_set), 1)


def category_overlap(candidate: ParsedProfile, job: ParsedProfile, category: str) -> float:
    return overlap_ratio(
        candidate.skills_by_category.get(category, []),
        job.skills_by_category.get(category, []),
    )


def build_fallback_match(engineer_text: str, job_text: str, fallback_reason: str) -> dict:
    candidate = build_profile(engineer_text, "engineer")
    job = build_profile(job_text, "job")

    skill_ratio = overlap_ratio(candidate.all_skills, job.all_skills)
    ai_ratio = category_overlap(candidate, job, "ai")
    workspace_ratio = category_overlap(candidate, job, "google_workspace")
    backend_ratio = category_overlap(candidate, job, "backend")
    frontend_ratio = category_overlap(candidate, job, "frontend")
    delivery_ratio = category_overlap(candidate, job, "delivery")

    breadth_bonus = min(len(candidate.all_skills), 12) * 0.9
    skill_score = clamp(58 + (skill_ratio * 36) + breadth_bonus)
    culture_score = clamp(64 + (delivery_ratio * 22) + (8 if candidate.experience_years >= 5 else 0))
    growth_score = clamp(62 + (ai_ratio * 18) + (workspace_ratio * 12) + (backend_ratio * 5))
    performing_score = clamp(56 + (backend_ratio * 16) + (frontend_ratio * 10) + (workspace_ratio * 12) + min(candidate.experience_years, 10))
    final_score = clamp((skill_score * 0.36) + (culture_score * 0.18) + (growth_score * 0.24) + (performing_score * 0.22))

    matched_skills = sorted(
        {item for item in candidate.all_skills if item.lower() in {skill.lower() for skill in job.all_skills}},
        key=str.lower,
    )
    missing_skills = sorted(
        {item for item in job.all_skills if item.lower() not in {skill.lower() for skill in candidate.all_skills}},
        key=str.lower,
    )
    matched_label = ", ".join(matched_skills[:8])
    gap_label = ", ".join(missing_skills[:6])

    # 一致/不足スキルが空のときはプレースホルダ文を質問文へ埋め込まず、
    # スキル名に依存しない文型へ切り替える (R47)
    if matched_label:
        summary_matches = f"主要一致スキルは {matched_label}。"
        qa_matches_question = f"{matched_label} を使った実装経験を、担当範囲・設計判断・成果指標に分けて説明してください。"
    else:
        summary_matches = "要件に対する明示的な一致スキルは限定的でした。"
        qa_matches_question = "これまでの代表的な実装経験を、担当範囲・設計判断・成果指標に分けて説明してください。"

    if gap_label:
        summary_gaps = f"確認すべきギャップは {gap_label}。"
        qa_gaps_question = f"現時点で不足候補として見えている {gap_label} を、着任後どの順番でキャッチアップしますか？"
        week1_focus = gap_label
    else:
        summary_gaps = "大きな未充足スキルは検出されていません。"
        qa_gaps_question = "この案件で新たに求められる技術領域を、着任後どの順番でキャッチアップしますか？"
        week1_focus = "必須スキル"

    summary = (
        f"{candidate.title} と {job.title} の適合度は {final_score}% です。"
        f"{summary_matches}"
        f"特に backend / AI / Google Workspace 連携の重なりを中心に評価しました。"
        f"{summary_gaps}"
        "Gemini live 復帰後は、この構造化プロファイルをそのままプロンプトへ渡すことで、"
        "より深い文脈評価に即時移行できます。"
    )

    qa = [
        {
            "question": qa_matches_question,
            "answer": "単なる利用経験ではなく、要件定義、API設計、認証、例外処理、運用時の監視までを一連の流れとして説明します。",
            "tip": "Google API や OAuth、バッチ更新、クォータ回避など、本プロジェクトで求められる実務上の判断を具体例に落とし込むと強いです。"
        },
        {
            "question": qa_gaps_question,
            "answer": "初週で既存仕様と認証フローを把握し、2週目で小さな検証実装、3週目で本番相当のエラーハンドリングとログ設計へ広げる計画を示します。",
            "tip": "不足を隠さず、検証単位・成果物・レビュー方法まで言語化すると信頼感が増します。"
        },
        {
            "question": "生成AIを外部業務システムへ組み込む際、どのように品質と安全性を担保しますか？",
            "answer": "構造化入出力、fallback、監査ログ、権限分離、手動確認ポイントを設計に含め、AI応答をそのまま業務データへ反映しない方針を説明します。",
            "tip": "AI live と deterministic fallback の二層構造を説明できると、今回の開発方針とよく噛み合います。"
        }
    ]

    return {
        "final_score": final_score,
        "scores": {
            "skill": skill_score,
            "culture": culture_score,
            "growth": growth_score,
            "performing": performing_score,
        },
        "summary": summary,
        "qa": qa,
        "roadmap_week1": f"{job.role} の業務範囲、OAuth / Google API 認証、既存 FastAPI 構成を把握し、{week1_focus} の検証観点を洗い出す",
        "roadmap_week2": "構造化パーサー、スキル分類、4軸スコアリングの小さな改善を行い、Sheets への診断ログ保存までを通す",
        "roadmap_week3": "Gemini live 復帰を想定し、プロンプト入力に渡す structured_profile / gap_analysis / scoring_context を安定化する",
        "roadmap_week4": "Browser Agent による主要シナリオ確認、エラー時 fallback、監査ログ、共有手順を整備して社長報告用デモ品質まで高める",
        "ai_mode": "deterministic_fallback",
        "fallback_reason": fallback_reason,
        "structured": {
            "candidate": asdict(candidate),
            "job": asdict(job),
            "matched_skills": matched_skills,
            "missing_skills": missing_skills,
        }
    }

# Dynamic Environment Diagnostics
API_KEY = os.environ.get("GEMINI_API_KEY")
AI_FORCE_MOCK = os.environ.get("AI_FORCE_MOCK", "").lower() in {"1", "true", "yes", "on"}
GEMINI_READY = API_KEY is not None and GEMINI_LIB_AVAILABLE and not AI_FORCE_MOCK
GEMINI_CLIENT = None
DETERMINISTIC_PARSE_DELAY_SECONDS = env_float("DETERMINISTIC_PARSE_DELAY_SECONDS", 0.0, 0.0, 5.0)
DETERMINISTIC_MATCH_DELAY_SECONDS = env_float("DETERMINISTIC_MATCH_DELAY_SECONDS", 0.0, 0.0, 5.0)

if GEMINI_READY:
    GEMINI_CLIENT = genai.Client(api_key=API_KEY)
    print(f"[+] Gemini API successfully configured via GEMINI_API_KEY using {GEMINI_MODEL}.")
elif AI_FORCE_MOCK:
    print("[!] AI_FORCE_MOCK enabled. Running in quota-safe mock fallback mode.")
elif not GEMINI_LIB_AVAILABLE:
    print("[!] google-genai is not installed. Running in deterministic fallback mode.")
else:
    print("[*] GEMINI_API_KEY not set. Running in deterministic fallback mode.")


def generate_gemini_content(contents, response_mime_type: Optional[str] = None):
    if GEMINI_CLIENT is None:
        raise RuntimeError("Gemini client is not configured.")
    config = None
    if response_mime_type:
        config = genai_types.GenerateContentConfig(response_mime_type=response_mime_type)
    return GEMINI_CLIENT.models.generate_content(
        model=GEMINI_MODEL,
        contents=contents,
        config=config,
    )


SEEDANCE_CONFIGURED = bool(SEEDANCE_API_KEY and SEEDANCE_API_URL)
SEEDANCE_READY = bool(SEEDANCE_API_ENABLED and SEEDANCE_CONFIGURED)
if SEEDANCE_READY:
    print("[+] Seedance API adapter enabled via environment variables.")
elif SEEDANCE_CONFIGURED:
    print("[*] Seedance API credentials detected, but external calls are disabled. Set SEEDANCE_API_ENABLED=1 to enable billing calls.")
else:
    print("[!] Seedance API credentials not set. Using local demo video fallback.")


class SeedanceVideoRequest(BaseModel):
    prompt: str
    aspect_ratio: str = "16:9"
    duration_seconds: int = 6


class StripeCustomerPortalSessionRequest(BaseModel):
    customer_id: str
    return_url: Optional[str] = None
    subscription_id: Optional[str] = None
    flow_type: Optional[str] = None
    configuration_id: Optional[str] = None
    locale: Optional[str] = None
    dry_run: bool = False


def stripe_customer_portal_return_url(requested_url: Optional[str]) -> str:
    return (
        (requested_url or "").strip()
        or os.environ.get("STRIPE_CUSTOMER_PORTAL_RETURN_URL", "").strip()
        or "https://mightylink-app.com/billing"
    )


def stripe_customer_portal_required_env() -> List[str]:
    return [
        "STRIPE_CUSTOMER_PORTAL_ENABLED=1",
        "STRIPE_SECRET_KEY",
        "STRIPE_CUSTOMER_PORTAL_RETURN_URL",
        "STRIPE_CUSTOMER_PORTAL_CONFIGURATION_ID (optional)",
    ]


BILLING_PORTAL_HTML = """<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Mighty Skill-Bridge Billing</title>
    <style>
        :root {
            --bg: #f7f8fb;
            --panel: #ffffff;
            --line: #d8dde7;
            --text: #18202f;
            --muted: #5d6676;
            --accent: #0f766e;
            --accent-strong: #0b5f59;
            --danger: #b42318;
        }
        * { box-sizing: border-box; }
        body {
            margin: 0;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            background: var(--bg);
            color: var(--text);
        }
        main {
            width: min(920px, calc(100% - 32px));
            margin: 32px auto;
            display: grid;
            gap: 18px;
        }
        header {
            display: flex;
            justify-content: space-between;
            gap: 16px;
            align-items: flex-end;
            border-bottom: 1px solid var(--line);
            padding-bottom: 16px;
        }
        h1 { font-size: 24px; margin: 0; }
        .home {
            color: var(--accent);
            text-decoration: none;
            font-weight: 700;
            white-space: nowrap;
        }
        .panel {
            background: var(--panel);
            border: 1px solid var(--line);
            border-radius: 8px;
            padding: 20px;
        }
        form {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 14px;
        }
        label {
            display: grid;
            gap: 6px;
            font-size: 13px;
            color: var(--muted);
            font-weight: 700;
        }
        input, select {
            min-height: 42px;
            border: 1px solid var(--line);
            border-radius: 6px;
            padding: 10px 12px;
            font: inherit;
            color: var(--text);
            background: #fff;
        }
        .full { grid-column: 1 / -1; }
        .check {
            display: flex;
            align-items: center;
            gap: 8px;
            min-height: 42px;
        }
        .check input { min-height: auto; }
        .actions {
            grid-column: 1 / -1;
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
            align-items: center;
        }
        button, .launch {
            min-height: 42px;
            border: 0;
            border-radius: 6px;
            padding: 0 16px;
            background: var(--accent);
            color: #fff;
            font: inherit;
            font-weight: 800;
            cursor: pointer;
            text-decoration: none;
            display: inline-flex;
            align-items: center;
        }
        button:hover, .launch:hover { background: var(--accent-strong); }
        .launch[hidden] { display: none; }
        pre {
            min-height: 164px;
            margin: 0;
            overflow: auto;
            white-space: pre-wrap;
            word-break: break-word;
            font-size: 13px;
            line-height: 1.55;
            color: var(--text);
        }
        .error { color: var(--danger); }
        @media (max-width: 720px) {
            main { width: min(100% - 24px, 920px); margin: 20px auto; }
            header { align-items: flex-start; flex-direction: column; }
            form { grid-template-columns: 1fr; }
        }
    </style>
</head>
<body>
    <main>
        <header>
            <h1>Billing Portal</h1>
            <a class="home" href="/">Mighty Skill-Bridge</a>
        </header>
        <section class="panel">
            <form id="portal-form">
                <label>Customer ID
                    <input id="customer-id" name="customer_id" required placeholder="cus_..." autocomplete="off">
                </label>
                <label>Flow
                    <select id="flow-type" name="flow_type">
                        <option value="">Portal home</option>
                        <option value="subscription_cancel">Cancel subscription</option>
                        <option value="subscription_update">Change plan</option>
                        <option value="payment_method_update">Payment method</option>
                    </select>
                </label>
                <label>Subscription ID
                    <input id="subscription-id" name="subscription_id" placeholder="sub_..." autocomplete="off">
                </label>
                <label>Configuration ID
                    <input id="configuration-id" name="configuration_id" placeholder="bpc_..." autocomplete="off">
                </label>
                <label class="full">Return URL
                    <input id="return-url" name="return_url" placeholder="https://mightylink-app.com/billing">
                </label>
                <label class="check full">
                    <input id="dry-run" name="dry_run" type="checkbox" checked>
                    Dry-run
                </label>
                <div class="actions">
                    <button type="submit">Create Session</button>
                    <a id="launch-link" class="launch" hidden target="_blank" rel="noopener">Open Portal</a>
                </div>
            </form>
        </section>
        <section class="panel">
            <pre id="result">Ready.</pre>
        </section>
    </main>
    <script>
        const form = document.getElementById("portal-form");
        const result = document.getElementById("result");
        const launch = document.getElementById("launch-link");

        form.addEventListener("submit", async (event) => {
            event.preventDefault();
            launch.hidden = true;
            launch.removeAttribute("href");
            result.className = "";
            const payload = {
                customer_id: document.getElementById("customer-id").value,
                return_url: document.getElementById("return-url").value || null,
                subscription_id: document.getElementById("subscription-id").value || null,
                configuration_id: document.getElementById("configuration-id").value || null,
                flow_type: document.getElementById("flow-type").value || null,
                dry_run: document.getElementById("dry-run").checked
            };
            try {
                const response = await fetch("/api/billing/customer-portal/session", {
                    method: "POST",
                    headers: {"Content-Type": "application/json"},
                    body: JSON.stringify(payload)
                });
                const data = await response.json();
                if (!response.ok) {
                    result.className = "error";
                }
                result.textContent = JSON.stringify(data, null, 2);
                if (data.url) {
                    launch.href = data.url;
                    launch.hidden = false;
                }
            } catch (error) {
                result.className = "error";
                result.textContent = String(error);
            }
        });
    </script>
</body>
</html>
"""

# Static Hosting route
@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    """Serves the project favicon for browser tabs and DevTools requests."""
    if not os.path.exists(FAVICON_FILE):
        raise HTTPException(status_code=404, detail="favicon.ico not found in project root.")
    return FileResponse(FAVICON_FILE, media_type="image/x-icon")


@app.get(CHROME_DEVTOOLS_WORKSPACE_PATH, include_in_schema=False)
async def chrome_devtools_workspace():
    """Lets Chrome DevTools connect this localhost app to the workspace without a 404."""
    return JSONResponse({
        "workspace": {
            "root": os.path.abspath(PROJECT_ROOT),
            "uuid": DEVTOOLS_WORKSPACE_UUID,
        }
    })


@app.get("/", response_class=HTMLResponse)
async def serve_index(username: str = Depends(verify_credentials)):
    """Serves the main frontend page index.html after HTTP Basic authentication."""
    try:
        index_path = os.path.join(ROOT_DIR, "index.html")
        if not os.path.exists(index_path):
            raise FileNotFoundError()
        
        content = None
        # 1. Try reading as strict UTF-8
        try:
            with open(index_path, "r", encoding="utf-8") as f:
                content = f.read()
            print("[+] Successfully loaded index.html in strict UTF-8.")
        except UnicodeDecodeError:
            # 2. Try reading as UTF-8, replacing illegal bytes instead of falling back to CP932
            # Since index.html is written in UTF-8, falling back to CP932 causes full text corruption.
            try:
                with open(index_path, "r", encoding="utf-8", errors="replace") as f:
                    content = f.read()
                print("[!] Loaded index.html in UTF-8 with errors='replace'.")
            except Exception as e:
                # 3. Last resort fallback to CP932
                print(f"[-] UTF-8 fallback failed: {e}. Trying CP932...")
                with open(index_path, "r", encoding="cp932", errors="ignore") as f:
                    content = f.read()
                    
        return HTMLResponse(
            content=content,
            headers={"Cache-Control": "private, no-store, max-age=0"},
        )
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="index.html not found in project workspace.")


@app.get("/billing", response_class=HTMLResponse)
async def serve_billing_portal():
    """Serves a local Stripe Customer Portal launcher."""
    return HTMLResponse(content=BILLING_PORTAL_HTML)


# API Auth User Information Endpoint
@app.get("/api/auth/me")
async def get_auth_me(current_user: dict = Depends(get_current_user)):
    """Returns the authenticated user details from Firebase Auth."""
    return {
        "status": "success",
        "user": current_user
    }


@app.post("/api/billing/customer-portal/session")
async def create_billing_customer_portal_session(
    req: StripeCustomerPortalSessionRequest,
    current_user: dict = Depends(get_current_user),
):
    """Create, or preview, a short-lived Stripe Customer Portal session."""
    del current_user
    return_url = stripe_customer_portal_return_url(req.return_url)
    configuration_id = (
        (req.configuration_id or "").strip()
        or os.environ.get("STRIPE_CUSTOMER_PORTAL_CONFIGURATION_ID", "").strip()
        or None
    )
    try:
        payload = build_customer_portal_payload(
            customer_id=req.customer_id,
            return_url=return_url,
            configuration_id=configuration_id,
            flow_type=req.flow_type,
            subscription_id=req.subscription_id,
            locale=req.locale,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    enabled = env_flag("STRIPE_CUSTOMER_PORTAL_ENABLED", default=False)
    secret_key = os.environ.get("STRIPE_SECRET_KEY", "").strip()

    if req.dry_run or not enabled or not secret_key:
        mode = "dry_run" if req.dry_run else "not_configured"
        return {
            "status": "preview",
            "mode": mode,
            "enabled": enabled,
            "stripe_secret_configured": bool(secret_key),
            "payload": sanitize_stripe_portal_payload(payload),
            "required_env": stripe_customer_portal_required_env(),
        }

    try:
        session = create_customer_portal_session(
            secret_key=secret_key,
            payload=payload,
        )
    except StripePortalError as exc:
        raise HTTPException(
            status_code=502,
            detail={
                "message": str(exc),
                "stripe_status": exc.status_code,
                "stripe_code": exc.code,
            },
        ) from exc

    return {
        "status": "success",
        "id": session.get("id"),
        "url": session["url"],
        "livemode": session.get("livemode"),
        "return_url": session.get("return_url") or return_url,
    }


@app.get("/api/user-data/export")
async def export_user_data(
    session_id: str = "",
    current_user: dict = Depends(get_current_user_for_data_export),
):
    """Return the authenticated user's scoped data export as JSON."""
    try:
        export_payload = build_user_data_export(current_user, session_id=session_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        print(f"[-] User data export failed: {exc}")
        raise HTTPException(status_code=500, detail="Failed to build user data export") from exc

    filename = "mighty-link-user-data-export.json"
    return JSONResponse(
        content=export_payload,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# API Health Check Endpoint
@app.get("/api/health")
async def health_check():
    """Provides client-side dynamic connection check."""
    return {
        "status": "healthy",
        "sheets_live": SHEETS_LIB_AVAILABLE and (os.path.exists(CREDENTIALS_FILE) or os.path.exists(CLIENT_SECRET_FILE)),
        "gemini_live": GEMINI_READY,
        "ai_mode": "gemini_live" if GEMINI_READY else "deterministic_fallback",
        "ai_force_mock": AI_FORCE_MOCK,
        "deterministic_parse_delay_seconds": DETERMINISTIC_PARSE_DELAY_SECONDS,
        "deterministic_match_delay_seconds": DETERMINISTIC_MATCH_DELAY_SECONDS,
        "seedance_live": SEEDANCE_READY,
        "seedance_api_enabled": SEEDANCE_API_ENABLED,
        "seedance_credentials_configured": SEEDANCE_CONFIGURED,
        "seedance_model": SEEDANCE_MODEL,
        "seedance_result_polling": bool(SEEDANCE_RESULT_API_URL_TEMPLATE),
        "seedance_poll_timeout_seconds": SEEDANCE_POLL_TIMEOUT_SECONDS,
        "seedance_demo_video": seedance_demo_video_url(),
        "stripe_customer_portal_enabled": env_flag("STRIPE_CUSTOMER_PORTAL_ENABLED", default=False),
        "stripe_secret_configured": bool(os.environ.get("STRIPE_SECRET_KEY", "").strip()),
    }


@app.get("/api/db-test")
async def db_test(username: str = Depends(verify_credentials)):
    import traceback
    results = {
        "postgres_available": POSTGRES_AVAILABLE,
        "use_supabase": USE_SUPABASE,
        "database_url_configured": bool(DATABASE_URL),
        "pool": get_supabase_pool_status(),
        "steps": []
    }
    if DATABASE_URL and USE_SUPABASE and POSTGRES_AVAILABLE:
        conn = None
        cur = None
        try:
            results["steps"].append("Borrowing pooled database connection...")
            conn, db_type = get_db_connection()
            results["steps"].append(f"Connected via {db_type}. Running test query...")
            cur = conn.cursor()
            cur.execute("SELECT 1;")
            val = cur.fetchone()[0]
            results["steps"].append(f"Query returned: {val}")
            results["direct_postgres_status"] = "success" if db_type == "postgres" else "fallback_sqlite"
        except Exception as e:
            results["direct_postgres_status"] = "error"
            results["direct_postgres_error"] = str(e)
            results["direct_postgres_traceback"] = traceback.format_exc()
        finally:
            if cur is not None:
                cur.close()
            if conn is not None:
                conn.close()
    else:
        results["direct_postgres_status"] = "no_url"
    return results


@app.get("/api/audit/recent")
async def recent_audit_events(limit: int = 20, username: str = Depends(verify_credentials)):
    """Returns recent local AI audit events without raw document bodies."""
    return {
        "status": "success",
        "audit_log": os.path.relpath(AUDIT_LOG_FILE, PROJECT_ROOT),
        "events": read_recent_audit_events(limit)
    }


ADMIN_DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Mighty Skill-Bridge API Guard</title>
    <link rel="icon" href="/favicon.ico" sizes="any">
    <style>
        :root { color-scheme: dark; --bg:#070807; --panel:#111312; --line:#2a2f2b; --text:#f3f5ef; --muted:#a6ada4; --ok:#9df56d; --warn:#ffd166; --bad:#ff7d7d; --blue:#8bdcff; }
        * { box-sizing: border-box; }
        body { margin:0; font-family: "Segoe UI", "Noto Sans JP", sans-serif; background: var(--bg); color: var(--text); }
        header { display:flex; justify-content:space-between; align-items:center; gap:16px; padding:18px 24px; border-bottom:1px solid var(--line); background:#050605; position:sticky; top:0; }
        main { max-width:1180px; margin:0 auto; padding:24px; display:grid; gap:18px; }
        h1 { margin:0; font-size:22px; }
        h2 { margin:0 0 12px; font-size:16px; color:var(--muted); font-weight:700; }
        .grid { display:grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap:14px; }
        .card { background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:16px; }
        .metric { font-size:32px; font-weight:800; line-height:1; }
        .label { color:var(--muted); font-size:13px; margin-top:8px; }
        .state { display:inline-flex; align-items:center; padding:5px 10px; border:1px solid var(--line); border-radius:999px; font-size:13px; color:var(--muted); }
        .closed { color:var(--ok); border-color:rgba(157,245,109,.35); }
        .open { color:var(--bad); border-color:rgba(255,125,125,.35); }
        table { width:100%; border-collapse:collapse; font-size:13px; }
        th, td { text-align:left; padding:9px 8px; border-bottom:1px solid var(--line); vertical-align:top; }
        th { color:var(--muted); font-weight:600; }
        code { color:var(--blue); }
        a { color:var(--blue); text-decoration:none; }
        .muted { color:var(--muted); }
        .toolbar { display:flex; gap:10px; align-items:center; flex-wrap:wrap; }
        button, .button { border:1px solid var(--line); border-radius:7px; padding:9px 12px; background:#191c1a; color:var(--text); cursor:pointer; font-weight:700; }
        button:hover, .button:hover { border-color:var(--blue); }
        @media (max-width: 860px) { .grid { grid-template-columns:1fr; } header { align-items:flex-start; flex-direction:column; } }
    </style>
</head>
<body>
    <header>
        <div>
            <h1>External API Guard</h1>
            <div class="muted">Mighty Skill-Bridge local billing safety dashboard</div>
        </div>
        <div class="toolbar">
            <a class="button" href="/">Demo</a>
            <a class="button" href="/api/admin/usage/export">Export JSONL</a>
            <button onclick="loadUsage()">Refresh</button>
        </div>
    </header>
    <main>
        <section class="grid" id="cards"></section>
        <section class="card">
            <h2>Circuit Breakers</h2>
            <div id="breakers"></div>
        </section>
        <section class="card">
            <h2>Saved Seedance Video</h2>
            <div id="saved-video" class="muted"></div>
        </section>
        <section class="card">
            <h2>Antigravity 2.0 Managed Agents Cost Simulator (T688 監視体制)</h2>
            <div class="muted" style="margin-bottom:12px;">Google Vertex AI Agent Builderの公式料金モデルに基づき、正式採用時の月額コストを動的にシミュレーションします。</div>
            <div style="display:grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap:18px; margin-bottom:14px;">
                <div>
                    <label style="display:block; font-size:12px; color:var(--muted); margin-bottom:4px;">月間稼働時間 (Active vCPU Hours)</label>
                    <input type="number" id="sim-hours" value="20" style="width:100%; border:1px solid var(--line); border-radius:5px; padding:6px; background:#191c1a; color:var(--text);" oninput="updateSim()">
                </div>
                <div>
                    <label style="display:block; font-size:12px; color:var(--muted); margin-bottom:4px;">想定月間会話セッション数</label>
                    <input type="number" id="sim-sessions" value="10000" style="width:100%; border:1px solid var(--line); border-radius:5px; padding:6px; background:#191c1a; color:var(--text);" oninput="updateSim()">
                </div>
                <div>
                    <label style="display:block; font-size:12px; color:var(--muted); margin-bottom:4px;">想定RAGクエリ数 (Vertex AI Search)</label>
                    <input type="number" id="sim-queries" value="5000" style="width:100%; border:1px solid var(--line); border-radius:5px; padding:6px; background:#191c1a; color:var(--text);" oninput="updateSim()">
                </div>
                <div>
                    <label style="display:block; font-size:12px; color:var(--muted); margin-bottom:4px;">Gemini 入力/出力トークン量 (百万 tokens)</label>
                    <div style="display:flex; gap:6px;">
                        <input type="number" id="sim-input-tokens" value="10" style="width:50%; border:1px solid var(--line); border-radius:5px; padding:6px; background:#191c1a; color:var(--text);" oninput="updateSim()" placeholder="入力 (M)">
                        <input type="number" id="sim-output-tokens" value="2" style="width:50%; border:1px solid var(--line); border-radius:5px; padding:6px; background:#191c1a; color:var(--text);" oninput="updateSim()" placeholder="出力 (M)">
                    </div>
                </div>
            </div>
            <div style="display:grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap:12px; border-top:1px solid var(--line); padding-top:14px;">
                <div>
                    <div id="sim-total" class="metric" style="color:var(--blue); font-size:28px;">$0.00</div>
                    <div class="label">想定月額合計コスト (USD)</div>
                </div>
                <div>
                    <div id="sim-budget-status" class="state" style="margin-top:6px; font-weight:700;">-</div>
                    <div class="label">予算アラート監視状態 (しきい値 $100.00)</div>
                </div>
                <div>
                    <div style="font-size:13px; color:var(--muted);">
                        • vCPU: 2個 / メモリ: 8GB 固定<br>
                        • 予算監視サーキット: 有効 (GCP Billing Alert)<br>
                        • Express Mode 評価: 90日間 (課金制限)
                    </div>
                </div>
            </div>
            <div id="sim-breakdown" style="font-size:12px; margin-top:14px; color:var(--muted); background:#0c0d0c; border:1px solid var(--line); border-radius:5px; padding:10px; display:grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap:8px;">
            </div>
        </section>
        <section class="card">
            <h2>Recent Events</h2>
            <div class="muted" style="margin-bottom:10px;">Prompts and API keys are not stored. Signed video URLs are not stored in this ledger.</div>
            <table>
                <thead><tr><th>Time</th><th>Provider</th><th>Operation</th><th>Outcome</th><th>Billable</th><th>Tokens</th><th>Task</th><th>Reason</th></tr></thead>
                <tbody id="events"></tbody>
            </table>
        </section>
        <section class="card muted" id="note"></section>
    </main>
    <script>
        const fmt = (value) => value === null || value === undefined ? "-" : value;
        function stateClass(value) { return value === "closed" ? "closed" : "open"; }
        async function loadUsage() {
            const response = await fetch("/api/admin/usage");
            const data = await response.json();
            const seedance = data.providers.seedance_api;
            const gemini = data.providers.gemini_api;
            document.getElementById("cards").innerHTML = `
                <div class="card"><div class="metric">${seedance.today_billable_calls}</div><div class="label">Seedance billable calls today</div></div>
                <div class="card"><div class="metric">${gemini.today_billable_calls}</div><div class="label">Gemini billable calls today</div></div>
                <div class="card"><div class="metric">${seedance.today_reported_total_tokens + gemini.today_reported_total_tokens}</div><div class="label">Provider-reported tokens today</div></div>
            `;
            const breakerRows = Object.entries(data.circuit_breakers).map(([name, breaker]) => `
                <tr>
                    <td><code>${name}</code></td>
                    <td><span class="state ${stateClass(breaker.state || (breaker.enabled ? "closed" : "open"))}">${breaker.state || (breaker.enabled ? "enabled" : "disabled")}</span></td>
                    <td>${fmt(breaker.today_billable_calls)} / ${fmt(breaker.daily_call_limit)}</td>
                    <td>${fmt(breaker.today_reported_total_tokens)} / ${fmt(breaker.daily_reported_token_limit)}</td>
                </tr>
            `).join("");
            document.getElementById("breakers").innerHTML = `<table><thead><tr><th>Name</th><th>State</th><th>Calls</th><th>Reported tokens</th></tr></thead><tbody>${breakerRows}</tbody></table>`;
            const saved = data.seedance_saved_default || {};
            document.getElementById("saved-video").innerHTML = `
                <div>Provider: <code>${fmt(saved.provider)}</code></div>
                <div>Model: <code>${fmt(saved.model)}</code></div>
                <div>Task: <code>${fmt(saved.task_id)}</code></div>
                <div>Video: <a href="/${fmt(saved.video)}" download>${fmt(saved.video)}</a></div>
                <div>Backup: <code>${fmt(saved.backup_video)}</code></div>
            `;
            document.getElementById("events").innerHTML = data.recent_events.map((event) => `
                <tr>
                    <td>${fmt(event.timestamp)}</td>
                    <td>${fmt(event.provider)}</td>
                    <td>${fmt(event.operation)}</td>
                    <td>${fmt(event.outcome)}</td>
                    <td>${event.billable ? "yes" : "no"}</td>
                    <td>${fmt(event.reported_total_tokens)}</td>
                    <td><code>${fmt(event.task_id)}</code></td>
                    <td>${fmt(event.reason)}</td>
                </tr>
            `).join("") || `<tr><td colspan="8" class="muted">No local usage events yet.</td></tr>`;
            document.getElementById("note").textContent = data.usage_note;
            updateSim();
        }
        async function updateSim() {
            const hours = document.getElementById("sim-hours").value || 0;
            const sessions = document.getElementById("sim-sessions").value || 0;
            const queries = document.getElementById("sim-queries").value || 0;
            const input = document.getElementById("sim-input-tokens").value || 0;
            const output = document.getElementById("sim-output-tokens").value || 0;
            
            const response = await fetch(`/api/admin/managed-agents/cost-simulation?hours=${hours}&sessions=${sessions}&queries=${queries}&input_tokens_million=${input}&output_tokens_million=${output}`);
            const data = await response.json();
            
            document.getElementById("sim-total").textContent = `$${data.total_cost.toFixed(2)}`;
            
            const statusEl = document.getElementById("sim-budget-status");
            statusEl.textContent = data.monitoring.budget_state.toUpperCase();
            statusEl.className = "state " + (data.monitoring.budget_state === "healthy" ? "closed" : "open");
            
            const bd = data.breakdown;
            document.getElementById("sim-breakdown").innerHTML = `
                <div>• vCPU コンピュート費: $${bd.vcpu_cost.toFixed(2)}</div>
                <div>• メモリ (8GB) リソース費: $${bd.memory_cost.toFixed(2)}</div>
                <div>• セッション履歴維持費 (1k): $${bd.session_cost.toFixed(2)}</div>
                <div>• RAG 検索 (Vertex AI Search) 費: $${bd.search_cost.toFixed(2)}</div>
                <div>• Gemini 入力トークン費: $${bd.gemini_input_cost.toFixed(2)}</div>
                <div>• Gemini 出力トークン費: $${bd.gemini_output_cost.toFixed(2)}</div>
            `;
        }
        loadUsage();
    </script>
</body>
</html>"""


@app.get("/admin", response_class=HTMLResponse)
async def admin_dashboard(username: str = Depends(verify_credentials)):
    """Local-only external API usage and circuit-breaker dashboard."""
    return HTMLResponse(content=ADMIN_DASHBOARD_HTML)


@app.get("/api/admin/usage")
async def admin_usage(username: str = Depends(verify_credentials)):
    return build_external_api_usage_summary()


@app.get("/admin/usage")
async def admin_usage_alias(username: str = Depends(verify_credentials)):
    """Human-friendly alias for users who type /admin/usage in the browser."""
    return build_external_api_usage_summary()


@app.get("/api/admin/usage/export", response_class=PlainTextResponse)
async def admin_usage_export(username: str = Depends(verify_credentials)):
    if not os.path.exists(EXTERNAL_API_USAGE_LOG_FILE):
        return PlainTextResponse("", media_type="application/jsonl")
    with open(EXTERNAL_API_USAGE_LOG_FILE, "r", encoding="utf-8") as f:
        return PlainTextResponse(f.read(), media_type="application/jsonl")


@app.get("/api/admin/operations-dashboard")
async def admin_operations_dashboard(limit: int = 20, username: str = Depends(verify_credentials)):
    """Authenticated T842/T800 dashboard summary across operations and usage analytics."""
    return {
        "status": "success",
        "viewer": username,
        **build_operations_dashboard_summary(limit=limit),
    }


@app.get("/api/admin/operations-dashboard/report.csv", response_class=PlainTextResponse)
async def admin_operations_dashboard_report(limit: int = 20, username: str = Depends(verify_credentials)):
    """CSV export for the T842/T800 admin dashboard without raw identifiers or source files."""
    summary = build_operations_dashboard_summary(limit=limit)
    csv_text = build_operations_dashboard_csv(summary)
    return PlainTextResponse(
        csv_text,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="mighty-operations-dashboard.csv"'},
    )


@app.get("/api/admin/managed-agents/cost-simulation")
async def managed_agents_cost_simulation(
    hours: float = 20.0,
    sessions: int = 10000,
    queries: int = 5000,
    input_tokens_million: float = 10.0,
    output_tokens_million: float = 2.0,
    username: str = Depends(verify_credentials)
):
    """
    Managed Agents (Vertex AI Agent Builder) cost estimation simulator.
    Provides detailed monthly costs based on standard Google Cloud billing dimensions.
    """
    VCPU_PRICE_PER_HOUR = 0.0864
    MEMORY_PRICE_PER_GB_HOUR = 0.0090
    SESSION_PRICE_PER_1K = 0.25
    SEARCH_PRICE_PER_1K = 4.00
    GEMINI_INPUT_PER_1M = 1.25
    GEMINI_OUTPUT_PER_1M = 3.75
    
    vcpu_hours = 2 * hours
    memory_gb_hours = 8 * hours
    
    vcpu_cost = vcpu_hours * VCPU_PRICE_PER_HOUR
    memory_cost = memory_gb_hours * MEMORY_PRICE_PER_GB_HOUR
    session_cost = (sessions / 1000) * SESSION_PRICE_PER_1K
    search_cost = (queries / 1000) * SEARCH_PRICE_PER_1K
    gemini_input_cost = input_tokens_million * GEMINI_INPUT_PER_1M
    gemini_output_cost = output_tokens_million * GEMINI_OUTPUT_PER_1M
    
    total_cost = vcpu_cost + memory_cost + session_cost + search_cost + gemini_input_cost + gemini_output_cost
    
    daily_budget = 5.00
    monthly_budget = 100.00
    
    budget_state = "healthy"
    if total_cost > monthly_budget:
        budget_state = "exceeded"
    elif total_cost > (monthly_budget * 0.8):
        budget_state = "warning"
        
    return {
        "status": "success",
        "parameters": {
            "monthly_hours": hours,
            "monthly_sessions": sessions,
            "monthly_queries": queries,
            "input_tokens_million": input_tokens_million,
            "output_tokens_million": output_tokens_million,
            "vcpu_count": 2,
            "memory_gb": 8
        },
        "breakdown": {
            "vcpu_cost": round(vcpu_cost, 2),
            "memory_cost": round(memory_cost, 2),
            "session_cost": round(session_cost, 2),
            "search_cost": round(search_cost, 2),
            "gemini_input_cost": round(gemini_input_cost, 2),
            "gemini_output_cost": round(gemini_output_cost, 2),
        },
        "total_cost": round(total_cost, 2),
        "currency": "USD",
        "monitoring": {
            "budget_state": budget_state,
            "daily_limit_usd": daily_budget,
            "monthly_limit_usd": monthly_budget,
            "gcp_billing_alerts_enabled": True
        }
    }


def read_knowledge_flow_manifest() -> Optional[dict]:
    if not os.path.exists(KNOWLEDGE_FLOW_MANIFEST):
        return None
    with open(KNOWLEDGE_FLOW_MANIFEST, "r", encoding="utf-8") as f:
        return json.load(f)


@app.get("/api/knowledge-flow/status")
async def knowledge_flow_status():
    """Returns current generated NotebookLM/Slack/Notion/Obsidian demo artifacts."""
    manifest = read_knowledge_flow_manifest()
    return {
        "status": "ready" if manifest else "not_generated",
        "output_dir": os.path.relpath(KNOWLEDGE_FLOW_DIR, PROJECT_ROOT),
        "manifest": manifest,
    }


@app.post("/api/knowledge-flow/generate")
async def generate_knowledge_flow_artifacts():
    """Generates safe, CEO-facing knowledge-flow demo artifacts locally."""
    if not os.path.exists(KNOWLEDGE_FLOW_SCRIPT):
        raise HTTPException(status_code=404, detail="Knowledge flow generator script not found.")

    result = subprocess.run(
        [sys.executable, KNOWLEDGE_FLOW_SCRIPT],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=60,
        check=False,
    )
    if result.returncode != 0:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Knowledge flow artifact generation failed.",
                "stdout": result.stdout,
                "stderr": result.stderr,
            },
        )

    manifest = read_knowledge_flow_manifest()
    return {
        "status": "success",
        "message": "NotebookLM / Slack / Notion / Obsidian demo artifacts generated.",
        "stdout": result.stdout,
        "manifest": manifest,
    }


@app.post("/api/seedance/video-demo")
async def generate_seedance_video(req: SeedanceVideoRequest):
    """Create a Seedance video when credentials are set, otherwise return a safe local preview."""
    prompt = normalize_text(req.prompt)
    if not prompt:
        raise HTTPException(status_code=400, detail="prompt is required.")

    if not seedance_demo_video_url():
        return {
            "status": "error",
            "mode": "missing_fallback_asset",
            "provider": "local_seedance_demo_asset",
            "message": "Run scripts/generate_seedance_demo_video.py before launching the demo.",
        }

    if not SEEDANCE_READY:
        reason = "Seedance API billing calls are disabled by default. Set SEEDANCE_API_ENABLED=1 before starting FastAPI to generate a new video."
        if not SEEDANCE_CONFIGURED:
            reason = "SEEDANCE_API_KEY and SEEDANCE_API_URL are not configured."
        append_external_api_event({
            "provider": "seedance_api",
            "operation": "generation_create",
            "billable": False,
            "outcome": "blocked",
            "model": SEEDANCE_MODEL,
            "reason": reason,
            "prompt_digest": stable_digest(prompt),
        })
        return seedance_fallback_response(
            reason,
            prompt,
        )

    allowed, circuit_reason, _stats = check_external_api_circuit(
        "seedance_api",
        "generation_create",
        SEEDANCE_DAILY_GENERATION_LIMIT,
        SEEDANCE_DAILY_REPORTED_TOKEN_LIMIT,
    )
    if not allowed:
        append_external_api_event({
            "provider": "seedance_api",
            "operation": "generation_create",
            "billable": False,
            "outcome": "blocked",
            "model": SEEDANCE_MODEL,
            "reason": circuit_reason,
            "prompt_digest": stable_digest(prompt),
        })
        return seedance_fallback_response(circuit_reason, prompt)

    headers = {
        "Authorization": f"Bearer {SEEDANCE_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = build_seedance_payload(prompt, req)

    try:
        create_response = requests.post(
            SEEDANCE_API_URL,
            headers=headers,
            json=payload,
            timeout=45,
        )
        if not create_response.ok:
            append_external_api_event({
                "provider": "seedance_api",
                "operation": "generation_create",
                "billable": True,
                "outcome": "http_error",
                "model": SEEDANCE_MODEL,
                "http_status": create_response.status_code,
                "reason": summarize_seedance_http_error(create_response),
                "prompt_digest": stable_digest(prompt),
            })
            return seedance_fallback_response(
                f"Seedance API request failed: {summarize_seedance_http_error(create_response)}",
                prompt,
            )
        create_payload = create_response.json()
        usage = find_token_usage(create_payload)
        video_url = find_video_url(create_payload)
        task_id = find_task_id(create_payload)

        if video_url:
            append_external_api_event({
                "provider": "seedance_api",
                "operation": "generation_create",
                "billable": True,
                "outcome": "live",
                "model": SEEDANCE_MODEL,
                "task_id": task_id,
                "http_status": create_response.status_code,
                "prompt_digest": stable_digest(prompt),
                **usage,
            })
            return {
                "status": "success",
                "mode": "live",
                "provider": "seedance_api",
                "model": SEEDANCE_MODEL,
                "video_url": video_url,
                "task_id": task_id,
                "raw_status": create_payload.get("status") if isinstance(create_payload, dict) else None,
            }

        if task_id:
            append_external_api_event({
                "provider": "seedance_api",
                "operation": "generation_create",
                "billable": True,
                "outcome": "task_created",
                "model": SEEDANCE_MODEL,
                "task_id": task_id,
                "http_status": create_response.status_code,
                "prompt_digest": stable_digest(prompt),
                **usage,
            })
            video_url, result_payload, poll_reason = poll_seedance_result(task_id, headers)
            if video_url:
                return {
                    "status": "success",
                    "mode": "live",
                    "provider": "seedance_api",
                    "model": SEEDANCE_MODEL,
                    "video_url": video_url,
                    "task_id": task_id,
                    "raw_status": result_payload.get("status") if isinstance(result_payload, dict) else None,
                }
            return seedance_pending_response(
                poll_reason,
                prompt,
                task_id,
                find_task_status(result_payload),
            )

        return seedance_fallback_response(
            "Seedance task was accepted but no video URL was returned yet.",
            prompt,
            task_id,
        )
    except requests.RequestException as exc:
        append_external_api_event({
            "provider": "seedance_api",
            "operation": "generation_create",
            "billable": False,
            "outcome": "request_exception",
            "model": SEEDANCE_MODEL,
            "reason": str(exc),
            "prompt_digest": stable_digest(prompt),
        })
        return seedance_fallback_response(f"Seedance API request failed: {exc}", prompt)
    except ValueError as exc:
        return seedance_fallback_response(f"Seedance API returned non-JSON response: {exc}", prompt)


@app.get("/api/seedance/video-task/{task_id}")
async def get_seedance_video_task(task_id: str):
    """Check an existing Seedance task once so the browser can continue polling."""
    if not SEEDANCE_READY:
        reason = "Seedance API billing calls are disabled by default. Set SEEDANCE_API_ENABLED=1 before starting FastAPI to poll a remote task."
        if not SEEDANCE_CONFIGURED:
            reason = "SEEDANCE_API_KEY and SEEDANCE_API_URL are not configured."
        append_external_api_event({
            "provider": "seedance_api",
            "operation": "task_poll",
            "billable": False,
            "outcome": "blocked",
            "model": SEEDANCE_MODEL,
            "task_id": task_id,
            "reason": reason,
        })
        return seedance_fallback_response(
            reason,
            task_id,
            task_id,
        )
    if not SEEDANCE_RESULT_API_URL_TEMPLATE:
        return seedance_pending_response(
            "SEEDANCE_RESULT_API_URL_TEMPLATE is not configured for result polling.",
            task_id,
            task_id,
            "unknown",
        )

    headers = {
        "Authorization": f"Bearer {SEEDANCE_API_KEY}",
        "Content-Type": "application/json",
    }
    result_url = SEEDANCE_RESULT_API_URL_TEMPLATE.format(task_id=task_id)
    try:
        result_response = requests.get(result_url, headers=headers, timeout=45)
        if not result_response.ok:
            append_external_api_event({
                "provider": "seedance_api",
                "operation": "task_poll",
                "billable": False,
                "outcome": "http_error",
                "model": SEEDANCE_MODEL,
                "task_id": task_id,
                "http_status": result_response.status_code,
                "reason": summarize_seedance_http_error(result_response),
            })
            return seedance_fallback_response(
                f"Seedance result request failed: {summarize_seedance_http_error(result_response)}",
                task_id,
                task_id,
            )
        result_payload = result_response.json()
        usage = find_token_usage(result_payload)
        video_url = find_video_url(result_payload)
        task_status = find_task_status(result_payload)
        if video_url:
            append_external_api_event({
                "provider": "seedance_api",
                "operation": "task_poll",
                "billable": False,
                "outcome": "live",
                "model": SEEDANCE_MODEL,
                "task_id": task_id,
                "http_status": result_response.status_code,
                **usage,
            })
            return {
                "status": "success",
                "mode": "live",
                "provider": "seedance_api",
                "model": SEEDANCE_MODEL,
                "video_url": video_url,
                "task_id": task_id,
                "task_status": task_status,
            }
        append_external_api_event({
            "provider": "seedance_api",
            "operation": "task_poll",
            "billable": False,
            "outcome": "pending",
            "model": SEEDANCE_MODEL,
            "task_id": task_id,
            "http_status": result_response.status_code,
            "reason": f"task_status={task_status or 'unknown'}",
            **usage,
        })
        return seedance_pending_response(
            "Seedance task is still running.",
            task_id,
            task_id,
            task_status,
        )
    except requests.RequestException as exc:
        return seedance_fallback_response(f"Seedance result request failed: {exc}", task_id, task_id)
    except ValueError as exc:
        return seedance_fallback_response(f"Seedance result returned non-JSON response: {exc}", task_id, task_id)


# 1. API: Multi-modal Resume/Job Parser
@app.post("/api/parse")
async def parse_document(
    file: UploadFile = File(None),
    text: str = Form(None),
    doc_type: str = Form("engineer"), # "engineer" or "job"
    legal_consent_accepted: bool = Form(False),
    legal_consent_version: str = Form(LEGAL_CONSENT_VERSION),
):
    """Parses text or binary files (PDF/Images) and structure them."""
    print(f"[*] API Parse Request received. Type: {doc_type}")
    legal_consent = validate_legal_consent(
        legal_consent_accepted,
        legal_consent_version,
        "api_parse",
    )
    
    file_bytes = None
    file_name = None
    file_type = None
    
    if file:
        file_bytes = await file.read()
        file_name = file.filename
        file_type = file.content_type
        print(f"  - File uploaded: {file_name} ({file_type}, {len(file_bytes)} bytes)")
    elif text:
        print(f"  - Direct text input received ({len(text)} characters)")
    else:
        raise HTTPException(status_code=400, detail="Either file or text must be provided.")

    source_text = text or decode_uploaded_text(file_bytes or b"")
    fallback_reason = "Gemini live mode is not configured."

    # --- Live Gemini Parsing Logic ---
    if GEMINI_READY:
        try:
            allowed, circuit_reason, _stats = check_external_api_circuit(
                "gemini_api",
                "parse",
                GEMINI_DAILY_CALL_LIMIT,
                GEMINI_DAILY_REPORTED_TOKEN_LIMIT,
            )
            if not allowed:
                append_external_api_event({
                    "provider": "gemini_api",
                    "operation": "parse",
                    "billable": False,
                    "outcome": "blocked",
                    "model": GEMINI_MODEL,
                    "reason": circuit_reason,
                    "prompt_digest": stable_digest(source_text),
                })
                raise RuntimeError(circuit_reason)
            local_profile = build_profile(source_text, doc_type)
            
            prompt = (
                f"You are a professional HR data extraction engine.\n"
                f"Parse the following {doc_type} document and extract a clean structured text summary in Japanese.\n"
                f"Focus on extracting: Name (if engineer), Job Title / Role Name, Core Skills, Cloud / Infra, Databases, and Career Goals.\n"
                f"Make sure to output clean Japanese, keeping important technical details.\n"
                f"Use this local deterministic pre-parse as hints, but correct it when the document says otherwise:\n"
                f"{json.dumps(asdict(local_profile), ensure_ascii=False)}"
            )
            
            response = None
            if file_bytes:
                # Multimodal API Input (PDF/Images)
                pdf_part = genai_types.Part.from_bytes(
                    data=file_bytes,
                    mime_type=file_type or "application/octet-stream",
                )
                response = generate_gemini_content([pdf_part, prompt])
            else:
                response = generate_gemini_content(f"{prompt}\n\nDocument Content:\n{text}")
                
            parsed_text = response.text.strip()
            append_external_api_event({
                "provider": "gemini_api",
                "operation": "parse",
                "billable": True,
                "outcome": "success",
                "model": GEMINI_MODEL,
                "prompt_digest": stable_digest(source_text),
                **find_token_usage(getattr(response, "usage_metadata", None)),
            })
            print(f"[+] Gemini Parser Sync completed successfully.")
            audit_event = write_audit_event(
                "parse",
                profile_audit_payload(local_profile, "gemini_live", "", source_text, file_name)
            )
            # Save to database
            db_id = 0
            if doc_type == "engineer":
                parsed_skills = local_profile.skills_by_category
                career_goals = {"strengths": local_profile.strengths, "risk_flags": local_profile.risk_flags}
                db_id = db_insert_engineer(local_profile.title, source_text, parsed_skills, career_goals)
            else:
                parsed_requirements = {"mandatory": local_profile.all_skills, "preferred": []}
                company_culture = {"summary": local_profile.summary}
                db_id = db_insert_job(local_profile.title, "Mighty-Link", source_text, parsed_requirements, company_culture)

            return {
                "status": "success",
                "ai_mode": "gemini_live",
                "parsed_content": parsed_text,
                "structured_profile": asdict(local_profile),
                "audit_event_id": audit_event["event_id"],
                "db_id": db_id,
                "legal_consent": legal_consent,
            }
            
        except Exception as e:
            fallback_reason = str(e)
            if "daily" not in fallback_reason.lower():
                append_external_api_event({
                    "provider": "gemini_api",
                    "operation": "parse",
                    "billable": False,
                    "outcome": "exception",
                    "model": GEMINI_MODEL,
                    "reason": fallback_reason,
                    "prompt_digest": stable_digest(source_text),
                })
            print(f"[-] Gemini live parser failed: {e}. Falling back to deterministic parser.")
    elif AI_FORCE_MOCK:
        fallback_reason = "AI_FORCE_MOCK is enabled to avoid Gemini quota usage."

    # --- Quota-safe deterministic parser fallback ---
    if DETERMINISTIC_PARSE_DELAY_SECONDS > 0:
        await asyncio.sleep(DETERMINISTIC_PARSE_DELAY_SECONDS)

    profile = await asyncio.to_thread(build_profile, source_text, doc_type)
    parsed_content = await asyncio.to_thread(format_profile, profile)
    audit_payload = profile_audit_payload(profile, "deterministic_fallback", fallback_reason, source_text, file_name)
    audit_event = await asyncio.to_thread(
        write_audit_event,
        "parse",
        audit_payload,
    )
    
    # Save to database
    db_id = 0
    if doc_type == "engineer":
        parsed_skills = profile.skills_by_category
        career_goals = {"strengths": profile.strengths, "risk_flags": profile.risk_flags}
        db_id = await asyncio.to_thread(db_insert_engineer, profile.title, source_text, parsed_skills, career_goals)
    else:
        parsed_requirements = {"mandatory": profile.all_skills, "preferred": []}
        company_culture = {"summary": profile.summary}
        db_id = await asyncio.to_thread(
            db_insert_job,
            profile.title,
            "Mighty-Link",
            source_text,
            parsed_requirements,
            company_culture,
        )
        
    print(f"[+] Deterministic parser fallback completed successfully.")
    return {
        "status": "success",
        "ai_mode": "deterministic_fallback",
        "fallback_reason": fallback_reason,
        "parsed_content": parsed_content,
        "structured_profile": asdict(profile),
        "audit_event_id": audit_event["event_id"],
        "db_id": db_id,
        "legal_consent": legal_consent,
    }


class EvaluationRequest(BaseModel):
    engineer_content: str
    job_content: str
    legal_consent_accepted: bool = False
    legal_consent_version: str = LEGAL_CONSENT_VERSION


class FeedbackRequest(BaseModel):
    match_id: Optional[int] = None
    rating: str
    nps_score: Optional[int] = None
    comment: Optional[str] = ""
    source: str = "diagnosis_report"
    page_url: Optional[str] = ""
    session_id: Optional[str] = ""


class AptitudeDemoQuestionRequest(BaseModel):
    # T876: session-only aptitude/motivation self-check. No answers here — this
    # only requests the question set. count is clamped to [10, 20].
    count: Optional[int] = None
    legal_consent_accepted: bool = False
    legal_consent_version: Optional[str] = None


class AptitudeDemoAnswer(BaseModel):
    dimension: Optional[str] = "general"
    value: int


class AptitudeDemoEvaluateRequest(BaseModel):
    answers: List[AptitudeDemoAnswer] = Field(default_factory=list)
    consented: bool = False
    legal_consent_accepted: bool = False
    legal_consent_version: Optional[str] = None


class UsageAnalyticsEventRequest(BaseModel):
    event_name: str
    event_surface: str = "public_demo"
    page_url: Optional[str] = ""
    session_id: Optional[str] = ""
    metadata: Optional[Dict[str, Any]] = None


class SalesEmailMatchReviewRequest(BaseModel):
    match_key: Optional[str] = ""
    project_key: Optional[str] = ""
    talent_key: Optional[str] = ""
    feedback_status: str
    corrected_score: Optional[float] = None
    corrected_notes: Optional[str] = ""
    corrected_fields: Optional[Dict[str, Any]] = None
    next_action: Optional[str] = ""


class EmployeeAssessmentResponseRequest(BaseModel):
    employee_identifier: str
    department: str
    motivation_level: int
    culture_level: int
    growth_feedback: str
    consented: bool
    consent_version: str = EMPLOYEE_ASSESSMENT_CONSENT_VERSION
    source: str = "employee_assessment_form"
    page_url: Optional[str] = ""
    session_id: Optional[str] = ""


class OnboardingProgressRequest(BaseModel):
    completed_step_ids: List[str] = []


class OnboardingActivateRequest(BaseModel):
    account_identifier: str
    completed_step_ids: List[str] = []
    legal_consent_accepted: bool = False
    legal_consent_version: str = ""
    source: str = "onboarding_wizard"
    page_url: Optional[str] = ""
    session_id: Optional[str] = ""


class AttendancePunchRequest(BaseModel):
    employee_identifier: str
    event_type: str
    consented: bool
    source: str = "attendance_widget"
    page_url: Optional[str] = ""
    session_id: Optional[str] = ""


class AttendanceTimesheetApprovalRequest(BaseModel):
    import_id: int
    decision: str = "approved"


class SupportRequest(BaseModel):
    category: str = "general"
    priority: Optional[str] = None
    contact_email: str
    subject: str
    message: str
    source: str = "support_form"
    page_url: Optional[str] = ""
    session_id: Optional[str] = ""


# 2. API: 4-Dimension AI Fit Evaluation
@app.post("/api/match")
async def evaluate_matching(req: EvaluationRequest):
    """Calculates multidimensional matching score and generate interview QA and roadmap."""
    print("[*] API Match Request received.")
    legal_consent = validate_legal_consent(
        req.legal_consent_accepted,
        req.legal_consent_version,
        "api_match",
    )
    
    fallback_reason = "Gemini live mode is not configured."

    # --- Live Gemini Evaluation Logic ---
    if GEMINI_READY:
        try:
            allowed, circuit_reason, _stats = check_external_api_circuit(
                "gemini_api",
                "match",
                GEMINI_DAILY_CALL_LIMIT,
                GEMINI_DAILY_REPORTED_TOKEN_LIMIT,
            )
            if not allowed:
                append_external_api_event({
                    "provider": "gemini_api",
                    "operation": "match",
                    "billable": False,
                    "outcome": "blocked",
                    "model": GEMINI_MODEL,
                    "reason": circuit_reason,
                    "prompt_digest": stable_digest(req.engineer_content + req.job_content),
                })
                raise RuntimeError(circuit_reason)
            fallback_context = build_fallback_match(
                req.engineer_content,
                req.job_content,
                "local deterministic pre-score for Gemini prompt context"
            )
            
            prompt = (
                "You are the Mighty-Link AI engine. Evaluate the fit between the Candidate Resume and the Job Description.\n"
                "You MUST return the response strictly as a JSON object with the following fields:\n"
                "{\n"
                "  \"final_score\": <Integer from 50 to 100>,\n"
                "  \"scores\": {\n"
                "    \"skill\": <Integer 50-100>,\n"
                "    \"culture\": <Integer 50-100>,\n"
                "    \"growth\": <Integer 50-100>,\n"
                "    \"performing\": <Integer 50-100>\n"
                "  },\n"
                "  \"summary\": \"<Detailed multi-dimensional evaluation summary paragraph in Japanese>\",\n"
                "  \"qa\": [\n"
                "    {\n"
                "      \"question\": \"<Technical interview question tailored for this specific match in Japanese>\",\n"
                "      \"answer\": \"<Best practice guide for candidate answer in Japanese>\",\n"
                "      \"tip\": \"<Tips to enhance points in Japanese>\"\n"
                "    },\n"
                "    {\n"
                "      \"question\": \"<Another relevant interview question in Japanese>\",\n"
                "      \"answer\": \"<Answer guide in Japanese>\",\n"
                "      \"tip\": \"<Tips in Japanese>\"\n"
                "    }\n"
                "  ],\n"
                "  \"roadmap_week1\": \"<Detailed week 1 roadmap actions in Japanese>\",\n"
                "  \"roadmap_week2\": \"<Detailed week 2 roadmap actions in Japanese>\",\n"
                "  \"roadmap_week3\": \"<Detailed week 3 roadmap actions in Japanese>\",\n"
                "  \"roadmap_week4\": \"<Detailed week 4 roadmap actions in Japanese>\"\n"
                "}\n"
                "Do NOT include any markdown code blocks (like ```json) or explanation text outside the JSON.\n\n"
                "Use this deterministic local analysis as structured context. Correct it when the source text provides better evidence:\n"
                f"{json.dumps(fallback_context.get('structured', {}), ensure_ascii=False)}\n\n"
                f"Candidate Resume Data:\n{req.engineer_content}\n\n"
                f"Job Description Data:\n{req.job_content}"
            )
            
            response = generate_gemini_content(
                prompt,
                response_mime_type="application/json"
            )
            
            res_text = response.text.strip()
            # Clean possible raw markdown block if LLM fails strict config
            if res_text.startswith("```"):
                res_text = res_text.split("\n", 1)[1].rsplit("\n", 1)[0]
                if res_text.startswith("json"):
                    res_text = res_text.split("\n", 1)[1]
            
            match_data = json.loads(res_text.strip())
            match_data["ai_mode"] = "gemini_live"
            match_data.setdefault("structured", fallback_context.get("structured", {}))
            append_external_api_event({
                "provider": "gemini_api",
                "operation": "match",
                "billable": True,
                "outcome": "success",
                "model": GEMINI_MODEL,
                "prompt_digest": stable_digest(req.engineer_content + req.job_content),
                **find_token_usage(getattr(response, "usage_metadata", None)),
            })
            # Resolve database records and save match result
            eng_id = resolve_or_insert_engineer(req.engineer_content)
            jb_id = resolve_or_insert_job(req.job_content)
            
            fit_ratio = float(match_data.get("final_score", 75)) / 100.0
            scores = match_data.get("scores", {})
            score_skill = scores.get("skill", 75)
            score_culture = scores.get("culture", 75)
            score_growth = scores.get("growth", 75)
            score_performing = scores.get("performing", 75)
            match_summary = match_data.get("summary", "")
            interview_questions = match_data.get("qa", [])
            
            db_match_id = db_insert_match_result(
                eng_id, jb_id, fit_ratio, score_skill, score_culture, score_growth, score_performing, match_summary, interview_questions
            )
            match_data["db_match_id"] = db_match_id
            match_data["legal_consent"] = legal_consent

            print("[+] Gemini Evaluator completed successfully.")
            return match_data
            
        except Exception as e:
            fallback_reason = str(e)
            if "daily" not in fallback_reason.lower():
                append_external_api_event({
                    "provider": "gemini_api",
                    "operation": "match",
                    "billable": False,
                    "outcome": "exception",
                    "model": GEMINI_MODEL,
                    "reason": fallback_reason,
                    "prompt_digest": stable_digest(req.engineer_content + req.job_content),
                })
            print(f"[-] Gemini live evaluation failed: {e}. Falling back to deterministic evaluator.")
    elif AI_FORCE_MOCK:
        fallback_reason = "AI_FORCE_MOCK is enabled to avoid Gemini quota usage."

    # --- Quota-safe deterministic evaluator fallback ---
    if DETERMINISTIC_MATCH_DELAY_SECONDS > 0:
        await asyncio.sleep(DETERMINISTIC_MATCH_DELAY_SECONDS)

    fallback_response = await asyncio.to_thread(
        build_fallback_match,
        req.engineer_content,
        req.job_content,
        fallback_reason,
    )
    audit_payload = match_audit_payload(fallback_response)
    audit_event = await asyncio.to_thread(write_audit_event, "match", audit_payload)
    fallback_response["audit_event_id"] = audit_event["event_id"]
    
    # Resolve database records and save match result
    eng_id, jb_id = await asyncio.gather(
        asyncio.to_thread(resolve_or_insert_engineer, req.engineer_content),
        asyncio.to_thread(resolve_or_insert_job, req.job_content),
    )
    
    fit_ratio = float(fallback_response.get("final_score", 75)) / 100.0
    scores = fallback_response.get("scores", {})
    score_skill = scores.get("skill", 75)
    score_culture = scores.get("culture", 75)
    score_growth = scores.get("growth", 75)
    score_performing = scores.get("performing", 75)
    match_summary = fallback_response.get("summary", "")
    interview_questions = fallback_response.get("qa", [])
    
    db_match_id = await asyncio.to_thread(
        db_insert_match_result,
        eng_id,
        jb_id,
        fit_ratio,
        score_skill,
        score_culture,
        score_growth,
        score_performing,
        match_summary,
        interview_questions,
    )
    fallback_response["db_match_id"] = db_match_id
    fallback_response["legal_consent"] = legal_consent
    
    print("[+] Deterministic evaluator fallback completed successfully.")
    return fallback_response


def load_extraction_report_from_postgres() -> Optional[dict]:
    """Helper to fetch and rebuild the extraction report dict directly from PostgreSQL DB."""
    db_url = os.environ.get("SUPABASE_DB_URL")
    if not db_url:
        return None
    try:
        import psycopg2
        from psycopg2.extras import RealDictCursor
        
        conn = psycopg2.connect(db_url)
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM sales_email_messages WHERE ingest_status = 'parsed';")
        messages = [dict(r) for r in cur.fetchall()]
        
        cur.execute("SELECT * FROM project_requirements;")
        projects = {p["message_id"]: dict(p) for p in cur.fetchall()}
        
        cur.execute("SELECT * FROM talent_profiles_from_email;")
        talents = {t["message_id"]: dict(t) for t in cur.fetchall()}
        
        cur.close()
        conn.close()
        
        extractions = []
        project_count = 0
        talent_count = 0
        skill_tag_count = 0
        
        for msg in messages:
            msg_id = msg["id"]
            proj = projects.get(msg_id)
            tal = talents.get(msg_id)
            
            email_kind = "unknown"
            proj_data = None
            if proj:
                email_kind = "project"
                project_count += 1
                skills_list = proj.get("required_skills", [])
                if isinstance(skills_list, str):
                    try:
                        skills_list = json.loads(skills_list)
                    except Exception:
                        skills_list = []
                skill_tag_count += len(skills_list)
                proj_data = {
                    "title": proj["title"],
                    "summary": proj["summary"],
                    "required_skills": skills_list,
                    "nice_to_have_skills": proj.get("nice_to_have_skills") or [],
                    "skill_categories": proj.get("skill_categories") or {},
                    "rate_min": proj["rate_min"],
                    "rate_max": proj["rate_max"],
                    "rate_unit": proj["rate_unit"],
                    "location": proj["location"],
                    "remote_type": proj["remote_type"],
                    "start_date_text": proj["start_date_text"],
                    "duration_text": proj["duration_text"],
                    "commercial_flow": proj["commercial_flow"],
                    "restrictions": proj["restrictions"],
                    "evidence_excerpt": (msg.get("body_excerpt") or "").strip() or proj["evidence_excerpt"],
                    "confidence": 1.0,
                    "review_status": proj["review_status"]
                }
                
            tal_data = None
            if tal:
                email_kind = "talent"
                talent_count += 1
                skills_list = tal.get("skills", [])
                if isinstance(skills_list, str):
                    try:
                        skills_list = json.loads(skills_list)
                    except Exception:
                        skills_list = []
                tal_data = {
                    "anonymized_talent_key": tal["anonymized_talent_key"],
                    "summary": tal["summary"],
                    "skills": skills_list,
                    "skill_categories": tal.get("skill_categories") or {},
                    "experience_years": tal["experience_years"],
                    "desired_rate_min": tal["desired_rate_min"],
                    "desired_rate_max": tal["desired_rate_max"],
                    "desired_location": tal["desired_location"],
                    "remote_preference": tal["remote_preference"],
                    "availability_text": tal["availability_text"],
                    "evidence_excerpt": (msg.get("body_excerpt") or "").strip() or tal["evidence_excerpt"],
                    "confidence": 1.0,
                    "review_status": tal["review_status"]
                }
                
            msg_received_at = "2026-06-18T00:00:00Z"
            if msg.get("received_at"):
                if isinstance(msg["received_at"], datetime):
                    msg_received_at = msg["received_at"].isoformat().replace("+00:00", "Z")
                else:
                    msg_received_at = str(msg["received_at"])
            elif msg.get("created_at"):
                if isinstance(msg["created_at"], datetime):
                    msg_received_at = msg["created_at"].isoformat().replace("+00:00", "Z")
                else:
                    msg_received_at = str(msg["created_at"])

            extractions.append({
                "source_path": msg["source_path"],
                "source_type": msg["source_type"],
                "dedupe_key": msg["dedupe_key"],
                "sender_domain": msg["sender_domain"],
                "normalized_subject": msg["normalized_subject"],
                "received_at": msg_received_at,
                "email_kind": email_kind,
                "model_name": "deterministic-sales-email-extractor-v1",
                "fallback_used": True,
                "project_requirement": proj_data,
                "talent_profile": tal_data
            })
            
        from datetime import datetime, timezone
        return {
            "task_id": "T817_4",
            "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "model_name": "deterministic-sales-email-extractor-v1",
            "fallback_used": True,
            "input_count": len(messages),
            "project_requirement_count": project_count,
            "talent_profile_count": talent_count,
            "skill_tag_count": skill_tag_count,
            "privacy_controls": [
                "raw_email_body_not_written",
                "email_phone_secret_patterns_redacted_from_evidence",
                "sender_hash_or_domain_only",
                "talent_identity_anonymized",
                "human_review_required_before_confirmed_status"
            ],
            "extractions": extractions
        }
    except Exception as exc:
        print(f"[-] PostgreSQL extraction report fetch failed: {exc}")
        return None


@app.get("/api/sales-email/matches")
async def list_sales_email_matches(
    direction: str = "project_to_talent",
    skills: str = "",
    remote: str = "",
    min_score: int = 0,
    limit: int = 20,
    project_key: str = "",
    talent_key: str = "",
    min_rate: Optional[int] = None,
    max_rate: Optional[int] = None,
    search_query: str = "",
    username: Optional[str] = Depends(verify_credentials_optional),
):
    """Return sanitized bidirectional candidate lists from T817_4 extraction output."""
    if criteria_from_values is None:
        raise HTTPException(status_code=503, detail="sales email matching module is unavailable")

    report_data = None
    source_report = "database"
    if os.environ.get("SUPABASE_DB_URL"):
        report_data = load_extraction_report_from_postgres()
        
    if report_data is None:
        report_path = Path(SALES_EMAIL_MATCH_REPORT_FILE)
        if report_path.exists():
            try:
                report_data = json.loads(report_path.read_text(encoding="utf-8"))
                source_report = os.path.relpath(str(report_path), PROJECT_ROOT)
            except Exception:
                pass
                
    if report_data is None:
        from datetime import datetime, timezone
        report_data = {
            "task_id": "T817_4",
            "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "extractions": []
        }

    try:
        criteria = criteria_from_values(
            direction=direction,
            skills=skills,
            remote=remote,
            min_score=min_score,
            limit=limit,
            project_key=project_key,
            talent_key=talent_key,
            min_rate=min_rate,
            max_rate=max_rate,
            search_query=search_query,
        )
        import sys
        sys.path.insert(0, str(Path(PROJECT_ROOT) / "src"))
        from sales_email_match import build_match_report
        report = build_match_report(report_data, criteria)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        print(f"[-] sales email match endpoint failed: {exc}")
        raise HTTPException(status_code=500, detail="sales email match report generation failed") from exc

    return {
        "status": "success",
        "source_report": source_report,
        **report,
    }


@app.get("/api/sales-email/analytics")
async def get_sales_email_analytics():
    """Return aggregated stats from extraction report for public dashboard analytics."""
    report_data = None
    if os.environ.get("SUPABASE_DB_URL"):
        report_data = load_extraction_report_from_postgres()
        
    if report_data is None:
        report_path = Path(SALES_EMAIL_MATCH_REPORT_FILE)
        if report_path.exists():
            try:
                report_data = json.loads(report_path.read_text(encoding="utf-8"))
            except Exception:
                pass
                
    if report_data is None:
        return {
            "status": "success",
            "daily_counts": {},
            "domain_counts": {},
            "skill_counts": {}
        }
    
    try:
        extractions = report_data.get("extractions", [])
        
        daily_counts = {}
        domain_counts = {}
        skill_counts = {}
        
        for item in extractions:
            dt_str = item.get("received_at") or report_data.get("generated_at") or "2026-06-18"
            dt = dt_str[:10]
            daily_counts[dt] = daily_counts.get(dt, 0) + 1
            
            dom = item.get("sender_domain", "unknown")
            domain_counts[dom] = domain_counts.get(dom, 0) + 1
            
            req = item.get("project_requirement")
            if req and isinstance(req, dict):
                skills = req.get("required_skills", [])
                for sk in skills:
                    skill_counts[sk] = skill_counts.get(sk, 0) + 1
                    
            tal = item.get("talent_profile")
            if tal and isinstance(tal, dict):
                skills = tal.get("skills", [])
                for sk in skills:
                    skill_counts[sk] = skill_counts.get(sk, 0) + 1
                    
        return {
            "status": "success",
            "daily_counts": daily_counts,
            "domain_counts": domain_counts,
            "skill_counts": skill_counts
        }
    except Exception as exc:
        print(f"[-] Analytics calculation failed: {exc}")
        return {"status": "error", "message": "Failed to calculate analytics"}


@app.post("/api/sales-email/sync")
async def sync_sales_emails(
    max_messages: int | None = Query(None, description="Maximum POP3 emails to fetch (default: 1000)"),
    retry_errors: bool = Query(False, description="Whether to include ingest_status='error' messages in parsing retry"),
    username: str = Depends(verify_credentials),
):
    """Sync POP3 emails to database, run AI parse pipeline, and rebuild review JSONs."""
    try:
        import sys
        sys.path.insert(0, str(Path(PROJECT_ROOT) / "scripts"))
        from sync_sales_emails import sync_sales_emails_pipeline
        result = sync_sales_emails_pipeline(max_messages=max_messages, retry_errors=retry_errors)
        return result
    except Exception as exc:
        print(f"[-] Sales email sync pipeline failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Sync pipeline failed: {str(exc)}")


@app.post("/api/sales-email/reviews")
async def submit_sales_email_match_review(
    req: SalesEmailMatchReviewRequest,
    username: str = Depends(verify_credentials),
):
    """Store a sanitized human review for a sales email match candidate."""
    required_helpers = [
        build_match_report_from_file,
        criteria_from_values,
        find_sales_email_match_for_review,
        build_sales_email_review_entry,
        sales_email_project_by_key,
        sales_email_talent_by_key,
        load_sales_email_review_report,
        upsert_sales_email_review_entry,
        write_sales_email_review_json,
        write_sales_email_review_markdown,
    ]
    if any(helper is None for helper in required_helpers):
        raise HTTPException(status_code=503, detail="sales email review module is unavailable")

    feedback_status = clean_feedback_text(req.feedback_status, 32).lower()
    if feedback_status not in SALES_EMAIL_REVIEW_STATUSES:
        raise HTTPException(status_code=400, detail="feedback_status must be accepted, rejected, needs_review, or corrected")
    if req.corrected_score is not None and not 0 <= req.corrected_score <= 100:
        raise HTTPException(status_code=400, detail="corrected_score must be between 0 and 100")

    report_path = Path(SALES_EMAIL_MATCH_REPORT_FILE)
    try:
        criteria = criteria_from_values(limit=100)
        report = build_match_report_from_file(report_path, criteria)
        match_row = find_sales_email_match_for_review(
            report,
            wanted_match_key=clean_feedback_text(req.match_key, 120),
            project_key=clean_feedback_text(req.project_key, 120),
            talent_key=clean_feedback_text(req.talent_key, 120),
        )
        review_entry = build_sales_email_review_entry(
            match_row,
            feedback_status=feedback_status,
            reviewer_id=username,
            corrected_score=req.corrected_score,
            corrected_notes=req.corrected_notes or "",
            corrected_fields=req.corrected_fields or {},
            next_action=req.next_action or "",
        )
        project = sales_email_project_by_key(report, review_entry["project_key"])
        talent = sales_email_talent_by_key(report, review_entry["talent_key"])
        db_result = db_insert_sales_email_match_review(
            match_row=match_row,
            project=project,
            talent=talent,
            review_entry=review_entry,
            report_direction=str(report.get("direction") or "project_to_talent"),
        )
        if db_result.get("error"):
            raise HTTPException(status_code=500, detail=storage_failure_detail("Failed to store sales email review"))

        review_log_path = Path(SALES_EMAIL_REVIEW_LOG_FILE)
        review_md_path = Path(SALES_EMAIL_REVIEW_MARKDOWN_FILE)
        current_log = load_sales_email_review_report(review_log_path)
        review_report = upsert_sales_email_review_entry(current_log, review_entry, replace=False)
        write_sales_email_review_json(review_report, review_log_path)
        write_sales_email_review_markdown(review_report, review_md_path)
        audit_event = write_audit_event(
            "sales_email_match_review",
            {
                "wbs_task": "T817_6",
                "review_id": review_entry["review_id"],
                "match_key": review_entry["match_key"],
                "feedback_status": review_entry["feedback_status"],
                "project_key": review_entry["project_key"],
                "talent_key": review_entry["talent_key"],
                "db_match_result_id": db_result["match_result_id"],
                "feedback_id": db_result["feedback_id"],
            },
        )
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        print(f"[-] sales email review endpoint failed: {exc}")
        raise HTTPException(status_code=500, detail="sales email review failed") from exc

    return {
        "status": "success",
        "review": review_entry,
        "db": db_result,
        "audit_event_id": audit_event["event_id"],
        "review_log": os.path.relpath(str(Path(SALES_EMAIL_REVIEW_LOG_FILE)), PROJECT_ROOT),
    }


@app.get("/api/sales-email/reviews/summary")
async def get_sales_email_match_review_summary(limit: int = 20, username: str = Depends(verify_credentials)):
    """Authenticated summary of T817_6 sales email match reviews."""
    file_summary = {}
    if load_sales_email_review_report is not None and build_sales_email_review_report is not None:
        try:
            file_report = load_sales_email_review_report(Path(SALES_EMAIL_REVIEW_LOG_FILE))
            file_summary = {
                "file_review_count": file_report.get("review_count", 0),
                "file_status_counts": file_report.get("status_counts", {}),
            }
        except Exception:
            file_summary = {"file_review_count": 0, "file_status_counts": {}}
    return {
        "status": "success",
        "review_log": os.path.relpath(str(Path(SALES_EMAIL_REVIEW_LOG_FILE)), PROJECT_ROOT),
        **db_get_sales_email_review_summary(limit=limit),
        **file_summary,
    }


@app.post("/api/analytics/event")
async def submit_usage_analytics_event(req: UsageAnalyticsEventRequest, request: Request):
    """Store a privacy-preserving product usage event for T800 KPI aggregation."""
    event_name = clean_feedback_text(req.event_name, 80).lower()
    event_surface = clean_feedback_text(req.event_surface, 80).lower() or "public_demo"
    if event_name not in VALID_USAGE_ANALYTICS_EVENTS:
        raise HTTPException(status_code=400, detail="unsupported analytics event_name")
    if event_surface not in VALID_USAGE_ANALYTICS_SURFACES:
        raise HTTPException(status_code=400, detail="unsupported analytics event_surface")

    db_result = db_insert_usage_analytics_event(
        event_name=event_name,
        event_surface=event_surface,
        page_url=req.page_url,
        session_id=req.session_id,
        user_agent=request.headers.get("user-agent", ""),
        metadata={
            **(req.metadata or {}),
            "api_version": "2026-06-27",
            "wbs_task": "T800",
        },
    )
    if not db_result.get("id"):
        raise HTTPException(status_code=500, detail=storage_failure_detail("Failed to store analytics event"))
    return {
        "status": "success",
        "event_id": db_result["id"],
        "event_name": db_result["event_name"],
        "privacy": {
            "session_pseudonymized": True,
            "ip_address_stored": False,
            "raw_user_agent_stored": False,
            "form_contents_stored": False,
        },
    }


@app.post("/api/feedback")
async def submit_feedback(req: FeedbackRequest):
    """Store post-diagnosis helpfulness and NPS feedback for quality review."""
    rating = clean_feedback_text(req.rating, 32).lower()
    if rating not in VALID_FEEDBACK_RATINGS:
        raise HTTPException(status_code=400, detail="rating must be helpful or not_helpful")
    if req.nps_score is not None and not 0 <= req.nps_score <= 10:
        raise HTTPException(status_code=400, detail="nps_score must be between 0 and 10")
    if req.comment and len(req.comment) > MAX_FEEDBACK_COMMENT_LENGTH:
        raise HTTPException(status_code=400, detail=f"comment must be {MAX_FEEDBACK_COMMENT_LENGTH} characters or fewer")

    feedback_id = db_insert_feedback_event(
        match_result_id=req.match_id,
        rating=rating,
        nps_score=req.nps_score,
        comment=req.comment,
        source=req.source,
        page_url=req.page_url,
        session_id=req.session_id,
        metadata={"api_version": "2026-06-16", "wbs_task": "T763"},
    )
    if not feedback_id:
        raise HTTPException(status_code=500, detail=storage_failure_detail("Failed to store feedback"))
    return {"status": "success", "feedback_id": feedback_id}


@app.get("/api/feedback/summary")
async def get_feedback_summary(limit: int = 20, username: str = Depends(verify_credentials)):
    """Authenticated feedback summary for operations and quality review."""
    return {"status": "success", **db_get_feedback_summary(limit=limit)}


def _aptitude_gemini_caller(prompt: str) -> Any:
    """Bridge the pure aptitude_demo module to the app's Gemini client."""
    if not GEMINI_READY or GEMINI_CLIENT is None:
        raise RuntimeError("Gemini live mode is not configured; using vetted fallback questions.")
    response = generate_gemini_content(prompt, response_mime_type="application/json")
    return getattr(response, "text", None)


@app.post("/api/aptitude-demo/questions")
async def aptitude_demo_questions(req: AptitudeDemoQuestionRequest):
    """T876: return a session-only motivation/condition self-check question set.

    Requires legal consent. NEVER persists anything (要配慮個人情報 protection,
    R119/QA-105); the aptitude_demo module has no storage access by construction.
    Only a non-identifying count is audited — never question or answer content.
    """
    validate_legal_consent(req.legal_consent_accepted, req.legal_consent_version, "api_aptitude_demo_questions")
    caller = _aptitude_gemini_caller if GEMINI_READY else None
    result = aptitude_demo.generate_questions(count=req.count, gemini_caller=caller)
    write_audit_event(
        "aptitude_demo_questions_generated",
        {"wbs_task": "T876", "question_count": result["count"], "source": result["source"],
         "persisted": False, "answers_stored": False},
    )
    return {"status": "success", **result}


@app.get("/api/aptitude-demo/legend")
async def aptitude_demo_legend():
    """T909: score bands (正常/注意/面談目安) and the answer scale.

    Served separately from evaluation so the UI can show the criteria BEFORE a
    person answers, and so the legend always comes from the same table the
    evaluator uses. Contains no personal data — no consent gate needed.
    """
    return {"status": "success", **aptitude_demo.score_legend()}


@app.post("/api/aptitude-demo/evaluate")
async def aptitude_demo_evaluate(req: AptitudeDemoEvaluateRequest):
    """T876: evaluate 1-5 answers on-screen only. Requires consent. NEVER persists.

    Answer values and the derived condition score are 要配慮個人情報 and are held
    only in this request's memory. The audit log records that an evaluation ran
    and how many items were answered — never the answers or the resulting score.
    """
    validate_legal_consent(req.legal_consent_accepted, req.legal_consent_version, "api_aptitude_demo_evaluate")
    if not req.consented:
        raise HTTPException(status_code=400, detail="consent is required before running the self-check evaluation")
    try:
        result = aptitude_demo.evaluate_responses([a.model_dump() for a in req.answers])
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    write_audit_event(
        "aptitude_demo_evaluated",
        {"wbs_task": "T876", "answered_count": result["answered_count"],
         "persisted": False, "answers_stored": False, "score_stored": False},
    )
    return {"status": "success", **result}


@app.post("/api/employee-assessment/responses")
async def submit_employee_assessment_response(req: EmployeeAssessmentResponseRequest):
    """Store a consented, pseudonymized employee self-report response for T840."""
    if not req.consented:
        raise HTTPException(status_code=400, detail="consent is required before storing the response")

    employee_identifier = clean_feedback_text(req.employee_identifier, MAX_EMPLOYEE_IDENTIFIER_LENGTH)
    if len(employee_identifier) < 3:
        raise HTTPException(status_code=400, detail="employee_identifier must be at least 3 characters")

    department = clean_feedback_text(req.department, MAX_EMPLOYEE_DEPARTMENT_LENGTH)
    if len(department) < 2:
        raise HTTPException(status_code=400, detail="department is required")

    if not 1 <= req.motivation_level <= 5:
        raise HTTPException(status_code=400, detail="motivation_level must be between 1 and 5")
    if not 1 <= req.culture_level <= 5:
        raise HTTPException(status_code=400, detail="culture_level must be between 1 and 5")

    feedback = clean_feedback_text(req.growth_feedback, MAX_EMPLOYEE_ASSESSMENT_FEEDBACK_LENGTH)
    if len(feedback) < 10:
        raise HTTPException(status_code=400, detail="growth_feedback must be at least 10 characters")
    if req.growth_feedback and len(req.growth_feedback) > MAX_EMPLOYEE_ASSESSMENT_FEEDBACK_LENGTH:
        raise HTTPException(
            status_code=400,
            detail=f"growth_feedback must be {MAX_EMPLOYEE_ASSESSMENT_FEEDBACK_LENGTH} characters or fewer",
        )

    db_result = db_insert_employee_assessment_response(
        employee_identifier=employee_identifier,
        department=department,
        motivation_level=req.motivation_level,
        culture_level=req.culture_level,
        growth_feedback=feedback,
        consent_version=req.consent_version,
        source=req.source,
        page_url=req.page_url,
        session_id=req.session_id,
    )
    if not db_result.get("id"):
        raise HTTPException(status_code=500, detail=storage_failure_detail("Failed to store employee assessment response"))

    audit_event = write_audit_event(
        "employee_assessment_response",
        {
            "wbs_task": "T840",
            "response_id": db_result["id"],
            "subject_pseudonym": db_result["subject_pseudonym"],
            "department_bucket": department,
            "motivation_level": req.motivation_level,
            "culture_level": req.culture_level,
            "raw_identifier_stored": False,
            "sensitive_text_redacted": True,
        },
    )

    return {
        "status": "success",
        "response_id": db_result["id"],
        "subject_pseudonym": db_result["subject_pseudonym"],
        "deletion_due_at": db_result["deletion_due_at"],
        "audit_event_id": audit_event["event_id"],
        "privacy_controls": {
            "raw_identifier_stored": False,
            "sensitive_text_redacted": True,
            "consent_version": clean_feedback_text(req.consent_version, 80) or EMPLOYEE_ASSESSMENT_CONSENT_VERSION,
        },
    }


@app.get("/api/employee-assessment/responses/summary")
async def get_employee_assessment_response_summary(limit: int = 20, username: str = Depends(verify_credentials)):
    """Authenticated, redacted summary of T840 employee self-report responses."""
    return {"status": "success", **db_get_employee_assessment_summary(limit=limit)}


@app.get("/api/onboarding/state")
async def get_onboarding_state():
    """Server-canonical onboarding wizard definition (T752, gate PUBLIC-06).

    The wizard renders from this catalogue and the activation validator checks
    the same list, so the UI cannot drift from the gate.
    """
    return {
        "status": "success",
        "flow_version": ONBOARDING_FLOW_VERSION,
        "legal_consent_version": LEGAL_CONSENT_VERSION,
        "steps": [dict(step) for step in ONBOARDING_STEPS],
        "required_step_ids": list(ONBOARDING_REQUIRED_STEP_IDS),
    }


@app.post("/api/onboarding/progress")
async def compute_onboarding_progress(req: OnboardingProgressRequest):
    """Progress for the submitted steps, evaluated against the canonical catalogue."""
    return build_onboarding_progress(req.completed_step_ids)


@app.post("/api/onboarding/activate")
async def activate_onboarding(req: OnboardingActivateRequest):
    """Activate an administratively issued account after the required steps.

    Refuses activation when a required step is outstanding or when the legal
    consent is absent/stale, so activation cannot outrun the consent gate that
    T745 enforces on the analysis APIs.
    """
    account_identifier = clean_feedback_text(
        req.account_identifier, MAX_ONBOARDING_IDENTIFIER_LENGTH
    )
    if len(account_identifier) < 3:
        raise HTTPException(
            status_code=400,
            detail="account_identifier must be at least 3 characters",
        )

    if not req.legal_consent_accepted:
        raise HTTPException(
            status_code=400,
            detail="legal consent is required before activation",
        )
    if clean_feedback_text(req.legal_consent_version, 80) != LEGAL_CONSENT_VERSION:
        raise HTTPException(
            status_code=400,
            detail=(
                "Invalid legal consent version: expected "
                f"{LEGAL_CONSENT_VERSION}"
            ),
        )

    progress = build_onboarding_progress(req.completed_step_ids)
    if progress["remaining_required_step_ids"]:
        raise HTTPException(
            status_code=400,
            detail=(
                "required onboarding steps are incomplete: "
                + ", ".join(progress["remaining_required_step_ids"])
            ),
        )

    subject_pseudonym = onboarding_pseudonym(account_identifier)
    session_token = f"sess_onb_{secrets.token_hex(16)}"
    activated_at = datetime.datetime.now(datetime.timezone.utc).isoformat()
    audit_event = write_audit_event(
        "onboarding_activated",
        {
            "wbs_task": "T752",
            "flow_version": ONBOARDING_FLOW_VERSION,
            "subject_pseudonym": subject_pseudonym,
            "completed_step_ids": progress["completed_step_ids"],
            "legal_consent_version": LEGAL_CONSENT_VERSION,
            "raw_identifier_stored": False,
        },
    )
    return {
        "status": "success",
        "activated": True,
        "auth_status": "authenticated",
        "session_token": session_token,
        "flow_version": ONBOARDING_FLOW_VERSION,
        "subject_pseudonym": subject_pseudonym,
        "completed_step_ids": progress["completed_step_ids"],
        "legal_consent_version": LEGAL_CONSENT_VERSION,
        "activated_at": activated_at,
        "audit_event_id": audit_event["event_id"],
        "privacy_controls": {"raw_identifier_stored": False},
    }


@app.post("/api/attendance/punch")
async def submit_attendance_punch(req: AttendancePunchRequest):
    """Store a pseudonymized attendance punch event for the T841 internal timecard."""
    if not req.consented:
        raise HTTPException(status_code=400, detail="consent is required before storing attendance data")

    employee_identifier = clean_feedback_text(req.employee_identifier, MAX_ATTENDANCE_IDENTIFIER_LENGTH)
    if len(employee_identifier) < 3:
        raise HTTPException(status_code=400, detail="employee_identifier must be at least 3 characters")

    event_type = normalize_attendance_event_type(req.event_type)
    if not event_type:
        raise HTTPException(status_code=400, detail="event_type must be one of in, out, rest-start, or rest-end")

    db_result = db_insert_attendance_punch(
        employee_identifier=employee_identifier,
        event_type=event_type,
        source=req.source,
        page_url=req.page_url,
        session_id=req.session_id,
    )
    if not db_result.get("id"):
        raise HTTPException(status_code=500, detail=storage_failure_detail("Failed to store attendance punch"))

    audit_event = write_audit_event(
        "attendance_punch",
        {
            "wbs_task": "T841",
            "punch_id": db_result["id"],
            "subject_pseudonym": db_result["subject_pseudonym"],
            "event_type": db_result["event_type"],
            "raw_identifier_stored": False,
        },
    )
    return {
        "status": "success",
        "punch_id": db_result["id"],
        "subject_pseudonym": db_result["subject_pseudonym"],
        "event_type": db_result["event_type"],
        "recorded_at": db_result["recorded_at"],
        "audit_event_id": audit_event["event_id"],
        "privacy_controls": {"raw_identifier_stored": False},
    }


@app.post("/api/attendance/timesheet/parse")
async def parse_attendance_timesheet(
    file: UploadFile = File(...),
    employee_identifier: str = Form(...),
    consented: bool = Form(...),
    consent_version: str = Form(ATTENDANCE_CONSENT_VERSION),
    source: str = Form("attendance_timesheet_upload"),
    page_url: str = Form(""),
    session_id: str = Form(""),
):
    """Parse and store a pending approval attendance timesheet import without saving the raw file."""
    if not consented:
        raise HTTPException(status_code=400, detail="consent is required before parsing attendance data")

    clean_identifier = clean_feedback_text(employee_identifier, MAX_ATTENDANCE_IDENTIFIER_LENGTH)
    if len(clean_identifier) < 3:
        raise HTTPException(status_code=400, detail="employee_identifier must be at least 3 characters")

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="timesheet file is empty")
    if len(raw_bytes) > MAX_ATTENDANCE_FILE_BYTES:
        raise HTTPException(status_code=400, detail=f"timesheet file must be {MAX_ATTENDANCE_FILE_BYTES} bytes or fewer")

    extension = Path(file.filename or "").suffix.lower()
    if extension not in {".csv", ".txt", ".xlsx", ".xls"}:
        raise HTTPException(
            status_code=400,
            detail="only CSV, text, or Excel (.xlsx/.xls) timesheets are supported (T841/T873/T874)",
        )

    try:
        if extension == ".xlsx":
            parse_result = parse_attendance_xlsx_bytes(raw_bytes)
        elif extension == ".xls":
            parse_result = parse_attendance_xls_bytes(raw_bytes)
        else:
            parse_result = parse_attendance_csv_bytes(raw_bytes)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    db_result = db_insert_attendance_timesheet_import(
        employee_identifier=clean_identifier,
        file_name=file.filename or "timesheet.csv",
        file_bytes=raw_bytes,
        parse_result=parse_result,
        consent_version=consent_version,
        source=source,
        page_url=page_url,
        session_id=session_id,
    )
    if not db_result.get("id"):
        raise HTTPException(status_code=500, detail=storage_failure_detail("Failed to store attendance timesheet import"))

    audit_event = write_audit_event(
        "attendance_timesheet_parse",
        {
            "wbs_task": "T841",
            "import_id": db_result["id"],
            "subject_pseudonym": db_result["subject_pseudonym"],
            "summary": db_result["summary"],
            "raw_file_stored": False,
            "original_filename_stored": False,
        },
    )
    return {
        "status": "success",
        "import_id": db_result["id"],
        "subject_pseudonym": db_result["subject_pseudonym"],
        "approval_status": db_result["status"],
        "summary": db_result["summary"],
        "audit_event_id": audit_event["event_id"],
        "privacy_controls": {
            "raw_identifier_stored": False,
            "raw_file_stored": False,
            "original_filename_stored": False,
        },
    }


@app.post("/api/attendance/timesheet/approve")
async def approve_attendance_timesheet(req: AttendanceTimesheetApprovalRequest, username: str = Depends(verify_credentials)):
    """Approve or reject a parsed attendance import through the admin-gated workflow."""
    decision = clean_feedback_text(req.decision, 32).lower()
    if decision not in VALID_ATTENDANCE_DECISIONS:
        raise HTTPException(status_code=400, detail="decision must be approved or rejected")
    if req.import_id <= 0:
        raise HTTPException(status_code=400, detail="import_id must be positive")

    db_result = db_update_attendance_timesheet_decision(req.import_id, decision)
    if db_result.get("error") == "not_found":
        raise HTTPException(status_code=404, detail="attendance import not found")
    if not db_result.get("id"):
        raise HTTPException(status_code=500, detail="Failed to update attendance import")

    audit_event = write_audit_event(
        "attendance_timesheet_approval",
        {
            "wbs_task": "T841",
            "import_id": db_result["id"],
            "subject_pseudonym": db_result["subject_pseudonym"],
            "decision": db_result["status"],
            "reviewer_id": username,
        },
    )
    return {"status": "success", "attendance_import": db_result, "audit_event_id": audit_event["event_id"]}


@app.get("/api/attendance/summary")
async def get_attendance_summary(limit: int = 20, username: str = Depends(verify_credentials)):
    """Authenticated attendance punch/import summary for T841 operations review."""
    return {"status": "success", **db_get_attendance_summary(limit=limit)}


@app.post("/api/support/request")
async def submit_support_request(req: SupportRequest):
    """Store user support inquiries for triage and escalation tracking."""
    category = clean_feedback_text(req.category, 32).lower()
    if category not in VALID_SUPPORT_CATEGORIES:
        raise HTTPException(status_code=400, detail="category must be general, technical, billing, privacy, or feedback")

    priority = clean_feedback_text(req.priority, 16).lower()
    if not priority:
        priority = "high" if category in {"technical", "privacy"} else "normal"
    if priority not in VALID_SUPPORT_PRIORITIES:
        raise HTTPException(status_code=400, detail="priority must be normal, high, or urgent")

    contact_email = clean_feedback_text(req.contact_email, 254).lower()
    if not SUPPORT_EMAIL_RE.match(contact_email):
        raise HTTPException(status_code=400, detail="valid contact_email is required")

    subject = clean_feedback_text(req.subject, MAX_SUPPORT_SUBJECT_LENGTH)
    if len(subject) < 3:
        raise HTTPException(status_code=400, detail="subject must be at least 3 characters")
    if req.subject and len(req.subject) > MAX_SUPPORT_SUBJECT_LENGTH:
        raise HTTPException(status_code=400, detail=f"subject must be {MAX_SUPPORT_SUBJECT_LENGTH} characters or fewer")

    message = clean_feedback_text(req.message, MAX_SUPPORT_MESSAGE_LENGTH)
    if len(message) < 10:
        raise HTTPException(status_code=400, detail="message must be at least 10 characters")
    if req.message and len(req.message) > MAX_SUPPORT_MESSAGE_LENGTH:
        raise HTTPException(status_code=400, detail=f"message must be {MAX_SUPPORT_MESSAGE_LENGTH} characters or fewer")

    support_id = db_insert_support_request(
        category=category,
        priority=priority,
        contact_email=contact_email,
        subject=subject,
        message=message,
        source=req.source,
        page_url=req.page_url,
        session_id=req.session_id,
        metadata={"api_version": "2026-06-16", "wbs_task": "T790"},
    )
    if not support_id:
        raise HTTPException(status_code=500, detail=storage_failure_detail("Failed to store support request"))
    return {"status": "success", "support_request_id": support_id, "priority": priority}


@app.get("/api/support/summary")
async def get_support_summary(limit: int = 20, username: str = Depends(verify_credentials)):
    """Authenticated support queue summary for operations review."""
    return {"status": "success", **db_get_support_summary(limit=limit)}


class SyncRequest(BaseModel):
    candidate_name: str
    job_name: str
    final_score: int
    skill_score: int
    culture_score: int
    growth_score: int
    performing_score: int
    summary: str

# 3. API: Google Sheets Synchronizer
@app.post("/api/sync")
async def sync_to_sheets(req: SyncRequest):
    """Appends matching evaluation records into Google Sheets with visual formatting."""
    print(f"[*] API Sheets Sync request: {req.candidate_name} <=> {req.job_name}")
    
    if not SHEETS_LIB_AVAILABLE:
        print("[-] gspread library is missing. Cannot perform Sheets Sync.")
        return {"status": "error", "message": "Google Sheets client library is not installed."}
        
    client = None
    auth_mode = None
    
    # 1. Google Workspace Service Account directly from Environment Variable (Best Practice for Serverless)
    service_account_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if service_account_json:
        try:
            scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
            info = json.loads(service_account_json)
            creds = ServiceCredentials.from_service_account_info(info, scopes=scopes)
            client = gspread.authorize(creds)
            auth_mode = "Service Account (Env)"
            print("[+] Authenticated Google Sheets via GOOGLE_SERVICE_ACCOUNT_JSON environment variable.")
        except Exception as e:
            print(f"[-] GOOGLE_SERVICE_ACCOUNT_JSON environment auth failed: {e}")

    # 2. Google Workspace OAuth 2.0 from Environment Variables
    client_secret_env = os.environ.get("GOOGLE_CLIENT_SECRET_JSON")
    authorized_user_env = os.environ.get("GOOGLE_AUTHORIZED_USER_JSON")
    if not client and client_secret_env and authorized_user_env:
        try:
            temp_secret_path = os.path.join(PROJECT_ROOT, "temp_client_secret.json")
            temp_user_path = os.path.join(PROJECT_ROOT, "temp_authorized_user.json")
            
            with open(temp_secret_path, "w", encoding="utf-8") as f:
                f.write(client_secret_env)
            with open(temp_user_path, "w", encoding="utf-8") as f:
                f.write(authorized_user_env)
                
            try:
                client = gspread.oauth(
                    credentials_filename=temp_secret_path,
                    authorized_user_filename=temp_user_path
                )
                assert_expected_google_account(credentials_from_gspread_client(client), USER_EMAIL)
                auth_mode = "OAuth 2.0 (Env)"
                print("[+] Authenticated Google Sheets via OAuth environment variables.")
            finally:
                if os.path.exists(temp_secret_path):
                    os.remove(temp_secret_path)
                if os.path.exists(temp_user_path):
                    os.remove(temp_user_path)
        except GoogleWorkspaceAccountError as e:
            print(f"[-] OAuth Env account verification failed: {e}")
            return {"status": "error", "message": str(e)}
        except Exception as e:
            print(f"[-] OAuth Env authentication failed: {e}")

    # 3. OAuth 2.0 File-based Fallback
    if not client and os.path.exists(CLIENT_SECRET_FILE):
        try:
            client = gspread.oauth(
                credentials_filename=CLIENT_SECRET_FILE,
                authorized_user_filename=AUTHORIZED_USER_FILE
            )
            assert_expected_google_account(credentials_from_gspread_client(client), USER_EMAIL)
            auth_mode = "OAuth 2.0"
        except GoogleWorkspaceAccountError as e:
            print(f"[-] OAuth 2.0 Workspace account verification failed in API: {e}")
            return {"status": "error", "message": str(e)}
        except Exception as e:
            print(f"[-] OAuth 2.0 authentication failed in API: {e}")

    # 4. Service Account File-based Fallback
    if not client and os.path.exists(CREDENTIALS_FILE):
        try:
            scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
            creds = ServiceCredentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
            client = gspread.authorize(creds)
            auth_mode = "Service Account"
        except Exception as e:
            print(f"[-] Service Account authentication failed in API: {e}")
            
    if not client:
        print("[-] Authentication credentials not found for Google Sheets.")
        return {"status": "error", "message": "Credentials file not found."}
        
    try:
        # 3. Open Spreadsheet and sheet
        sh = client.open_by_key(SPREADSHEET_ID)
        
        # Check if tab exists, otherwise create it
        tab_name = "Mighty Match Logs"
        try:
            worksheet = sh.worksheet(tab_name)
            print(f"[+] Opened existing sheet: '{tab_name}'")
        except gspread.exceptions.WorksheetNotFound:
            worksheet = sh.add_worksheet(title=tab_name, rows="100", cols="10")
            print(f"[+] Created new sheet: '{tab_name}'")
            
        # Get current data to check headers
        existing_values = worksheet.get_all_values()
        
        headers = ["診断日時", "候補者氏名", "案件・求人名", "総合マッチ度 (%)", "技術 (Skill)", "文化 (Culture)", "キャリア (Growth)", "即戦力 (Performing)", "分析レポート概要"]
        
        # 4. Prepare batch requests list to prevent 429
        requests_list = []
        row_index = 1
        
        if not existing_values:
            # Append headers
            worksheet.append_row(headers)
            existing_values = [headers]
            row_index = 2
        else:
            row_index = len(existing_values) + 1
            
        # Append data row
        jst = datetime.timezone(datetime.timedelta(hours=9))
        jst_now = datetime.datetime.now(jst).strftime("%Y-%m-%d %H:%M:%S")
        data_row = [
            jst_now,
            req.candidate_name,
            req.job_name,
            req.final_score,
            req.skill_score,
            req.culture_score,
            req.growth_score,
            req.performing_score,
            req.summary
        ]
        worksheet.append_row(data_row)
        print(f"[+] Successfully appended matching record for {req.candidate_name} into row {row_index}")
        
        # 5. Apply Visual Formatting to Sheets (Mighty Blue Design) via batch update
        sheet_id = worksheet.id
        
        # Gridlines enable
        requests_list.append({
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridlinesVisible": True
                },
                "fields": "gridlinesVisible"
            }
        })
        
        # Format Header Row (Mighty Blue, Bold, White text, Centered)
        requests_list.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": COLORS["header_bg"],
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "textFormat": {
                            "foregroundColor": COLORS["header_text"],
                            "bold": True,
                            "fontSize": 11
                        }
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
            }
        })
        
        # Format Data Row Just Inserted (Center align scores, gray background if even)
        is_even = row_index % 2 == 0
        cell_format = {
            "verticalAlignment": "MIDDLE",
            "textFormat": {"fontSize": 10}
        }
        if is_even:
            cell_format["backgroundColor"] = COLORS["row_even"]
            
        requests_list.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_index - 1,
                    "endRowIndex": row_index,
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                },
                "cell": {
                    "userEnteredFormat": cell_format
                },
                "fields": "userEnteredFormat(backgroundColor,verticalAlignment,textFormat)"
            }
        })
        
        # Center align scores specifically (columns index 3 to 7: final_score, skill, culture, growth, performing)
        requests_list.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_index - 1,
                    "endRowIndex": row_index,
                    "startColumnIndex": 3,
                    "endColumnIndex": 8
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER",
                        "textFormat": {
                            "bold": True,
                            "foregroundColor": COLORS["accent_green"]
                        }
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment,textFormat)"
            }
        })
        
        # Auto-resize columns width
        requests_list.append({
            "autoResizeDimensions": {
                "dimensions": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": len(headers)
                }
            }
        })
        
        # Set specific height for rows
        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": 0,
                    "endIndex": 1
                },
                "properties": {
                    "pixelSize": 36
                },
                "fields": "pixelSize"
            }
        })
        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": row_index - 1,
                    "endIndex": row_index
                },
                "properties": {
                    "pixelSize": 28
                },
                "fields": "pixelSize"
            }
        })
        
        # Execute Visual Formatting Batch Update
        sh.batch_update({"requests": requests_list})
        print(f"[+] Visual styles and formatting applied successfully to '{tab_name}'!")
        
        return {"status": "success", "message": f"Successfully synced matching record into Google Sheets via {auth_mode}."}
        
    except Exception as e:
        print(f"[-] Sheets synchronization failed: {e}")
        return {"status": "error", "message": str(e)}


@app.get("/api/engineers")
async def list_engineers(username: Optional[str] = Depends(verify_credentials_optional)):
    if SUPABASE_SDK_ACTIVE:
        try:
            data = sdk_get_engineers()
            engineers = []
            for r in data:
                engineers.append({
                    "id": r.get("id"),
                    "name": r.get("name"),
                    "parsed_skills": json.loads(r.get("parsed_skills")) if isinstance(r.get("parsed_skills"), str) else (r.get("parsed_skills") or {}),
                    "career_goals": json.loads(r.get("career_goals")) if isinstance(r.get("career_goals"), str) else (r.get("career_goals") or {}),
                    "created_at": r.get("created_at")
                })
            return {"status": "success", "engineers": engineers}
        except Exception as e:
            print(f"[-] Supabase SDK list_engineers failed: {e}")

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, name, parsed_skills, career_goals, created_at FROM engineers ORDER BY id DESC;")
        rows = cursor.fetchall()
        engineers = []
        for r in rows:
            if db_type == "postgres":
                engineers.append({
                    "id": r[0],
                    "name": r[1],
                    "parsed_skills": json.loads(r[2]) if r[2] else {},
                    "career_goals": json.loads(r[3]) if r[3] else {},
                    "created_at": r[4].isoformat() if hasattr(r[4], "isoformat") else str(r[4])
                })
            else:
                engineers.append({
                    "id": r["id"],
                    "name": r["name"],
                    "parsed_skills": json.loads(r["parsed_skills"]) if r["parsed_skills"] else {},
                    "career_goals": json.loads(r["career_goals"]) if r["career_goals"] else {},
                    "created_at": r["created_at"]
                })
        return {"status": "success", "engineers": engineers}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        cursor.close()
        conn.close()

@app.get("/api/jobs")
async def list_jobs(username: Optional[str] = Depends(verify_credentials_optional)):
    if SUPABASE_SDK_ACTIVE:
        try:
            data = sdk_get_jobs()
            jobs = []
            for r in data:
                jobs.append({
                    "id": r.get("id"),
                    "title": r.get("title"),
                    "company": r.get("company"),
                    "job_description": r.get("job_description") or "",
                    "parsed_requirements": json.loads(r.get("parsed_requirements")) if isinstance(r.get("parsed_requirements"), str) else (r.get("parsed_requirements") or {}),
                    "company_culture": json.loads(r.get("company_culture")) if isinstance(r.get("company_culture"), str) else (r.get("company_culture") or {}),
                    "created_at": r.get("created_at")
                })
            return {"status": "success", "jobs": jobs}
        except Exception as e:
            print(f"[-] Supabase SDK list_jobs failed: {e}")

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, title, company, job_description, parsed_requirements, company_culture, created_at FROM jobs ORDER BY id DESC;")
        rows = cursor.fetchall()
        jobs = []
        for r in rows:
            if db_type == "postgres":
                jobs.append({
                    "id": r[0],
                    "title": r[1],
                    "company": r[2],
                    "job_description": r[3] or "",
                    "parsed_requirements": json.loads(r[4]) if r[4] else {},
                    "company_culture": json.loads(r[5]) if r[5] else {},
                    "created_at": r[6].isoformat() if hasattr(r[6], "isoformat") else str(r[6])
                })
            else:
                jobs.append({
                    "id": r["id"],
                    "title": r["title"],
                    "company": r["company"],
                    "job_description": r["job_description"] or "",
                    "parsed_requirements": json.loads(r["parsed_requirements"]) if r["parsed_requirements"] else {},
                    "company_culture": json.loads(r["company_culture"]) if r["company_culture"] else {},
                    "created_at": r["created_at"]
                })
        return {"status": "success", "jobs": jobs}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        cursor.close()
        conn.close()

@app.get("/api/matches")
async def list_matches(username: str = Depends(verify_credentials)):
    if SUPABASE_SDK_ACTIVE:
        try:
            from supabase_client import get_supabase_client
            client = get_supabase_client()
            if client:
                response = client.table("match_results").select(
                    "id, engineer_id, job_id, fit_ratio, score_skill, score_culture, score_growth, score_performing, match_summary, analyzed_at, engineers(name), jobs(title, company)"
                ).order("id", desc=True).execute()
                if response and response.data:
                    matches = []
                    for r in response.data:
                        matches.append({
                            "id": r.get("id"),
                            "engineer_id": r.get("engineer_id"),
                            "engineer_name": r.get("engineers", {}).get("name") if r.get("engineers") else "Unknown",
                            "job_id": r.get("job_id"),
                            "job_title": r.get("jobs", {}).get("title") if r.get("jobs") else "Unknown",
                            "company": r.get("jobs", {}).get("company") if r.get("jobs") else "Unknown",
                            "fit_ratio": r.get("fit_ratio"),
                            "scores": {
                                "skill": r.get("score_skill"),
                                "culture": r.get("score_culture"),
                                "growth": r.get("score_growth"),
                                "performing": r.get("score_performing")
                            },
                            "match_summary": r.get("match_summary"),
                            "analyzed_at": r.get("analyzed_at")
                        })
                    return {"status": "success", "matches": matches}
        except Exception as e:
            print(f"[-] Supabase SDK list_matches failed: {e}")

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    try:
        query = """
            SELECT mr.id, mr.engineer_id, e.name as engineer_name, mr.job_id, j.title as job_title, j.company as company,
                   mr.fit_ratio, mr.score_skill, mr.score_culture, mr.score_growth, mr.score_performing, mr.match_summary, mr.analyzed_at
            FROM match_results mr
            LEFT JOIN engineers e ON mr.engineer_id = e.id
            LEFT JOIN jobs j ON mr.job_id = j.id
            ORDER BY mr.id DESC;
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        matches = []
        for r in rows:
            if db_type == "postgres":
                matches.append({
                    "id": r[0],
                    "engineer_id": r[1],
                    "engineer_name": r[2] or "Unknown",
                    "job_id": r[3],
                    "job_title": r[4] or "Unknown",
                    "company": r[5] or "Unknown",
                    "fit_ratio": r[6],
                    "scores": {
                        "skill": r[7],
                        "culture": r[8],
                        "growth": r[9],
                        "performing": r[10]
                    },
                    "match_summary": r[11],
                    "analyzed_at": r[12].isoformat() if hasattr(r[12], "isoformat") else str(r[12])
                })
            else:
                matches.append({
                    "id": r["id"],
                    "engineer_id": r["engineer_id"],
                    "engineer_name": r["engineer_name"] or "Unknown",
                    "job_id": r["job_id"],
                    "job_title": r["job_title"] or "Unknown",
                    "company": r["company"] or "Unknown",
                    "fit_ratio": r["fit_ratio"],
                    "scores": {
                        "skill": r["score_skill"],
                        "culture": r["score_culture"],
                        "growth": r["score_growth"],
                        "performing": r["score_performing"]
                    },
                    "match_summary": r["match_summary"],
                    "analyzed_at": r["analyzed_at"]
                })
        return {"status": "success", "matches": matches}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        cursor.close()
        conn.close()


if __name__ == "__main__":
    if sys.platform.startswith("win"):
        try:
            import asyncio
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        except (AttributeError, RuntimeError):
            pass
    import uvicorn
    print("[*] Starting Mighty Skill-Bridge FastAPI local server...")
    uvicorn.run(app, host="127.0.0.1", port=8000)
